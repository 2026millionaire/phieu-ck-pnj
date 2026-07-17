# -*- coding: utf-8 -*-
"""Kho CCCD/CMND/hộ chiếu độc lập, mã hóa tại chỗ.

File XLSX chỉ được đọc theo thứ tự dòng. Mỗi Vendor lấy giá trị CMND không
trống đầu tiên; dữ liệu rõ không được ghi vào log hay bảng chỉ mục.
"""

from __future__ import annotations

import hashlib
import hmac
import json
import os
import re
import secrets
import sqlite3
import time
import zipfile
from contextlib import closing
from datetime import date, datetime
from pathlib import Path

from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from cryptography.hazmat.primitives.kdf.hkdf import HKDF
from openpyxl import load_workbook

from customer_lookup import (
    CustomerLookupError,
    default_data_dir,
    load_master_key,
)


IDENTITY_DB_ENV_NAME = "CUSTOMER_IDENTITY_DB"
MAX_XLSX_UNCOMPRESSED_BYTES = 300 * 1024 * 1024
MAX_XLSX_ENTRIES = 5000


def normalize_identity_customer_code(value: object) -> str | None:
    """Chuẩn hóa Vendor dùng cho kho CCCD, gồm cả mã khách hàng dạng E."""
    raw = str(value or "").strip().upper()
    if re.fullmatch(r"1[0-9]{8}", raw):
        return raw
    if re.fullmatch(r"0[0-9]{9}", raw):
        return raw[1:]
    if re.fullmatch(r"E[A-Z0-9]{7}", raw):
        return raw
    return None


def default_identity_db_path() -> Path:
    configured = os.environ.get(IDENTITY_DB_ENV_NAME)
    if configured:
        return Path(configured).expanduser()
    return default_data_dir() / "customer_identity.db"


def _derive_identity_keys(master_key: bytes) -> tuple[bytes, bytes, bytes]:
    material = HKDF(
        algorithm=hashes.SHA256(),
        length=96,
        salt=b"pnj-customer-identity-v1",
        info=b"separate identity lookup encryption digest keys",
    ).derive(master_key)
    return material[:32], material[32:64], material[64:96]


def _source_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).casefold()


def _cell_text(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        raw = str(value)
    elif isinstance(value, float) and value.is_integer():
        raw = str(int(value))
    else:
        return str(value).strip()
    number_format = str(getattr(cell, "number_format", "") or "")
    if re.fullmatch(r"0+", number_format):
        raw = raw.zfill(len(number_format))
    return raw


def _date_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = str(value or "").strip()
    for pattern in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, pattern).date().isoformat()
        except ValueError:
            continue
    return ""


def _validate_xlsx_container(path: Path) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            entries = archive.infolist()
            if not entries or len(entries) > MAX_XLSX_ENTRIES:
                raise CustomerLookupError("File XLSX có cấu trúc không hợp lệ.")
            total = sum(max(0, item.file_size) for item in entries)
            if total > MAX_XLSX_UNCOMPRESSED_BYTES:
                raise CustomerLookupError("Dung lượng giải nén của file XLSX vượt giới hạn.")
            if any(".." in Path(item.filename).parts for item in entries):
                raise CustomerLookupError("File XLSX có đường dẫn nội bộ không hợp lệ.")
    except (zipfile.BadZipFile, OSError) as exc:
        raise CustomerLookupError("Không đọc được cấu trúc file XLSX.") from exc


def select_identity_records(path: Path | str) -> tuple[list[dict[str, str]], dict]:
    """Dùng SQLite trong RAM để chọn CMND không trống đầu tiên của mỗi Vendor."""
    source_path = Path(path)
    _validate_xlsx_container(source_path)
    try:
        workbook = load_workbook(
            source_path, read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        raise CustomerLookupError("Không thể mở file XLSX bảng kê.") from exc

    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    connection.execute(
        """
        CREATE TABLE staging (
            row_number INTEGER PRIMARY KEY,
            vendor TEXT NOT NULL,
            customer_name TEXT NOT NULL,
            identity_value TEXT NOT NULL,
            plant TEXT NOT NULL,
            source_date TEXT NOT NULL
        )
        """
    )
    source_rows = 0
    invalid_vendor_rows = 0
    source_date_max = ""
    try:
        worksheet = workbook.active
        header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
        headers = [_normalize_header(cell.value) for cell in header_cells]
        required = {
            "inv.date": "source_date",
            "vendor": "vendor",
            "tênvendor": "customer_name",
            "cmnd": "identity_value",
            "plant": "plant",
        }
        if not all(name in headers for name in required):
            raise CustomerLookupError(
                "File thiếu một trong các cột bắt buộc: Inv.Date, Vendor, Tên Vendor, CMND, Plant."
            )
        indexes = {target: headers.index(name) for name, target in required.items()}

        rows_to_insert = []
        for row_number, cells in enumerate(worksheet.iter_rows(min_row=2), 2):
            vendor_raw = _cell_text(cells[indexes["vendor"]])
            if not vendor_raw:
                continue
            canonical = normalize_identity_customer_code(vendor_raw)
            if canonical is None:
                invalid_vendor_rows += 1
                continue
            plant = _cell_text(cells[indexes["plant"]])
            source_date = _date_text(cells[indexes["source_date"]].value)
            source_date_max = max(source_date_max, source_date)
            rows_to_insert.append(
                (
                    row_number,
                    canonical,
                    _cell_text(cells[indexes["customer_name"]]),
                    _cell_text(cells[indexes["identity_value"]]),
                    plant,
                    source_date,
                )
            )
            source_rows += 1
            if len(rows_to_insert) >= 2000:
                connection.executemany(
                    "INSERT INTO staging VALUES (?, ?, ?, ?, ?, ?)", rows_to_insert
                )
                rows_to_insert.clear()
        if rows_to_insert:
            connection.executemany(
                "INSERT INTO staging VALUES (?, ?, ?, ?, ?, ?)", rows_to_insert
            )
        if source_rows == 0:
            raise CustomerLookupError("File không có dòng bảng kê hợp lệ.")

        unique_vendors = int(
            connection.execute("SELECT COUNT(DISTINCT vendor) FROM staging").fetchone()[0]
        )
        selected = connection.execute(
            """
            WITH ranked AS (
                SELECT row_number, vendor, customer_name, identity_value, plant, source_date,
                       ROW_NUMBER() OVER (
                           PARTITION BY vendor ORDER BY row_number ASC
                       ) AS position
                FROM staging
                WHERE TRIM(identity_value) <> ''
            )
            SELECT vendor, customer_name, identity_value, plant, source_date
            FROM ranked
            WHERE position = 1
            ORDER BY row_number ASC
            """
        ).fetchall()
        records = [dict(row) for row in selected]
        summary = {
            "source_rows": source_rows,
            "unique_vendors": unique_vendors,
            "with_identity": len(records),
            "missing_identity": unique_vendors - len(records),
            "invalid_vendor_rows": invalid_vendor_rows,
            "source_date_max": source_date_max,
            "source_sha256": _source_sha256(source_path),
        }
        return records, summary
    finally:
        connection.close()
        workbook.close()


class CustomerIdentityStore:
    def __init__(self, db_path: Path | str, master_key: bytes):
        self.db_path = Path(db_path)
        encryption_key, self.lookup_hmac_key, self.digest_key = _derive_identity_keys(master_key)
        self.aesgcm = AESGCM(encryption_key)

    @classmethod
    def from_environment(cls, create: bool = False) -> "CustomerIdentityStore":
        path = default_identity_db_path()
        if create:
            path.parent.mkdir(parents=True, exist_ok=True)
        elif not path.exists():
            raise CustomerLookupError("Chưa có CSDL CCCD.")
        return cls(path, load_master_key(create=create))

    def connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.db_path, timeout=30)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA journal_mode=WAL")
        connection.execute("PRAGMA synchronous=FULL")
        return connection

    def initialize(self) -> None:
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        with closing(self.connect()) as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS identity_records (
                    lookup_key BLOB PRIMARY KEY,
                    payload_nonce BLOB NOT NULL,
                    payload_ciphertext BLOB NOT NULL,
                    payload_digest BLOB NOT NULL,
                    updated_at REAL NOT NULL
                ) WITHOUT ROWID;
                CREATE TABLE IF NOT EXISTS identity_import_batches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_sha256 TEXT NOT NULL UNIQUE,
                    imported_at REAL NOT NULL,
                    mode TEXT NOT NULL,
                    source_rows INTEGER NOT NULL,
                    unique_vendors INTEGER NOT NULL,
                    with_identity INTEGER NOT NULL,
                    missing_identity INTEGER NOT NULL,
                    inserted_rows INTEGER NOT NULL,
                    updated_rows INTEGER NOT NULL,
                    unchanged_rows INTEGER NOT NULL,
                    source_date_max TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS identity_metadata (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                ) WITHOUT ROWID;
                """
            )
            connection.commit()

    def lookup_key(self, customer_code: str) -> bytes:
        return hmac.new(
            self.lookup_hmac_key, customer_code.encode("ascii"), hashlib.sha256
        ).digest()

    def _encrypt(self, payload: dict, key: bytes) -> tuple[bytes, bytes, bytes]:
        plaintext = json.dumps(
            payload, ensure_ascii=False, separators=(",", ":"), sort_keys=True
        ).encode("utf-8")
        nonce = secrets.token_bytes(12)
        ciphertext = self.aesgcm.encrypt(nonce, plaintext, b"customer-identity-v1|" + key)
        digest = hmac.new(self.digest_key, plaintext, hashlib.sha256).digest()
        return nonce, ciphertext, digest

    def _decrypt(self, row: sqlite3.Row) -> dict:
        try:
            plaintext = self.aesgcm.decrypt(
                bytes(row["payload_nonce"]),
                bytes(row["payload_ciphertext"]),
                b"customer-identity-v1|" + bytes(row["lookup_key"]),
            )
            payload = json.loads(plaintext.decode("utf-8"))
        except Exception as exc:
            raise CustomerLookupError("Không thể giải mã bản ghi CCCD.") from exc
        return payload

    def get_record(self, customer_code: str) -> dict | None:
        canonical = normalize_identity_customer_code(customer_code)
        if canonical is None or not self.db_path.exists():
            return None
        self.initialize()
        key = self.lookup_key(canonical)
        with closing(self.connect()) as connection:
            row = connection.execute(
                "SELECT lookup_key, payload_nonce, payload_ciphertext FROM identity_records WHERE lookup_key = ?",
                (key,),
            ).fetchone()
        return self._decrypt(row) if row else None

    def get_summary(self) -> dict:
        self.initialize()
        with closing(self.connect()) as connection:
            metadata = {
                row["key"]: row["value"]
                for row in connection.execute("SELECT key, value FROM identity_metadata")
            }
            count = int(connection.execute("SELECT COUNT(*) FROM identity_records").fetchone()[0])
        return {
            "record_count": count,
            "source_date_max": metadata.get("source_date_max", ""),
            "last_import_at": float(metadata.get("last_import_at") or 0),
            "mode": "initial" if count == 0 else "periodic",
        }

    def preview_file(self, path: Path | str) -> dict:
        records, source = select_identity_records(path)
        self.initialize()
        inserted = updated = unchanged = name_changes = 0
        mode = "initial" if self.get_summary()["record_count"] == 0 else "periodic"
        with closing(self.connect()) as connection:
            imported = connection.execute(
                "SELECT 1 FROM identity_import_batches WHERE source_sha256 = ?",
                (source["source_sha256"],),
            ).fetchone()
            if imported:
                raise CustomerLookupError("File này đã được cập nhật trước đó.")
            for item in records:
                key = self.lookup_key(item["vendor"])
                row = connection.execute(
                    "SELECT lookup_key, payload_nonce, payload_ciphertext FROM identity_records WHERE lookup_key = ?",
                    (key,),
                ).fetchone()
                if row is None:
                    inserted += 1
                    continue
                current = self._decrypt(row)
                old_date = str(current.get("source_date") or "")
                new_date = str(item.get("source_date") or "")
                if old_date and new_date and new_date < old_date:
                    unchanged += 1
                    continue
                name_changed = bool(mode == "periodic" and
                    item["customer_name"].strip()
                    and current.get("verified_name", "") != item["customer_name"].strip()
                )
                changed = any((
                    current.get("identity_value", "") != item["identity_value"],
                    current.get("source_name", "") != item["customer_name"].strip(),
                    current.get("source_date", "") != new_date,
                    current.get("plant", "") != item["plant"],
                    name_changed,
                ))
                if changed:
                    updated += 1
                    name_changes += int(name_changed)
                else:
                    unchanged += 1
        return {
            **source,
            "inserted_rows": inserted,
            "updated_rows": updated,
            "unchanged_rows": unchanged,
            "name_changes": name_changes,
            "mode": mode,
        }

    def create_backup(self) -> Path | None:
        if not self.db_path.exists():
            return None
        backup_dir = self.db_path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        target = backup_dir / f"customer_identity-before-import-{time.strftime('%Y%m%d-%H%M%S')}-{secrets.token_hex(4)}.db"
        with closing(sqlite3.connect(self.db_path)) as source, closing(sqlite3.connect(target)) as destination:
            source.backup(destination)
        try:
            os.chmod(target, 0o600)
        except OSError:
            pass
        return target

    def import_file(self, path: Path | str, expected_sha256: str, mode: str) -> dict:
        records, source = select_identity_records(path)
        if not hmac.compare_digest(source["source_sha256"], expected_sha256):
            raise CustomerLookupError("File xác nhận không trùng với file đã kiểm tra.")
        if mode not in ("initial", "periodic"):
            raise CustomerLookupError("Chế độ cập nhật CCCD không hợp lệ.")
        current_mode = self.get_summary()["mode"]
        if mode != current_mode:
            raise CustomerLookupError("Trạng thái CSDL đã thay đổi; vui lòng kiểm tra lại file.")
        self.create_backup()
        inserted = updated = unchanged = name_updated = 0
        connection = self.connect()
        try:
            connection.execute("BEGIN IMMEDIATE")
            locked_count = int(
                connection.execute("SELECT COUNT(*) FROM identity_records").fetchone()[0]
            )
            locked_mode = "initial" if locked_count == 0 else "periodic"
            if mode != locked_mode:
                raise CustomerLookupError(
                    "CSDL vừa được cập nhật ở phiên khác; vui lòng kiểm tra lại file."
                )
            if connection.execute(
                "SELECT 1 FROM identity_import_batches WHERE source_sha256 = ?",
                (source["source_sha256"],),
            ).fetchone():
                raise CustomerLookupError("File này đã được cập nhật trước đó.")
            for item in records:
                key = self.lookup_key(item["vendor"])
                row = connection.execute(
                    "SELECT lookup_key, payload_nonce, payload_ciphertext, payload_digest FROM identity_records WHERE lookup_key = ?",
                    (key,),
                ).fetchone()
                current = self._decrypt(row) if row else {}
                old_date = str(current.get("source_date") or "")
                new_date = str(item.get("source_date") or "")
                if old_date and new_date and new_date < old_date:
                    unchanged += 1
                    continue
                verified_name = str(current.get("verified_name") or "")
                source_name = item["customer_name"].strip()
                if mode == "periodic" and source_name and source_name != verified_name:
                    verified_name = source_name
                    name_updated += 1
                payload = {
                    "customer_code": item["vendor"],
                    "identity_value": item["identity_value"],
                    "source_name": source_name,
                    "verified_name": verified_name,
                    "source_date": new_date,
                    "plant": item["plant"],
                }
                nonce, ciphertext, digest = self._encrypt(payload, key)
                if row is None:
                    connection.execute(
                        "INSERT INTO identity_records VALUES (?, ?, ?, ?, ?)",
                        (key, nonce, ciphertext, digest, time.time()),
                    )
                    inserted += 1
                elif hmac.compare_digest(bytes(row["payload_digest"]), digest):
                    unchanged += 1
                else:
                    connection.execute(
                        "UPDATE identity_records SET payload_nonce=?, payload_ciphertext=?, payload_digest=?, updated_at=? WHERE lookup_key=?",
                        (nonce, ciphertext, digest, time.time(), key),
                    )
                    updated += 1
            imported_at = time.time()
            connection.execute(
                """
                INSERT INTO identity_import_batches (
                    source_sha256, imported_at, mode, source_rows, unique_vendors,
                    with_identity, missing_identity, inserted_rows, updated_rows,
                    unchanged_rows, source_date_max
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    source["source_sha256"], imported_at, mode, source["source_rows"],
                    source["unique_vendors"], source["with_identity"], source["missing_identity"],
                    inserted, updated, unchanged, source["source_date_max"],
                ),
            )
            previous_max_row = connection.execute(
                "SELECT value FROM identity_metadata WHERE key='source_date_max'"
            ).fetchone()
            previous_max = previous_max_row[0] if previous_max_row else ""
            connection.execute(
                "INSERT OR REPLACE INTO identity_metadata VALUES ('source_date_max', ?)",
                (max(previous_max, source["source_date_max"]),),
            )
            connection.execute(
                "INSERT OR REPLACE INTO identity_metadata VALUES ('last_import_at', ?)",
                (str(imported_at),),
            )
            connection.commit()
        except Exception:
            connection.rollback()
            raise
        finally:
            connection.close()
        return {
            **source,
            "inserted_rows": inserted,
            "updated_rows": updated,
            "unchanged_rows": unchanged,
            "name_updated": name_updated,
            "dataset": self.get_summary(),
        }

# -*- coding: utf-8 -*-
"""Excel Template TT writer for ĐNCK thanh toán chi phí khác."""

import io
import html
import zipfile


UNICODE_PROBE = "ăâêôơưđÁÀẢÃẠ"


class TemplateValidationError(RuntimeError):
    pass


def detail_rows_for_template(hd_rows):
    rows = []
    for row in hd_rows or []:
        note_parts = []
        if row.get("lookup_url"):
            note_parts.append(str(row.get("lookup_url")))
        if row.get("lookup_code"):
            note_parts.append("Mã: " + str(row.get("lookup_code")))
        rows.append({
            "description": str(row.get("description") or ""),
            "amount": int(row.get("amount") or 0),
            "invoice_no_full": str(row.get("invoice_no_full") or ""),
            "supplier_tax_id": str(row.get("supplier_tax_id") or ""),
            "note": " | ".join(note_parts),
        })
    return rows


def _validate_xlsx_structure(blob):
    with zipfile.ZipFile(io.BytesIO(blob), "r") as archive:
        names = set(archive.namelist())
        required = {
            "xl/workbook.xml",
            "xl/worksheets/sheet1.xml",
            "xl/styles.xml",
        }
        missing = required - names
        if missing:
            raise TemplateValidationError("Template TT thiếu cấu trúc bắt buộc: " + ", ".join(sorted(missing)))
        workbook = archive.read("xl/workbook.xml").decode("utf-8", errors="replace")
        if "Sheet2" in workbook and 'state="hidden"' not in workbook:
            raise TemplateValidationError("Sheet2 không còn trạng thái ẩn.")
        if any(name.startswith("xl/tables/") for name in names):
            table_xml = "\n".join(
                archive.read(name).decode("utf-8", errors="replace")
                for name in names
                if name.startswith("xl/tables/")
            )
            if "Table1" not in table_xml and "table1" not in table_xml.lower():
                raise TemplateValidationError("Không xác minh được Table1 trong template.")
        sheet_xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8", errors="replace")
        sheet_plain = html.unescape(sheet_xml)
        if "Chi phí phòng chờ khách VIP" not in sheet_plain and "sharedStrings.xml" in names:
            shared = archive.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
            if "Chi phí phòng chờ khách VIP" not in html.unescape(shared):
                raise TemplateValidationError("Không đọc lại được nội dung Unicode trong XML.")
        xml_text = sheet_xml
        if "xl/sharedStrings.xml" in names:
            xml_text += archive.read("xl/sharedStrings.xml").decode("utf-8", errors="replace")
        if "?" in xml_text and "Chi phí" not in xml_text:
            raise TemplateValidationError("Nghi ngờ lỗi Unicode trong Template TT.")


def build_template_tt_bytes(template_path, hd_rows):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise TemplateValidationError("Thiếu thư viện openpyxl để ghi Template TT.") from exc

    wb = load_workbook(template_path)
    if "Sheet1" not in wb.sheetnames:
        raise TemplateValidationError("Template TT không có Sheet1.")
    ws = wb["Sheet1"]
    if "Sheet2" in wb.sheetnames and wb["Sheet2"].sheet_state != "hidden":
        raise TemplateValidationError("Sheet2 trong template gốc không ở trạng thái ẩn.")

    rows = detail_rows_for_template(hd_rows)
    for idx in range(30):
        excel_row = 5 + idx
        ws.cell(row=excel_row, column=1, value=idx + 1)
        for column in range(2, 7):
            ws.cell(row=excel_row, column=column, value=None)
        ws.cell(row=excel_row, column=7, value=True)

    for idx, row in enumerate(rows):
        excel_row = 5 + idx
        ws.cell(row=excel_row, column=2, value=row["description"])
        ws.cell(row=excel_row, column=3, value=row["amount"])
        ws.cell(row=excel_row, column=4, value=row["invoice_no_full"])
        ws.cell(row=excel_row, column=5, value=row["supplier_tax_id"])
        ws.cell(row=excel_row, column=6, value=row["note"])

    output = io.BytesIO()
    wb.save(output)
    blob = output.getvalue()

    check_wb = load_workbook(io.BytesIO(blob), data_only=False)
    check_ws = check_wb["Sheet1"]
    if rows:
        if check_ws.cell(row=5, column=2).value != rows[0]["description"]:
            raise TemplateValidationError("Read-back Template TT không khớp diễn giải dòng đầu.")
        if check_ws.cell(row=5, column=4).value != rows[0]["invoice_no_full"]:
            raise TemplateValidationError("Read-back Template TT không khớp số hóa đơn dòng đầu.")
    _validate_xlsx_structure(blob)
    return blob

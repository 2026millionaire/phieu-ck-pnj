# -*- coding: utf-8 -*-
"""
Phieu Xac Nhan Chuyen Khoan - PNJ Store 1305
Flask web app for creating transfer confirmation slips.
"""

import json
import io
import os
import re
import secrets
import shutil
import sqlite3
import subprocess
import tempfile
import threading
import time
import unicodedata
import urllib.parse
import urllib.request
import webbrowser
import zipfile
from datetime import datetime, timedelta
from pathlib import Path

from functools import wraps

from flask import Flask, g, jsonify, redirect, render_template, request, send_file, session, url_for
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from werkzeug.security import check_password_hash, generate_password_hash

import shared_auth
import erp_billing
import erp_business_partner
import erp_purchase_orders

try:
    from customer_lookup import (
        CustomerLookupError,
        CustomerLookupStore,
        default_db_path as customer_lookup_db_path,
        normalize_customer_code,
        suggestion_for_field,
    )
except ImportError:
    CustomerLookupStore = None
    CustomerLookupError = RuntimeError
    customer_lookup_db_path = None
    normalize_customer_code = None
    suggestion_for_field = None

try:
    from customer_identity import (
        CustomerIdentityStore,
        default_identity_db_path,
        select_identity_records,
    )
except ImportError:
    CustomerIdentityStore = None
    default_identity_db_path = None
    select_identity_records = None

try:
    from employee_lookup import EmployeeLookupStore, default_employee_db_path, normalize_employee_code
except ImportError:
    EmployeeLookupStore = None
    default_employee_db_path = None
    normalize_employee_code = None

# ---------------------------------------------------------------------------
# App config
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.secret_key = os.environ.get("SECRET_KEY", "pnj1305-xnck-secret-key-2026")

# Auth: only require login when running on server (not localhost)
REQUIRE_LOGIN = os.environ.get("REQUIRE_LOGIN", "0") == "1"

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "phieu_ck.db")
PORT = 5050
CUSTOMER_LOOKUP_TURNSTILE_SITEKEY = os.environ.get(
    "CUSTOMER_LOOKUP_TURNSTILE_SITEKEY", ""
).strip()
CUSTOMER_LOOKUP_TURNSTILE_SECRET = os.environ.get(
    "CUSTOMER_LOOKUP_TURNSTILE_SECRET", ""
).strip()
CUSTOMER_LOOKUP_TURNSTILE_HOSTNAME = os.environ.get(
    "CUSTOMER_LOOKUP_TURNSTILE_HOSTNAME", ""
).strip()
_customer_lookup_store = None
_customer_identity_store = None
_employee_lookup_store = None
CUSTOMER_IMPORT_MAX_BYTES = 500 * 1024 * 1024
CUSTOMER_IMPORT_MAX_FILES = 10
# Bảng kê miền Trung có thể lớn hơn nhiều lần bộ Huế. Giới hạn vẫn hữu hạn để
# tránh một tài khoản Admin bị lạm dụng làm đầy ổ đĩa bằng file nén độc hại.
CUSTOMER_IDENTITY_IMPORT_MAX_BYTES = 150 * 1024 * 1024
_customer_import_jobs_lock = threading.Lock()
PDF_TOKEN_MAX_AGE = 15 * 60
DEFAULT_QT82_FORM_URL = (
    "https://eoffice.pnj.com.vn/workflow/SitePages/NewWorkflow.aspx"
    "?mode=1&LID=4ABC02AE-DF6F-4EE7-95CC-E431993C78BB&wid=1346"
)
QT82_EXTENSION_DIR = Path(__file__).resolve().parent / "chrome-extension" / "qt82-helper"
QT82_EXTENSION_FILES = (
    "manifest.json",
    "background.js",
    "web-bridge.js",
    "eoffice-kendo-main.js",
    "eoffice-fill.js",
    "README.md",
)

CENTRAL_PLANTS_SEED = (
    ("Huế", "1303"), ("Huế", "1304"), ("Huế", "1305"), ("Huế", "1394"), ("Huế", "1398"),
    ("Huế", "1465"), ("Huế", "1570"), ("Huế", "1613"),
    ("Quảng Trị", "1339"), ("Quảng Trị", "1530"), ("Quảng Trị", "1671"),
    ("Quảng Bình", "1397"), ("Quảng Bình", "1331"), ("Quảng Bình", "1540"),
    ("Quảng Bình", "1490"), ("Quảng Bình", "1619"), ("Quảng Bình", "1657"),
    ("Đà Nẵng", "1267"), ("Đà Nẵng", "1268"), ("Đà Nẵng", "1269"),
    ("Đà Nẵng", "1270"), ("Đà Nẵng", "1271"), ("Đà Nẵng", "1272"),
    ("Đà Nẵng", "1273"), ("Đà Nẵng", "1274"), ("Đà Nẵng", "1275"),
    ("Đà Nẵng", "1276"), ("Đà Nẵng", "1369"), ("Đà Nẵng", "1395"),
    ("Đà Nẵng", "1416"), ("Đà Nẵng", "1427"), ("Đà Nẵng", "1428"),
    ("Đà Nẵng", "1454"), ("Đà Nẵng", "1563"), ("Đà Nẵng", "1668"), ("Đà Nẵng", "1669"),
    ("Bình Định", "1241"), ("Bình Định", "1242"), ("Bình Định", "1243"),
    ("Bình Định", "1366"), ("Bình Định", "1464"), ("Bình Định", "1500"),
    ("Bình Định", "1531"), ("Bình Định", "1598"), ("Bình Định", "1622"),
    ("Quảng Nam", "1332"), ("Quảng Nam", "1512"), ("Quảng Nam", "1529"),
    ("Quảng Nam", "1333"), ("Quảng Nam", "1396"), ("Quảng Nam", "1588"),
    ("Quảng Ngãi", "1334"), ("Quảng Ngãi", "1335"), ("Quảng Ngãi", "1429"), ("Quảng Ngãi", "1444"), ("Quảng Ngãi", "1448"),
    ("Quảng Ngãi", "1466"), ("Quảng Ngãi", "1515"), ("Quảng Ngãi", "1612"),
)

DNCK_EMPLOYEE_DEMO_SEED = (
    ("E01F7743", "HỒ THỊ HÀ MY", "106873221304", "Vietinbank", "046184002275"),
    ("E0124764", "CHÂU ĐĂNG KHOA", "109876756206", "Vietinbank", "046093004708"),
    ("E0116169", "TÔN NỮ MINH THANH", "0772400486", "Vietinbank", "046196009728"),
    ("E0112425", "LÊ THỊ KIM TUYẾN", "105871228613", "Vietinbank", "046195009339"),
    ("E01F4212", "ĐOÀN THỊ THU HẰNG", "0398800224", "Vietinbank", "046188010576"),
    ("E0124136", "LÊ THỊ CÚC", "107878423202", "Vietinbank", "046198003167"),
    ("E0130126", "TRẦN QUANG TRINH", "106885680781", "Vietinbank", "046094015432"),
    ("E0111358", "BÙI KHẮC KIM LIÊN", "103871216615", "Vietinbank", "046194021724"),
    ("E0121620", "THẨM THỊ NGỌC DUNG", "104868399506", "Vietinbank", "046197003617"),
    ("E0126423", "NGUYỄN THỊ MỸ UYÊN", "105881616099", "Vietinbank", "046302009946"),
    ("E0116165", "LÊ THỊ THÙY NHƯ", "0364921308", "Vietinbank", "046195012263"),
    ("E0123813", "HỒ VĂN TRUNG", "107878018535", "Vietinbank", "046082007734"),
    ("E0123806", "HÀ VĂN RIN", "0359050023", "Vietinbank", "046093000857"),
    ("E01F9014", "HUỲNH THỊ THU HƯƠNG", "0935052054", "Vietinbank", ""),
    ("E0111919", "LÊ THỊ MỸ TUYỀN", "100872810158", "Vietinbank", "046190005028"),
    ("E0117010", "NGUYỄN ĐỨC NHẬT QUANG", "105873763741", "Vietinbank", "046097011168"),
    ("E0125140", "NGUYỄN THỊ MAI YẾN", "105880005837", "Vietinbank", "046500005467"),
    ("E01M6688", "TRẦN XUÂN HẢI", "100873771126", "Vietinbank", "046081013617"),
    ("E01F6890", "TRƯƠNG THANH THANH", "0932599103", "Vietinbank", "046190013521"),
)


def normalize_qt82_form_url(value):
    """Chỉ cho phép URL HTTPS thuộc workflow eOffice PNJ, không nhận credential/fragment."""
    try:
        parsed = urllib.parse.urlsplit(str(value or "").strip())
        if (
            parsed.scheme.lower() != "https"
            or parsed.hostname != "eoffice.pnj.com.vn"
            or parsed.username
            or parsed.password
            or parsed.port not in (None, 443)
            or not parsed.path.lower().startswith("/workflow/")
            or parsed.fragment
        ):
            return ""
        netloc = "eoffice.pnj.com.vn" if parsed.port is None else "eoffice.pnj.com.vn:443"
        return urllib.parse.urlunsplit(("https", netloc, parsed.path, parsed.query, ""))
    except (TypeError, ValueError):
        return ""


def build_qt82_extension_archive():
    """Đóng gói đúng mã nguồn extension đang chạy, không đưa dữ liệu runtime vào ZIP."""
    manifest_path = QT82_EXTENSION_DIR / "manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    version = str(manifest.get("version", "")).strip()
    if not re.fullmatch(r"\d+(?:\.\d+){1,3}", version):
        raise ValueError("Phiên bản extension trong manifest không hợp lệ.")

    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for relative_name in QT82_EXTENSION_FILES:
            source = (QT82_EXTENSION_DIR / relative_name).resolve()
            if source.parent != QT82_EXTENSION_DIR.resolve() or not source.is_file():
                raise FileNotFoundError(f"Thiếu tệp extension: {relative_name}")
            archive.write(source, arcname=f"PNJ-QT82-Draft-Helper/{relative_name}")
    output.seek(0)
    return output, version

# ---------------------------------------------------------------------------
# Load data from Excel
# ---------------------------------------------------------------------------
import pandas as pd

DATA_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "phieu-ck-data.xlsx")


def load_bank_data():
    """Load bank list from Excel sheet 'bank'."""
    try:
        df = pd.read_excel(DATA_XLSX, sheet_name="bank", dtype=str).fillna("")
        banks = []
        for _, row in df.iterrows():
            banks.append({
                "eoffice": row.get("Mã eoffice", "").strip(),
                "bin": row.get("BIN", "").strip(),
                "ten_tra_cuu": row.get("Tên tra cứu", "").strip(),
                "ten_day_du": row.get("Tên đầy đủ", "").strip(),
                "ten_gd": row.get("Tên GD", "").strip(),
                "code": row.get("Code", "").strip(),
            })
        return banks
    except Exception:
        return []


def load_tvv_data():
    """Load TVV list from Excel sheet 'tvv'."""
    try:
        df = pd.read_excel(DATA_XLSX, sheet_name="tvv", dtype=str).fillna("")
        tvvs = []
        for _, row in df.iterrows():
            tvvs.append({
                "ma": row.get("mã TVV", "").strip(),
                "ten": row.get("tên TVV", "").strip(),
            })
        return tvvs
    except Exception:
        return []


BANK_LIST = load_bank_data()
TVV_LIST = load_tvv_data()

# Legacy BANK_BINS for backward compat (QR generation)
BANK_BINS = {b["ten_gd"]: b["bin"] for b in BANK_LIST}
BANK_BINS.update({b["code"]: b["bin"] for b in BANK_LIST})

# MBBank bankId mapping (BIN → bankId for account lookup)
def load_mb_bank_map():
    try:
        df = pd.read_excel(DATA_XLSX, sheet_name="mb_bank", dtype=str).fillna("")
        return {row["BIN"]: row["bankId_MB"] for _, row in df.iterrows() if row["BIN"]}
    except Exception:
        return {}

MB_BANK_MAP = load_mb_bank_map()

# ---------------------------------------------------------------------------
# Staff defaults (store 1305)
# ---------------------------------------------------------------------------
STAFF = {
    "nv_ke_toan": ["CHÂU ĐĂNG KHOA", "LÊ THỊ MỸ TUYỀN"],
    "cua_hang_truong": "HỒ THỊ HÀ MY",
    "tvv": "NGUYỄN THỊ MỸ UYÊN",
}
PAYMENT_PLANNING_STORE_NAME = "TTKH 27 Hà Nội,Huế"
PAYMENT_PLANNING_TAX_CODE = "0300521758"
PAYMENT_PLANNING_PNJ_ADDRESS = "170E Phan Đăng Lưu, Phường Đức Nhuận, Thành phố Hồ Chí Minh, Việt Nam"
PAYMENT_PLANNING_PNJ_CONTACT = "+ 84 (028) 39951703"
PAYMENT_PLANNING_PROFILES = {
    "pnj": {
        "abbr": "PNJ",
        "representative_abbr": "PNJ",
        "company_name": 'CÔNG TY CỔ PHẦN VÀNG BẠC ĐÁ QUÝ PHÚ NHUẬN (Sau đây gọi tắt là "PNJ")',
        "store_name": PAYMENT_PLANNING_STORE_NAME,
        "tax_code": PAYMENT_PLANNING_TAX_CODE,
        "address": PAYMENT_PLANNING_PNJ_ADDRESS,
        "contact": PAYMENT_PLANNING_PNJ_CONTACT,
        "has_working_day_definition": True,
        "definition_rows": {
            "ngay_t": "Là ngày PNJ hoàn tất tiếp nhận tài sản, hồ sơ/chứng từ hợp lệ, hai bên ký xác nhận Bảng kê mua lại và Thoả Thuận này; Trường hợp các sự kiện này phát sinh vào các thời điểm khác nhau, Ngày T được xác định là ngày hoàn thành sự kiện cuối cùng.\nVD: BKML ký ngày 1/7, Thoả thuận thu đổi sản phẩm ký ngày 2/7, Tiếp nhận tài sản ngày 3/7. Như vậy T là ngày 3/7",
            "ngay_lam_viec": "Là ngày từ thứ Hai đến thứ Sáu, không bao gồm ngày nghỉ hằng tuần, ngày nghỉ lễ, tết và ngày PNJ/ngân hàng phục vụ thanh toán không làm việc theo quy định hoặc thông báo hợp lệ.",
            "cach_tinh": "“T+n” là ngày làm việc thứ n kể từ ngày liền sau Ngày T. Nếu ngày dự kiến thanh toán rơi vào ngày không phải Ngày làm việc, thời hạn được chuyển sang Ngày làm việc tiếp theo.",
            "hoan_tat": "Đối với chuyển khoản, nghĩa vụ thanh toán được xem là hoàn tất khi PNJ đã phát lệnh chuyển tiền hợp lệ đến đúng thông tin tài khoản nhận tiền của Khách Hàng ở phần đầu Thoả Thuận này; thời điểm tiền ghi Có phụ thuộc quy trình xử lý của ngân hàng, trừ trường hợp lỗi thuộc PNJ.",
        },
        "effectiveness_items": [
            "Thoả thuận này có hiệu lực từ ngày ký.",
            "Trường hợp có sự khác biệt giữa Thỏa thuận này và các tài liệu, chứng từ khác liên quan đến giao dịch thu đổi sản phẩm về phương thức thanh toán, thời hạn thanh toán hoặc các nội dung được điều chỉnh tại Thỏa thuận này, Thỏa thuận này được ưu tiên áp dụng.",
            "Các bên ưu tiên trao đổi, đối chiếu và thương lượng thiện chí khi phát sinh vướng mắc. Trường hợp không giải quyết được, tranh chấp được xử lý tại cơ quan có thẩm quyền theo quy định pháp luật.",
            "Phụ lục được lập thành 02 bản có giá trị như nhau, mỗi bên giữ 01 bản",
        ],
    },
    "cao": {
        "abbr": "CAF",
        "representative_abbr": "CAF",
        "company_name": 'CÔNG TY TRÁCH NHIỆM HỮU HẠN MỘT THÀNH VIÊN THỜI TRANG CAO (CAF)',
        "store_name": PAYMENT_PLANNING_STORE_NAME,
        "tax_code": "0309279212",
        "address": "170E Phan Đăng Lưu, Phường Đức Nhuận, Thành phố Hồ Chí Minh",
        "contact": PAYMENT_PLANNING_PNJ_CONTACT,
        "has_working_day_definition": False,
        "definition_rows": {
            "ngay_t": "Là ngày CAF hoàn tất tiếp nhận tài sản, hồ sơ/chứng từ hợp lệ, hai bên ký xác nhận Bảng kê mua lại và Thoả Thuận này; Trường hợp các sự kiện này phát sinh vào các thời điểm khác nhau, Ngày T được xác định là ngày hoàn thành sự kiện cuối cùng.\nVD: BKML ký ngày 1/7, Thoả thuận thu đổi sản phẩm ký ngày 2/7, Tiếp nhận tài sản ngày 3/7. Như vậy T là ngày 3/7",
            "cach_tinh": "“T+n” là ngày thứ n kể từ ngày liền sau Ngày T.",
            "hoan_tat": "Đối với chuyển khoản, nghĩa vụ thanh toán được xem là hoàn tất khi CAF đã phát lệnh chuyển tiền hợp lệ đến đúng thông tin tài khoản nhận tiền của Khách Hàng ở phần đầu Thoả Thuận này; thời điểm tiền ghi Có phụ thuộc quy trình xử lý của ngân hàng, trừ trường hợp lỗi thuộc CAF.",
        },
        "effectiveness_items": [
            "Thoả thuận này có hiệu lực từ ngày ký. Quyền sở hữu đối với Sản phẩm thu đổi được chuyển giao ngay lập tức và hoàn toàn từ Khách Hàng sang CAF kể từ thời điểm CAF nhận bàn giao Tài sản trên thực tế và hai bên ký Bảng Kê Mua lại/Thỏa Thuận Thu đổi sản phẩm",
            "Trường hợp có sự khác biệt giữa Thỏa thuận này và các tài liệu, chứng từ khác liên quan đến giao dịch thu đổi sản phẩm về phương thức thanh toán, thời hạn thanh toán hoặc các nội dung được điều chỉnh tại Thỏa thuận này, Thỏa thuận này được ưu tiên áp dụng.",
            "Các bên ưu tiên trao đổi, đối chiếu và thương lượng thiện chí khi phát sinh vướng mắc. Trường hợp không giải quyết được, tranh chấp được xử lý tại cơ quan có thẩm quyền theo quy định pháp luật.",
            "Thỏa Thuận này chấm dứt khi xảy ra một trong các trường hợp sau:\n(a) CAF đã hoàn thành đầy đủ nghĩa vụ thanh toán theo Thỏa Thuận này; hoặc\n(b) Các Bên thống nhất bằng văn bản về việc không tiếp tục thực hiện Thỏa Thuận này và khách hàng lựa chọn phương án thu đổi sản phẩm khác theo chính sách của CAF.\nPhụ lục được lập thành 02 bản có giá trị như nhau, mỗi bên giữ 01 bản",
        ],
    },
}

# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_db():
    """Get a database connection for the current request."""
    if "db" not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db


def get_customer_lookup_store():
    """Mở kho tra cứu nếu đã được nhập và có thể mở khóa trên máy hiện tại."""
    global _customer_lookup_store
    if CustomerLookupStore is None or customer_lookup_db_path is None:
        return None
    if _customer_lookup_store is not None:
        return _customer_lookup_store
    try:
        if not customer_lookup_db_path().is_file():
            return None
        store = CustomerLookupStore.from_environment(create=False)
        store.initialize()
        _customer_lookup_store = store
        return store
    except CustomerLookupError:
        return None


def get_customer_identity_store(create=False):
    """Mở kho CCCD độc lập; chỉ Admin/import mới được phép tạo kho mới."""
    global _customer_identity_store
    if CustomerIdentityStore is None or default_identity_db_path is None:
        return None
    if _customer_identity_store is not None:
        return _customer_identity_store
    try:
        if not create and not default_identity_db_path().is_file():
            return None
        store = CustomerIdentityStore.from_environment(create=create)
        store.initialize()
        _customer_identity_store = store
        return store
    except CustomerLookupError:
        return None


def get_employee_lookup_store(create=False):
    global _employee_lookup_store
    if EmployeeLookupStore is None or default_employee_db_path is None:
        return None
    if _employee_lookup_store is not None:
        return _employee_lookup_store
    try:
        if not create and not default_employee_db_path().is_file():
            return None
        store = EmployeeLookupStore.from_environment(create=create)
        store.initialize()
        _employee_lookup_store = store
        return store
    except CustomerLookupError:
        return None


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    """Create tables if they don't exist."""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS phieu (
            id              INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at      TEXT NOT NULL,
            ma_kh           TEXT,
            ten_kh          TEXT,
            sdt             TEXT,
            cccd            TEXT,
            dia_chi         TEXT DEFAULT '',
            so_tk           TEXT,
            ten_tk          TEXT,
            ngan_hang       TEXT,
            so_bk           TEXT,
            tvv_code        TEXT,
            tvv_name        TEXT,
            cht_name        TEXT,
            plant           TEXT DEFAULT '1305',
            chung_tu_json   TEXT,
            tong_ck         REAL DEFAULT 0,
            ngay_tt         TEXT,
            status          TEXT DEFAULT 'draft',
            qr_url          TEXT,
            noi_dung        TEXT,
            nguoi_ki        TEXT DEFAULT 'tvv',
            da_trinh        INTEGER DEFAULT 0,
            sap_document_override TEXT DEFAULT '',
            use_bk_ref      INTEGER DEFAULT 0,
            show_payment_dates INTEGER DEFAULT 0
        )
    """)
    # Add columns if missing (for existing DBs)
    for col, ctype, default in [
        ("nguoi_ki", "TEXT", "'tvv'"),
        ("da_trinh", "INTEGER", "0"),
        ("user_id", "INTEGER", "1"),
        ("sap_document_override", "TEXT", "''"),
        ("dia_chi", "TEXT", "''"),
        ("use_bk_ref", "INTEGER", "0"),
        ("show_payment_dates", "INTEGER", "0"),
    ]:
        try:
            conn.execute(f"ALTER TABLE phieu ADD COLUMN {col} {ctype} DEFAULT {default}")
        except Exception:
            pass

    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            name     TEXT DEFAULT ''
        )
    """)
    # Create default admin user if no users exist
    row = conn.execute("SELECT COUNT(*) FROM users").fetchone()
    if row[0] == 0:
        conn.execute(
            "INSERT OR IGNORE INTO users (username, password, name) VALUES (?, ?, ?)",
            ("admin", generate_password_hash("pnj1305"), "Admin"),
        )

    # TVV table (replaces Excel sheet)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS tvv (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            ma   TEXT NOT NULL,
            ten  TEXT NOT NULL
        )
    """)
    # Migrate TVV from Excel → DB if table is empty
    row_tvv = conn.execute("SELECT COUNT(*) FROM tvv").fetchone()
    if row_tvv[0] == 0:
        for t in TVV_LIST:
            if t["ma"] and t["ten"]:
                conn.execute("INSERT INTO tvv (ma, ten) VALUES (?, ?)", (t["ma"], t["ten"]))

    # Lý do hủy BK table
    conn.execute("""
        CREATE TABLE IF NOT EXISTS lydo_huy (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            noi_dung TEXT NOT NULL
        )
    """)
    row_ld = conn.execute("SELECT COUNT(*) FROM lydo_huy").fetchone()
    if row_ld[0] == 0:
        for ld in [
            "Khách hàng đổi ý.",
            "Sai thông tin khách hàng.",
            "Sai mã sản phẩm.",
            "Sai trọng lượng sản phẩm.",
            "Sai giá trị mua lại.",
        ]:
            conn.execute("INSERT INTO lydo_huy (noi_dung) VALUES (?)", (ld,))

    conn.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS dnck (
            id               INTEGER PRIMARY KEY AUTOINCREMENT,
            created_at       TEXT DEFAULT CURRENT_TIMESTAMP,
            object_type      TEXT DEFAULT 'customer',
            object_code      TEXT,
            object_name      TEXT,
            identity_value   TEXT,
            phone            TEXT,
            account_number   TEXT,
            account_name     TEXT,
            bank             TEXT,
            purpose          TEXT,
            approval_level   TEXT,
            expense_type     TEXT,
            cost_group       TEXT,
            request_content  TEXT,
            sap_document     TEXT,
            amount           REAL DEFAULT 0,
            payment_tag      TEXT DEFAULT 'Bảng Kê',
            approver_option  TEXT DEFAULT 'none',
            detail_json      TEXT,
            hashtags_json    TEXT,
            reference_note   TEXT,
            reference_links_json TEXT,
            cost_limit_ref   TEXT,
            da_trinh         INTEGER DEFAULT 0,
            user_id          INTEGER DEFAULT 1
        )
    """)
    try:
        conn.execute("ALTER TABLE dnck ADD COLUMN identity_value TEXT DEFAULT ''")
    except Exception:
        pass
    for statement in (
        "ALTER TABLE dnck ADD COLUMN hashtags_json TEXT DEFAULT '[]'",
        "ALTER TABLE dnck ADD COLUMN reference_note TEXT DEFAULT ''",
        "ALTER TABLE dnck ADD COLUMN reference_links_json TEXT DEFAULT '[]'",
        "ALTER TABLE dnck ADD COLUMN cost_limit_ref TEXT DEFAULT ''",
        "ALTER TABLE dnck ADD COLUMN approval_level TEXT DEFAULT ''",
        "ALTER TABLE dnck ADD COLUMN expense_type TEXT DEFAULT ''",
    ):
        try:
            conn.execute(statement)
        except Exception:
            pass
    conn.execute("""
        CREATE TABLE IF NOT EXISTS dnck_object_lookup (
            id                 INTEGER PRIMARY KEY AUTOINCREMENT,
            object_code        TEXT NOT NULL,
            object_name        TEXT NOT NULL,
            account_number     TEXT NOT NULL,
            bank               TEXT NOT NULL,
            identity_value     TEXT DEFAULT '',
            bank_eoffice_code  TEXT DEFAULT '',
            is_primary         INTEGER DEFAULT 1,
            source             TEXT DEFAULT 'demo'
        )
    """)
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_dnck_object_lookup_account
        ON dnck_object_lookup (object_code, account_number)
    """)
    for code, name, account, bank, identity in DNCK_EMPLOYEE_DEMO_SEED:
        conn.execute("""
            INSERT OR REPLACE INTO dnck_object_lookup
                (object_code, object_name, account_number, bank, identity_value,
                 bank_eoffice_code, is_primary, source)
            VALUES (?, ?, ?, ?, ?, ?, 1, 'demo')
        """, (
            remove_all_whitespace(code).upper(),
            name,
            normalize_account_number(account),
            bank,
            remove_all_whitespace(identity),
            find_eoffice_bank_code(bank),
        ))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS central_plants (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            province TEXT NOT NULL,
            plant    TEXT NOT NULL UNIQUE
        )
    """)
    conn.executemany(
        "INSERT OR IGNORE INTO central_plants (province, plant) VALUES (?, ?)",
        CENTRAL_PLANTS_SEED,
    )
    # Default settings
    defaults = {
        "cht_name": "HỒ THỊ HÀ MY",
        "cht_short_name": "Hà My",
        "qt82_store_manager_query": "my.hth",
        "qt82_form_url": DEFAULT_QT82_FORM_URL,
        "kt1_name": "CHÂU ĐĂNG KHOA",
        "kt1_short_name": "Khoa",
        "kt2_name": "LÊ THỊ MỸ TUYỀN",
        "kt2_short_name": "Tuyền",
        "thoi_gian_ck": "48",
        "show_payment_time": "1",
        "plant": "1305",
        "vietqr_client_id": "",
        "vietqr_api_key": "",
        "mb_username": "",
        "mb_password": "",
        "mb_account": "",
        "bk_prefix": "4403",
        "invoice_prefix": "901",
        "deposit_prefix": "16",
        "hbtl_prefix": "990",
        "tvv_button_color_mode": "0",
        "billing_invoice_days": "2",
        "use_bk_ref_default": "0",
        "show_payment_dates_default": "1",
    }
    for k, v in defaults.items():
        conn.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", (k, v))
    conn.commit()
    conn.close()


def get_settings():
    """Load all settings as dict."""
    db = get_db()
    rows = db.execute("SELECT key, value FROM settings").fetchall()
    return {r["key"]: r["value"] for r in rows}


def settings_flag(settings, key, default="0"):
    """Return a boolean-like integer for settings/data flags."""
    value = settings.get(key, default) if isinstance(settings, dict) else default
    return 1 if str(value).strip().lower() in {"1", "true", "yes", "on"} else 0


def row_to_dict(row):
    """Convert a sqlite3.Row to a plain dict."""
    if row is None:
        return None
    return dict(row)

# ---------------------------------------------------------------------------
# Number to Vietnamese words
# ---------------------------------------------------------------------------

def _doc_so_hang(n):
    """Read a number 0-999 in Vietnamese."""
    ones = [
        "", "mot", "hai", "ba", "bon", "nam",
        "sau", "bay", "tam", "chin",
    ]
    vn = [
        "", "m\u1ed9t", "hai", "ba", "b\u1ed1n", "n\u0103m",
        "s\u00e1u", "b\u1ea3y", "t\u00e1m", "ch\u00edn",
    ]

    if n == 0:
        return ""

    tram = n // 100
    chuc = (n % 100) // 10
    donvi = n % 10

    parts = []

    if tram > 0:
        parts.append(f"{vn[tram]} tr\u0103m")
        if chuc == 0 and donvi > 0:
            parts.append(f"linh {vn[donvi]}")
        elif chuc == 1:
            parts.append("m\u01b0\u1eddi")
            if donvi > 0:
                if donvi == 5:
                    parts.append("l\u0103m")
                elif donvi == 1:
                    parts.append("m\u1ed1t")
                else:
                    parts.append(vn[donvi])
        elif chuc > 1:
            parts.append(f"{vn[chuc]} m\u01b0\u01a1i")
            if donvi > 0:
                if donvi == 5:
                    parts.append("l\u0103m")
                elif donvi == 1:
                    parts.append("m\u1ed1t")
                else:
                    parts.append(vn[donvi])
    else:
        # tram == 0
        if chuc == 0:
            parts.append(vn[donvi])
        elif chuc == 1:
            parts.append("m\u01b0\u1eddi")
            if donvi > 0:
                if donvi == 5:
                    parts.append("l\u0103m")
                elif donvi == 1:
                    parts.append("m\u1ed1t")
                else:
                    parts.append(vn[donvi])
        else:
            parts.append(f"{vn[chuc]} m\u01b0\u01a1i")
            if donvi > 0:
                if donvi == 5:
                    parts.append("l\u0103m")
                elif donvi == 1:
                    parts.append("m\u1ed1t")
                else:
                    parts.append(vn[donvi])

    return " ".join(parts)


def so_thanh_chu(n):
    """
    Convert an integer to Vietnamese words.
    E.g. 2302410 -> "hai triệu, ba trăm linh hai nghìn, bốn trăm mười đồng"
    """
    if n is None or n == 0:
        return "kh\u00f4ng \u0111\u1ed3ng"

    n = int(round(n))
    if n < 0:
        return "(\u00e2m) " + so_thanh_chu(-n)

    # Split into groups of 3 digits from right
    units = ["", "ngh\u00ecn", "tri\u1ec7u", "t\u1ef7", "ngh\u00ecn t\u1ef7"]
    groups = []
    temp = n
    while temp > 0:
        groups.append(temp % 1000)
        temp //= 1000

    parts = []
    for i in range(len(groups) - 1, -1, -1):
        if groups[i] == 0:
            continue
        text = _doc_so_hang(groups[i])
        if text:
            unit = units[i] if i < len(units) else ""
            parts.append(f"{text} {unit}".strip())

    result = ", ".join(parts) if parts else "kh\u00f4ng"
    # Capitalize first letter
    result = result[0].upper() + result[1:]
    return result + " \u0111\u1ed3ng"


PAYMENT_SCHEDULE_RATES = (
    ("T/T+1", 10),
    ("T+30", 20),
    ("T+60", 25),
    ("T+90", 25),
    ("T+120", 20),
)
PAYMENT_SCHEDULE_DATE_OFFSETS = (1, 30, 60, 90, 120)


def build_payment_schedule(total_amount):
    """Tạo lịch thanh toán 5 đợt; đợt cuối bù phần lệch làm tròn."""
    total = max(0, int(round(float(total_amount or 0))))
    schedule = []
    allocated = 0
    last_index = len(PAYMENT_SCHEDULE_RATES) - 1
    for index, (label, percent) in enumerate(PAYMENT_SCHEDULE_RATES):
        amount = total - allocated if index == last_index else int((total * percent / 100) + 0.5)
        allocated += amount
        schedule.append({
            "label": label,
            "percent": percent,
            "amount": amount,
        })
    return schedule


def payment_planning_base_date(p):
    """Return the date used as T for Payment Planning, based on the phieu date."""
    try:
        return datetime.strptime(
            f"{p.get('nam')}-{p.get('thang')}-{p.get('ngay')}", "%Y-%m-%d"
        ).date()
    except (TypeError, ValueError):
        try:
            return datetime.strptime(str(p.get("created_at") or "")[:10], "%Y-%m-%d").date()
        except (TypeError, ValueError):
            return None


def payment_planning_amounts(chung_tu_list, cash_amount):
    """Tính 3 giá trị chính của phụ lục Payment Planning từ chứng từ XNCK."""
    total_trade = 0
    for item in chung_tu_list or []:
        if item.get("loai") == "Bảng kê":
            total_trade += abs(float(item.get("gia_tri") or 0))
    cash = int(round(float(cash_amount or 0)))
    product_conversion = int(round(total_trade - cash))
    return {
        "total_trade": int(round(total_trade)),
        "product_conversion": product_conversion,
        "cash_amount": cash,
    }


def payment_planning_profile_for_plant(plant):
    """Return the Payment Planning legal text profile for the saved store plant."""
    return PAYMENT_PLANNING_PROFILES["cao"] if str(plant or "").strip() == "2122" else PAYMENT_PLANNING_PROFILES["pnj"]


def prepare_payment_planning_for_output(row, settings=None):
    """Return display data for Payment Planning HTML/PDF/XLSX."""
    p = prepare_phieu_for_output(row, settings)
    profile = payment_planning_profile_for_plant(p.get("plant"))
    amounts = payment_planning_amounts(p.get("chung_tu", []), p.get("tong_ck", 0))
    use_bk_ref = bool(int(p.get("use_bk_ref") or 0))
    bk_numbers = []
    for item in p.get("chung_tu", []):
        if item.get("loai") != "Bảng kê":
            continue
        so_ct = remove_all_whitespace((item.get("bk_ref") if use_bk_ref else "") or item.get("so_ct", ""))
        if so_ct and so_ct not in bk_numbers:
            bk_numbers.append(so_ct)
    signer_title_map = {
        "tvv": "Tư Vấn Viên",
        "cht": "Cửa Hàng Trưởng",
        "kt1": "Kế Toán Cửa Hàng",
        "kt2": "Kế Toán Cửa Hàng",
    }
    p.update(amounts)
    p["total_trade_words"] = so_thanh_chu(amounts["total_trade"])
    p["product_conversion_words"] = so_thanh_chu(amounts["product_conversion"])
    p["cash_amount_words"] = so_thanh_chu(amounts["cash_amount"])
    p["seller_signature_name"] = str(p.get("ten_kh") or "").upper()
    p["planning_representative_name"] = str((settings or {}).get("cht_name") or STAFF["cua_hang_truong"]).strip()
    p["buyer_signature_name"] = p["planning_representative_name"].upper()
    p["buyer_signature_title"] = "Cửa Hàng Trưởng"
    p["planning_profile"] = profile
    p["planning_abbr"] = profile["abbr"]
    p["planning_representative_abbr"] = profile["representative_abbr"]
    p["planning_company_name"] = profile["company_name"]
    p["planning_store_name"] = profile["store_name"]
    p["planning_tax_code"] = profile["tax_code"]
    p["planning_pnj_address"] = profile["address"]
    p["planning_pnj_contact"] = profile["contact"]
    p["planning_definition_rows"] = profile["definition_rows"]
    p["planning_has_working_day_definition"] = profile["has_working_day_definition"]
    p["planning_effectiveness_items"] = profile["effectiveness_items"]
    p["planning_bk_numbers"] = ",".join(bk_numbers) or p.get("so_bk") or "__________"
    p["planning_sign_date"] = f"{p.get('ngay') or '__'} / {p.get('thang') or '__'} / {p.get('nam') or '____'}"
    p["show_payment_dates"] = 1 if int(p.get("show_payment_dates") or 0) else 0
    base_date = payment_planning_base_date(p) if p["show_payment_dates"] else None
    planning_schedule = []
    for index, item in enumerate(build_payment_schedule(amounts["cash_amount"])):
        due_date = ""
        if base_date and index < len(PAYMENT_SCHEDULE_DATE_OFFSETS):
            due_date = (base_date + timedelta(days=PAYMENT_SCHEDULE_DATE_OFFSETS[index])).strftime("%d/%m/%Y")
        planning_schedule.append({
            **item,
            "date": due_date,
        })
    p["planning_schedule"] = planning_schedule
    p["payment_method_label"] = "☐ Chuyển khoản    ☐ Khác"
    p["planning_file_title"] = f"Thoa thuan thu doi {ascii_filename_part(p.get('ten_kh'), 30)} {p.get('id')}"
    return p


# ---------------------------------------------------------------------------
# Business logic helpers
# ---------------------------------------------------------------------------

def parse_sap_paste(raw_text):
    """
    Parse tab-separated SAP ZFIE0029 data.

    Columns:
      0: Document Number (14xx=HĐ, 25xx=BK, 16xx=Cọc)
      1: Reference
      2: Amount (trailing '-' = negative = phải CK; positive = phải thu HĐ)
      3: Sp. G/L Trans.Type (usually empty)
      4: Sales Document (90xxxxxxxx for HĐ)
      5: Purchasing Document (optional)

    Logic:
      - Amount positive → Hóa đơn, số CT from col4 (90xx), CK -= amount
      - Amount negative + col0 25xx → Bảng kê, CK += amount
      - Amount negative + col0 16xx → Biên nhận cọc, CK += amount
    """
    records = []
    if not raw_text or not raw_text.strip():
        return records

    for line in raw_text.strip().split("\n"):
        line = line.strip()
        if not line:
            continue

        cols = line.split("\t")
        if len(cols) < 3:
            continue

        doc_num = cols[0].strip()
        so_ref = cols[1].strip() if len(cols) > 1 else ""

        # Parse amount
        amt_str = cols[2].strip() if len(cols) > 2 else "0"
        is_negative = amt_str.endswith("-")
        cleaned = re.sub(r"[.,\s\-]", "", amt_str)
        try:
            amount = abs(float(cleaned))
        except ValueError:
            amount = 0

        # Column 4: Sales Document (for HĐ)
        so_hd = cols[4].strip() if len(cols) > 4 and cols[4].strip() else ""

        # Determine type based on sign and document number
        # so_ct = số hiển thị (90xx cho HĐ, 44xx cho BK lấy từ so_bk input)
        # doc_num = số SAP gốc (14xx, 25xx, 16xx...)
        if not is_negative:
            loai = "Hóa đơn"
            so_ct = so_hd if so_hd else doc_num  # 90xxxxxxxx
            gia_tri = amount
        elif re.match(r"^25\d{8}$", doc_num):
            loai = "Bảng kê"
            so_ct = doc_num  # tạm dùng 25xx, frontend sẽ thay bằng so_bk (44xx)
            gia_tri = amount
        elif re.match(r"^16\d{8}$", doc_num):
            loai = "Biên nhận cọc"
            so_ct = doc_num
            gia_tri = amount
        elif re.match(r"^15\d{8}$", doc_num):
            loai = "HBTL"
            so_ct = doc_num
            gia_tri = amount
        elif re.match(r"^26\d{8}$", doc_num):
            loai = "Phải CK khác"
            so_ct = doc_num
            gia_tri = amount
        else:
            loai = "Phải CK khác" if is_negative else "Phải thu khác"
            so_ct = doc_num
            gia_tri = amount

        records.append({
            "doc_num": doc_num,  # số SAP gốc
            "loai": loai,
            "so_ct": so_ct,
            "gia_tri": gia_tri,
            "gio": "",
        })

    return records


def calc_tong_ck(chung_tu_list):
    """
    Total CK = BK + Cọc - HĐ
    """
    total = 0
    for r in chung_tu_list:
        if r["loai"] in ("Hóa đơn", "Phải thu khác", "Thuế TNCN"):
            total -= r["gia_tri"]
        else:
            total += r["gia_tri"]
    return total


def calc_ngay_tt(created_at_str=None):
    """
    Payment date rule:
    - If created before 16:30 on a weekday -> same day
    - If after 16:30 or weekend -> next business day
    """
    if created_at_str:
        now = datetime.strptime(created_at_str, "%Y-%m-%d %H:%M:%S")
    else:
        now = datetime.now()

    cutoff = now.replace(hour=16, minute=30, second=0, microsecond=0)

    if now.weekday() < 5 and now <= cutoff:
        # Weekday before 16:30 -> same day
        return now.strftime("%Y-%m-%d")
    else:
        # Find next business day
        candidate = now + timedelta(days=1)
        while candidate.weekday() >= 5:  # Skip Saturday(5) and Sunday(6)
            candidate += timedelta(days=1)
        return candidate.strftime("%Y-%m-%d")


def normalize_account_number(value):
    """Giữ chữ/số tài khoản, chỉ bỏ khoảng trắng và viết hoa."""
    return re.sub(r"\s+", "", str(value or "").strip()).upper()


def remove_all_whitespace(value):
    """Bỏ toàn bộ ký tự trắng trong các mã định danh để tránh trùng do khoảng trắng."""
    return re.sub(r"\s+", "", str(value or "").strip())


def sanitize_chung_tu_list(chung_tu_list):
    """Chuẩn hóa chứng từ trước khi lưu/in: mã số bỏ khoảng trắng, nội dung khác chỉ trim."""
    cleaned = []
    for item in chung_tu_list or []:
        if not isinstance(item, dict):
            continue
        row = dict(item)
        row["loai"] = str(row.get("loai", "")).strip()
        row["doc_num"] = remove_all_whitespace(row.get("doc_num", ""))
        row["so_ct"] = remove_all_whitespace(row.get("so_ct", ""))
        row["bk_ref"] = remove_all_whitespace(row.get("bk_ref", ""))
        row["gio"] = str(row.get("gio", "")).strip()
        cleaned.append(row)
    return cleaned


def ascii_filename_part(value, max_length=32):
    """Chuyển tiếng Việt có dấu sang ASCII an toàn cho tên file tải xuống."""
    text = str(value or "").strip().upper().replace("Đ", "D").replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:max_length].strip() or "PHIEU"


def print_html_for_pdf(html):
    """Chuẩn hóa HTML in trực tiếp để renderer server xuất PDF từ cùng template."""
    static_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static").replace("\\", "/")
    static_url = "file:///" + static_dir.lstrip("/")
    html = html.replace('src="/static/', f'src="{static_url}/')
    html = html.replace("src='/static/", f"src='{static_url}/")
    html = html.replace('href="/static/', f'href="{static_url}/')
    html = html.replace("href='/static/", f"href='{static_url}/")
    pdf_css = """
    <style>
        .no-print { display: none !important; }
        body { background: #fff !important; border: 0 !important; box-shadow: none !important; }
    </style>
    """
    return html.replace("</head>", pdf_css + "</head>", 1)


def find_pdf_renderer():
    """Tìm browser/renderer dùng để xuất PDF từ print.html."""
    env_path = os.environ.get("PDF_RENDERER_PATH")
    if env_path and os.path.exists(env_path):
        return env_path
    for name in ("chromium", "chromium-browser", "google-chrome", "google-chrome-stable", "wkhtmltopdf"):
        path = shutil.which(name)
        if path:
            return path
    return None


def make_pdf_from_print_html(html):
    """Xuất PDF từ chính HTML của trang in để tránh lệch layout giữa In và Tải PDF."""
    html = print_html_for_pdf(html)
    try:
        from weasyprint import HTML

        buf = io.BytesIO()
        HTML(string=html, base_url=os.path.dirname(os.path.abspath(__file__))).write_pdf(buf)
        buf.seek(0)
        return buf
    except Exception:
        pass

    renderer = find_pdf_renderer()
    if not renderer:
        raise RuntimeError("PDF_RENDERER_MISSING")

    with tempfile.TemporaryDirectory(prefix="phieu_ck_pdf_") as tmpdir:
        html_path = os.path.join(tmpdir, "print.html")
        pdf_path = os.path.join(tmpdir, "print.pdf")
        with open(html_path, "w", encoding="utf-8") as f:
            f.write(html)

        renderer_name = os.path.basename(renderer).lower()
        if "wkhtmltopdf" in renderer_name:
            cmd = [
                renderer,
                "--quiet",
                "--encoding", "utf-8",
                "--print-media-type",
                "--enable-local-file-access",
                "--page-size", "A5",
                "--orientation", "Portrait",
                "--margin-top", "6mm",
                "--margin-bottom", "6mm",
                "--margin-left", "8mm",
                "--margin-right", "8mm",
                html_path,
                pdf_path,
            ]
        else:
            cmd = [
                renderer,
                "--headless",
                "--no-sandbox",
                "--disable-gpu",
                "--no-pdf-header-footer",
                f"--print-to-pdf={pdf_path}",
                html_path,
            ]

        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=30)
        if result.returncode != 0 or not os.path.exists(pdf_path):
            raise RuntimeError("PDF_RENDER_FAILED")
        with open(pdf_path, "rb") as f:
            buf = io.BytesIO(f.read())
        buf.seek(0)
        return buf


def build_qr_url(ngan_hang, so_tk, amount=None, memo=None):
    """
    Build VietQR image URL.
    Format: https://img.vietqr.io/image/{BIN}-{account}-qr_only.png
    """
    # Try exact match first, then search keywords in bank name
    bin_code = BANK_BINS.get(ngan_hang, "")
    if not bin_code:
        nh_upper = ngan_hang.upper()
        for key, val in BANK_BINS.items():
            if key.upper() in nh_upper:
                bin_code = val
                break
    account = normalize_account_number(so_tk)
    if not bin_code or not account:
        return ""

    url = f"https://img.vietqr.io/image/{bin_code}-{urllib.parse.quote(account)}-qr_only.png"

    params = []
    if amount:
        params.append(f"amount={int(amount)}")
    if memo:
        params.append(f"addInfo={memo}")
    if params:
        url += "?" + "&".join(params)

    return url


def format_vnd_amount(amount):
    """Format VND amount with Vietnamese thousand separators for transfer descriptions."""
    try:
        value = int(round(float(amount or 0)))
    except (TypeError, ValueError):
        value = 0
    return f"{value:,}".replace(",", ".")


def build_noi_dung(plant, so_bk, ngay, ten_kh, tong_ck=None):
    """
    Build eOffice QT82 noi_dung string.
    Format: '1305_CK BK {so_bk} ngày {date} cho {ten_kh} - {amount} VND'
    """
    base = f"{plant}_CK BK {so_bk} ngày {ngay} cho {ten_kh}"
    if tong_ck is None:
        return base
    return f"{base} - {format_vnd_amount(tong_ck)} VND"


def find_eoffice_bank_code(ngan_hang):
    """Tìm mã dùng để chọn ngân hàng trên QT82, không gửi cả danh mục ra client."""
    bank_name = str(ngan_hang or "").strip()
    for bank in BANK_LIST:
        if bank["ten_tra_cuu"] and bank["ten_tra_cuu"] in bank_name:
            return bank["eoffice"]
    for bank in BANK_LIST:
        if bank["ten_gd"] and bank["ten_gd"].lower() in bank_name.lower():
            return bank["eoffice"]
    return ""


def build_qt82_payload(phieu, settings):
    """Chuẩn bị bản nháp QT82; không lưu, gửi hay chứa thông tin đăng nhập eOffice."""
    chung_tu = phieu.get("chung_tu") or []
    override_sap_document = re.sub(
        r"\s*,\s*", ", ", remove_all_whitespace(phieu.get("sap_document_override", "")).replace(",", ",")
    ).strip(" ,")
    doc_nums = []
    for item in chung_tu:
        doc_num = remove_all_whitespace(item.get("doc_num", ""))
        if doc_num and doc_num not in doc_nums:
            doc_nums.append(doc_num)

    sap_placeholder = not override_sap_document and not doc_nums
    sap_document = override_sap_document or (", ".join(doc_nums) if doc_nums else "1234")
    account_name = str(phieu.get("ten_tk", "") or "").strip()
    account_number = normalize_account_number(phieu.get("so_tk", ""))
    bank_query = find_eoffice_bank_code(phieu.get("ngan_hang", ""))
    customer_code = remove_all_whitespace(phieu.get("ma_kh", ""))
    customer_name = str(phieu.get("ten_kh", "") or "").strip()
    cccd = re.sub(r"\D", "", str(phieu.get("cccd", "") or ""))
    total_amount = int(round(float(phieu.get("tong_ck", 0) or 0)))
    detail_documents = []
    for item in chung_tu:
        detail_document = remove_all_whitespace(item.get("so_ct", ""))
        if detail_document and detail_document not in detail_documents:
            detail_documents.append(detail_document)

    checks = [
        {"key": "customer", "label": "Mã và tên khách hàng", "ok": bool(customer_code and customer_name)},
        {"key": "account_name", "label": "Tên tài khoản đã xác minh", "ok": bool(account_name)},
        {
            "key": "account_number",
            "label": "Số tài khoản không có khoảng trắng hoặc dấu gạch",
            "ok": bool(re.fullmatch(r"[A-Za-z0-9]+", account_number)),
        },
        {"key": "bank", "label": "Mã ngân hàng eOffice", "ok": bool(bank_query)},
        {"key": "cccd", "label": "CCCD đủ 12 số", "ok": bool(re.fullmatch(r"\d{12}", cccd))},
        {"key": "amount", "label": "Tổng chuyển khoản lớn hơn 0", "ok": total_amount > 0},
        {"key": "details", "label": "Có chi tiết thanh toán", "ok": bool(chung_tu)},
    ]

    return {
        "version": 1,
        "phieuId": int(phieu.get("id") or 0),
        "formUrl": normalize_qt82_form_url(settings.get("qt82_form_url")) or DEFAULT_QT82_FORM_URL,
        "purpose": "Thanh toán cho khách hàng(Mua lại)",
        "currency": "VND",
        "managerApproval": "Có",
        "storeManagerQuery": settings.get("qt82_store_manager_query", "my.hth").strip() or "my.hth",
        "storeManagerName": settings.get("cht_name", "").strip(),
        "paymentObjectName": customer_name,
        "paymentObjectCode": customer_code,
        "costGroup": "Hàng hóa(ML)",
        "requestContent": str(phieu.get("eo_noi_dung", "") or ""),
        "paymentMethod": "Bank transfer – Chuyển khoản",
        "paymentAmount": total_amount,
        "sapDocument": sap_document,
        "sapPlaceholder": sap_placeholder,
        "companyCode": "1000",
        "desiredDateMode": "browser_today",
        "accountName": account_name,
        "accountNumber": account_number,
        "bankQuery": bank_query,
        "customerId": cccd,
        "detailCount": len(chung_tu),
        "detailDocuments": detail_documents,
        "ready": all(item["ok"] for item in checks),
        "checks": checks,
    }


DNCK_OBJECT_TYPES = {
    "employee": "Nhân viên",
    "vendor": "Nhà cung cấp",
    "customer": "Khách hàng",
}

DNCK_APPROVER_OPTIONS = {
    "hoang": {"query": "hoang.vp", "name": "hoang.vp@pnj.com.vn"},
    "lanh": {"query": "lanh.nt01", "name": "lanh.nt01@pnj.com.vn"},
    "none": None,
}


def infer_dnck_object_type(object_code, requested_type=None):
    """Suy luận loại đối tượng DNCK từ mã, nhưng vẫn cho giao diện ghi đè."""
    requested = str(requested_type or "").strip().lower()
    if requested in DNCK_OBJECT_TYPES:
        return requested
    code = remove_all_whitespace(object_code).upper()
    if re.fullmatch(r"E01[A-Z0-9]*", code):
        return "employee"
    requested_purpose = str(requested_type or "").strip().lower()
    if "nhà cung cấp" in requested_purpose or "nha cung cap" in requested_purpose:
        return "vendor"
    if "nhân viên" in requested_purpose or "nhan vien" in requested_purpose:
        return "employee"
    return "customer"


def default_dnck_purpose(object_type):
    if object_type == "employee":
        return "Thanh toán cho nhân viên"
    if object_type == "vendor":
        return "Thanh toán cho nhà cung cấp"
    return "Thanh toán cho khách hàng"


def sanitize_dnck_detail(detail_list, fallback_label, fallback_amount, sap_document):
    """Chuẩn hóa chi tiết thanh toán DNCK để dùng chung Template TT."""
    cleaned = []
    for item in detail_list or []:
        if not isinstance(item, dict):
            continue
        amount = int(float(item.get("amount") or item.get("gia_tri") or 0))
        label = str(item.get("label") or item.get("loai") or fallback_label or "").strip()
        document = remove_all_whitespace(item.get("document") or item.get("so_ct") or "")
        identity = remove_all_whitespace(item.get("identity") or item.get("tax_id") or "")
        note = str(item.get("note") or item.get("ghi_chu") or "").strip()
        if amount == 0 and not label and not document and not identity and not note:
            continue
        cleaned.append({
            "label": label or fallback_label or "Thanh Toán Khác",
            "amount": amount,
            "document": document or remove_all_whitespace(sap_document) or "DNCK",
            "identity": identity,
            "note": note,
        })
    if not cleaned and fallback_amount > 0:
        cleaned.append({
            "label": fallback_label or "Thanh toán khác",
            "amount": int(fallback_amount),
            "document": remove_all_whitespace(sap_document) or "DNCK",
            "identity": "",
            "note": "",
        })
    return cleaned


def normalize_dnck_hashtags(value):
    """Chuẩn hóa hashtag DNCK: chữ thường, không dấu # lưu DB, giữ thứ tự nhập."""
    if isinstance(value, list):
        raw_items = value
    else:
        raw_text = str(value or "")
        raw_items = re.split(r"[\s,;]+", raw_text)
    tags = []
    seen = set()
    for item in raw_items:
        tag = str(item or "").strip().lower()
        tag = tag.lstrip("#")
        tag = re.sub(r"[^0-9a-zA-Z_À-ỹ-]+", "", tag)
        if not tag or tag in seen:
            continue
        seen.add(tag)
        tags.append(tag)
    return tags[:20]


def dnck_row_to_dict(row):
    d = row_to_dict(row)
    d["source"] = "dnck"
    if not d:
        return None
    try:
        d["detail"] = json.loads(d.get("detail_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        d["detail"] = []
    try:
        d["hashtags"] = normalize_dnck_hashtags(json.loads(d.get("hashtags_json") or "[]"))
    except (json.JSONDecodeError, TypeError):
        d["hashtags"] = normalize_dnck_hashtags(d.get("hashtags_json") or "")
    try:
        d["reference_links"] = json.loads(d.get("reference_links_json") or "[]")
    except (json.JSONDecodeError, TypeError):
        d["reference_links"] = []
    return d


def build_dnck_qt82_payload(dnck, settings):
    """Chuẩn bị draft QT82 cho đề nghị CK khác, tách khỏi phiếu BK mua lại."""
    object_type = infer_dnck_object_type(dnck.get("object_code"), dnck.get("object_type") or dnck.get("purpose"))
    amount = int(round(float(dnck.get("amount", 0) or 0)))
    account_number = normalize_account_number(dnck.get("account_number", ""))
    bank_query = find_eoffice_bank_code(dnck.get("bank", ""))
    sap_document = remove_all_whitespace(dnck.get("sap_document", "")) or "1234"
    detail = sanitize_dnck_detail(
        dnck.get("detail") or [],
        dnck.get("payment_tag") or "Thanh Toán Khác",
        amount,
        sap_document,
    )
    approvers = []
    option = DNCK_APPROVER_OPTIONS.get(str(dnck.get("approver_option") or "none").strip(), None)
    if option:
        approvers.append(option)
    approvers.append({"query": "my.hth", "name": "my.hth@pnj.com.vn"})
    approval_people = []
    if object_type not in ("employee", "vendor"):
        store_manager_name = settings.get("cht_name", "").strip()
        approval_people.append({
            "role": "Cửa hàng trưởng",
            "query": settings.get("qt82_store_manager_query", "my.hth").strip() or "my.hth",
            "name": store_manager_name,
        })
    for approver in approvers:
        approval_people.append({
            "role": "Người phê duyệt",
            "query": approver.get("query", ""),
            "name": approver.get("name", approver.get("query", "")),
        })
    approval_emails = [
        person.get("name") or person.get("query", "")
        for person in approval_people
        if person.get("name") or person.get("query")
    ]
    checks = [
        {"key": "object", "label": "Có mã và tên đối tượng", "ok": bool(dnck.get("object_code") and dnck.get("object_name"))},
        {"key": "account_name", "label": "Có tên tài khoản", "ok": bool(dnck.get("account_name"))},
        {"key": "account_number", "label": "Số tài khoản hợp lệ", "ok": bool(re.fullmatch(r"[A-Za-z0-9]+", account_number))},
        {"key": "bank", "label": "Có mã ngân hàng eOffice", "ok": bool(bank_query)},
        {"key": "amount", "label": "Số tiền lớn hơn 0", "ok": amount > 0},
        {"key": "content", "label": "Có nội dung đề nghị thanh toán", "ok": bool(str(dnck.get("request_content") or "").strip())},
        {"key": "details", "label": "Có chi tiết thanh toán", "ok": bool(detail)},
    ]
    return {
        "version": 1,
        "source": "dnck",
        "qt82Mode": "thanh_toan_khac",
        "phieuId": int(dnck.get("id") or 0),
        "formUrl": normalize_qt82_form_url(settings.get("qt82_form_url")) or DEFAULT_QT82_FORM_URL,
        "purpose": dnck.get("purpose") or default_dnck_purpose(object_type),
        "approvalLevel": dnck.get("approval_level") or "Cấp cửa hàng",
        "approvalPeople": approval_people,
        "approvalEmails": approval_emails,
        "approvalEmailsText": ", ".join(approval_emails),
        "expenseType": dnck.get("expense_type") or dnck.get("cost_group") or "Khác",
        "currency": "VND",
        "managerApproval": "Có",
        "skipStoreManager": object_type in ("employee", "vendor"),
        "storeManagerQuery": settings.get("qt82_store_manager_query", "my.hth").strip() or "my.hth",
        "storeManagerName": settings.get("cht_name", "").strip(),
        "approvers": approvers,
        "paymentObjectName": str(dnck.get("object_name") or "").strip(),
        "paymentObjectCode": remove_all_whitespace(dnck.get("object_code", "")),
        "costGroup": dnck.get("cost_group") or "Khác",
        "hashtags": dnck.get("hashtags") or [],
        "costLimitRef": dnck.get("cost_limit_ref") or "",
        "referenceLinks": dnck.get("reference_links") or [],
        "referenceNote": dnck.get("reference_note") or "",
        "requestContent": str(dnck.get("request_content") or "").strip(),
        "paymentMethod": "Bank transfer – Chuyển khoản",
        "paymentAmount": amount,
        "sapDocument": sap_document,
        "sapPlaceholder": sap_document == "1234",
        "companyCode": "1000",
        "desiredDateMode": "browser_today",
        "accountName": str(dnck.get("account_name") or "").strip(),
        "accountNumber": account_number,
        "bankQuery": bank_query,
        "customerId": str(dnck.get("identity_value") or "").strip(),
        "detailCount": len(detail),
        "detailDocuments": [item["document"] for item in detail],
        "ready": all(item["ok"] for item in checks),
        "checks": checks,
    }


def build_created_at_from_form(data, chung_tu_list):
    """
    Use the user-selected form date as the source of truth.

    The VPS clock can drift after provider/domain incidents. If we rely only on
    datetime.now(), print/eOffice dates can be one day behind the browser UI.
    """
    ngay_lap = (data.get("ngay_lap") or "").strip()
    created_date = None
    if ngay_lap:
        try:
            created_date = datetime.strptime(ngay_lap, "%Y-%m-%d").date()
        except ValueError:
            created_date = None

    doc_times = []
    for ct in chung_tu_list:
        gio = (ct.get("gio") or "").strip()
        if not gio:
            continue
        for fmt in ("%d/%m/%Y %H:%M", "%H:%M"):
            try:
                parsed = datetime.strptime(gio, fmt)
                doc_times.append(parsed.time())
                break
            except ValueError:
                pass

    now = datetime.now()
    if created_date:
        created_time = max(doc_times) if doc_times else now.time().replace(microsecond=0)
        return datetime.combine(created_date, created_time).strftime("%Y-%m-%d %H:%M:%S")
    return now.strftime("%Y-%m-%d %H:%M:%S")


def find_recent_duplicate_phieu(db, user_id, data, ten_kh, so_bk, tong_ck):
    """
    Return a recent matching phieu if the browser sends the same save request
    repeatedly. This protects iPhone/Safari flows where opening the print page
    can fail silently and users tap Save again.
    """
    ma_kh = remove_all_whitespace(data.get("ma_kh"))
    cccd = remove_all_whitespace(data.get("cccd"))
    so_tk = normalize_account_number(data.get("so_tk"))
    if not ma_kh or not so_bk:
        return None
    cutoff = (datetime.now() - timedelta(minutes=30)).strftime("%Y-%m-%d %H:%M:%S")

    return db.execute(
        """
        SELECT id FROM phieu
        WHERE user_id = ?
          AND ma_kh = ?
          AND ten_kh = ?
          AND cccd = ?
          AND so_tk = ?
          AND so_bk = ?
          AND ABS(COALESCE(tong_ck, 0) - ?) < 1
          AND created_at >= ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (user_id, ma_kh, ten_kh, cccd, so_tk, so_bk, float(tong_ck or 0), cutoff),
    ).fetchone()


def pdf_token_serializer():
    """Serializer for short-lived PDF download tokens."""
    return URLSafeTimedSerializer(app.secret_key, salt="phieu-ck-pdf-download")


def create_pdf_token(phieu_id, user_id):
    """Create a short-lived token scoped to one phieu PDF."""
    return pdf_token_serializer().dumps({
        "purpose": "pdf",
        "phieu_id": int(phieu_id),
        "user_id": int(user_id),
    })


def verify_output_token(token):
    """Return token payload if valid for one printable/downloadable phieu."""
    token = (token or "").strip()
    if not token:
        return None
    try:
        payload = pdf_token_serializer().loads(token, max_age=PDF_TOKEN_MAX_AGE)
    except (BadSignature, SignatureExpired, TypeError, ValueError):
        return None
    try:
        payload_phieu_id = int(payload.get("phieu_id"))
        payload_user_id = int(payload.get("user_id"))
    except (TypeError, ValueError):
        return None
    if payload.get("purpose") != "pdf" or payload_phieu_id <= 0 or payload_user_id <= 0:
        return None
    return {"purpose": "pdf", "phieu_id": payload_phieu_id, "user_id": payload_user_id}


def verify_pdf_token(phieu_id):
    """Return token payload if valid for this phieu, otherwise None."""
    payload = verify_output_token(request.args.get("token"))
    if not payload:
        return None
    try:
        return payload if payload["phieu_id"] == int(phieu_id) else None
    except (TypeError, ValueError):
        return None


def verify_print_token():
    """Return token payload for clean HTML print URLs without exposing row IDs."""
    return verify_output_token(request.view_args.get("token") if request.view_args else "")


def prepare_phieu_for_output(row, settings=None):
    """Return a phieu dict with all display fields used by print/PDF output."""
    d = row_to_dict(row)
    d["tong_ck_chu"] = so_thanh_chu(d["tong_ck"])
    try:
        d["chung_tu"] = json.loads(d["chung_tu_json"]) if d["chung_tu_json"] else []
    except (json.JSONDecodeError, TypeError):
        d["chung_tu"] = []
    d["chung_tu"] = sanitize_chung_tu_list(d["chung_tu"])
    d["ma_kh"] = remove_all_whitespace(d.get("ma_kh", ""))
    d["sdt"] = remove_all_whitespace(d.get("sdt", ""))
    d["cccd"] = remove_all_whitespace(d.get("cccd", ""))
    d["dia_chi"] = str(d.get("dia_chi", "") or "").strip()
    d["so_bk"] = remove_all_whitespace(d.get("so_bk", ""))
    d["plant"] = remove_all_whitespace(d.get("plant", ""))

    try:
        dt = datetime.strptime(d["created_at"], "%Y-%m-%d %H:%M:%S")
        d["created_at_fmt"] = dt.strftime("%H:%M_%d/%m/%Y")
        d["created_at_date"] = dt.strftime("%d/%m/%Y %H:%M")
        d["ngay"] = dt.strftime("%d")
        d["thang"] = dt.strftime("%m")
        d["nam"] = dt.strftime("%Y")
    except (ValueError, TypeError):
        d["created_at_fmt"] = d["created_at"]
        d["created_at_date"] = d["created_at"]
        d["ngay"] = ""
        d["thang"] = ""
        d["nam"] = ""

    raw_sdt = re.sub(r"\D", "", d.get("sdt", ""))
    d["sdt_fmt"] = f"{raw_sdt[:4]} {raw_sdt[4:7]} {raw_sdt[7:]}" if len(raw_sdt) == 10 else d.get("sdt", "")

    raw_tk = normalize_account_number(d.get("so_tk", ""))
    d["so_tk_fmt"] = " ".join([raw_tk[i:i+4] for i in range(0, len(raw_tk), 4)]) if raw_tk.isdigit() else raw_tk

    raw_cccd = re.sub(r"\D", "", d.get("cccd", ""))
    d["cccd_fmt"] = " ".join([raw_cccd[i:i+3] for i in range(0, len(raw_cccd), 3)]) if raw_cccd else d.get("cccd", "")

    ngay_tt_raw = d.get("ngay_tt", "")
    try:
        if " " in ngay_tt_raw:
            dt_tt = datetime.strptime(ngay_tt_raw, "%Y-%m-%d %H:%M")
            d["ngay_tt_fmt"] = dt_tt.strftime("%d/%m/%Y %H:%M")
        else:
            dt_tt = datetime.strptime(ngay_tt_raw, "%Y-%m-%d")
            dt_ca = datetime.strptime(d["created_at"], "%Y-%m-%d %H:%M:%S")
            d["ngay_tt_fmt"] = dt_tt.strftime("%d/%m/%Y") + " " + dt_ca.strftime("%H:%M")
    except (ValueError, TypeError):
        d["ngay_tt_fmt"] = ngay_tt_raw

    bk_times = []
    for ct in d["chung_tu"]:
        if ct.get("loai") in ("Bảng kê", "Biên nhận cọc") and ct.get("gio"):
            try:
                t = datetime.strptime(ct["gio"], "%d/%m/%Y %H:%M")
                bk_times.append(t)
            except ValueError:
                pass
    if bk_times:
        latest_bk = max(bk_times)
        d["created_at_fmt"] = latest_bk.strftime("%H:%M_%d/%m/%Y")

    settings = settings or get_settings()
    nguoi_ki = d.get("nguoi_ki", "tvv")
    nguoi_ki_map = {
        "tvv": d.get("tvv_name", ""),
        "cht": settings.get("cht_name", ""),
        "kt1": settings.get("kt1_name", ""),
        "kt2": settings.get("kt2_name", ""),
    }
    d["nguoi_ki_name"] = nguoi_ki_map.get(nguoi_ki, d.get("tvv_name", ""))
    d["show_payment_time"] = settings.get("show_payment_time", "1") == "1"
    d["payment_schedule"] = build_payment_schedule(d.get("tong_ck", 0))
    d["file_title"] = f"CK {ascii_filename_part(d.get('ten_kh'), 30)} {d.get('id')}"
    return d


def make_phieu_pdf(p):
    """Create a printable PDF file for browsers that block window.print()."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from reportlab.lib.pagesizes import A5
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfbase.pdfmetrics import registerFontFamily
        from reportlab.platypus import (
            Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        )
    except ImportError as exc:
        raise RuntimeError("REPORTLAB_MISSING") from exc

    font_sets = [
        {
            "regular": r"C:\Windows\Fonts\times.ttf",
            "bold": r"C:\Windows\Fonts\timesbd.ttf",
            "italic": r"C:\Windows\Fonts\timesi.ttf",
            "bold_italic": r"C:\Windows\Fonts\timesbi.ttf",
        },
        {
            "regular": "/usr/share/fonts/truetype/msttcorefonts/Times_New_Roman.ttf",
            "bold": "/usr/share/fonts/truetype/msttcorefonts/Times_New_Roman_Bold.ttf",
            "italic": "/usr/share/fonts/truetype/msttcorefonts/Times_New_Roman_Italic.ttf",
            "bold_italic": "/usr/share/fonts/truetype/msttcorefonts/Times_New_Roman_Bold_Italic.ttf",
        },
        {
            "regular": "/usr/share/fonts/truetype/liberation2/LiberationSerif-Regular.ttf",
            "bold": "/usr/share/fonts/truetype/liberation2/LiberationSerif-Bold.ttf",
            "italic": "/usr/share/fonts/truetype/liberation2/LiberationSerif-Italic.ttf",
            "bold_italic": "/usr/share/fonts/truetype/liberation2/LiberationSerif-BoldItalic.ttf",
        },
        {
            "regular": "/usr/share/fonts/truetype/liberation/LiberationSerif-Regular.ttf",
            "bold": "/usr/share/fonts/truetype/liberation/LiberationSerif-Bold.ttf",
            "italic": "/usr/share/fonts/truetype/liberation/LiberationSerif-Italic.ttf",
            "bold_italic": "/usr/share/fonts/truetype/liberation/LiberationSerif-BoldItalic.ttf",
        },
        {
            "regular": "/usr/share/fonts/truetype/freefont/FreeSerif.ttf",
            "bold": "/usr/share/fonts/truetype/freefont/FreeSerifBold.ttf",
            "italic": "/usr/share/fonts/truetype/freefont/FreeSerifItalic.ttf",
            "bold_italic": "/usr/share/fonts/truetype/freefont/FreeSerifBoldItalic.ttf",
        },
        {
            "regular": "/usr/share/fonts/truetype/noto/NotoSerif-Regular.ttf",
            "bold": "/usr/share/fonts/truetype/noto/NotoSerif-Bold.ttf",
            "italic": "/usr/share/fonts/truetype/noto/NotoSerif-Italic.ttf",
            "bold_italic": "/usr/share/fonts/truetype/noto/NotoSerif-BoldItalic.ttf",
        },
        {
            "regular": "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
            "bold": "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf",
            "italic": "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Italic.ttf",
            "bold_italic": "/usr/share/fonts/truetype/dejavu/DejaVuSerif-BoldItalic.ttf",
        },
    ]
    base_font = "Times-Roman"
    bold = "Times-Bold"
    italic = "Times-Italic"
    for font_paths in font_sets:
        if not os.path.exists(font_paths["regular"]):
            continue
        regular_font = font_paths["regular"]
        bold_font = font_paths["bold"] if os.path.exists(font_paths["bold"]) else regular_font
        italic_font = font_paths["italic"] if os.path.exists(font_paths["italic"]) else regular_font
        bold_italic_font = font_paths["bold_italic"] if os.path.exists(font_paths["bold_italic"]) else bold_font
        try:
            pdfmetrics.registerFont(TTFont("PNJSerif", regular_font))
            pdfmetrics.registerFont(TTFont("PNJSerif-Bold", bold_font))
            pdfmetrics.registerFont(TTFont("PNJSerif-Italic", italic_font))
            pdfmetrics.registerFont(TTFont("PNJSerif-BoldItalic", bold_italic_font))
            registerFontFamily(
                "PNJSerif",
                normal="PNJSerif",
                bold="PNJSerif-Bold",
                italic="PNJSerif-Italic",
                boldItalic="PNJSerif-BoldItalic",
            )
        except Exception:
            continue
        base_font = "PNJSerif"
        bold = "PNJSerif-Bold"
        italic = "PNJSerif-Italic"
        break

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A5,
        rightMargin=9 * mm,
        leftMargin=9 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    styles = getSampleStyleSheet()
    normal = ParagraphStyle("NormalPNJ", parent=styles["Normal"], fontName=base_font, fontSize=8.8, leading=10.5)
    small = ParagraphStyle("SmallPNJ", parent=normal, fontSize=7.2, leading=8.5)
    bold_style = ParagraphStyle("BoldPNJ", parent=normal, fontName=bold)
    center_bold = ParagraphStyle("CenterBoldPNJ", parent=bold_style, alignment=TA_CENTER)
    title_style = ParagraphStyle("TitlePNJ", parent=center_bold, fontSize=10.8, leading=13, spaceBefore=3, spaceAfter=5)
    right_italic = ParagraphStyle("RightItalicPNJ", parent=normal, fontName=italic, alignment=TA_RIGHT)
    story = []

    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "logo_pnj.webp")
    logo = Image(logo_path, width=12 * mm, height=7 * mm) if os.path.exists(logo_path) else Paragraph("PNJ", bold_style)
    header_left = Table([[logo, Paragraph("<b>CÔNG TY CP VÀNG BẠC<br/>ĐÁ QUÝ PHÚ NHUẬN</b>", small)]], colWidths=[14 * mm, 46 * mm])
    header_right = Paragraph("<b>CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM</b><br/><b>Độc lập - Tự do - Hạnh phúc</b><br/>______", center_bold)
    story.append(Table([[header_left, header_right]], colWidths=[64 * mm, 70 * mm]))
    story.append(Paragraph(f"Số: CH PNJ NEXT 27 Hà Nội, Huế - {p.get('plant') or '1305'}_{p.get('created_at_fmt') or ''}", small))
    story.append(Paragraph("PHIẾU XÁC NHẬN THÔNG TIN THANH TOÁN CHUYỂN KHOẢN", title_style))

    story.append(Paragraph("<b>1. Thông tin Khách Hàng</b>", bold_style))
    story.append(Table([
        [Paragraph(f"Họ & Tên: <b>{p.get('ten_kh') or ''}</b>", normal), Paragraph(f"Mã KH (Vendor)*: <b>{p.get('ma_kh') or ''}</b>", normal)],
        [Paragraph(f"Số điện thoại: <b>{p.get('sdt_fmt') or ''}</b>", normal), Paragraph(f"Số CCCD: <b>{p.get('cccd_fmt') or ''}</b>", normal)],
    ], colWidths=[70 * mm, 64 * mm]))

    if p.get("dia_chi"):
        story.append(Paragraph(f"Địa chỉ: <b>{p.get('dia_chi') or ''}</b>", normal))

    story.append(Paragraph("<b>2. Thông tin thanh toán / Ủy Quyền chuyển khoản</b>", bold_style))
    payment_rows = [
        [Paragraph("Tên tài khoản thụ hưởng:", normal), Paragraph(f"<b>{p.get('ten_tk') or p.get('ten_kh') or ''}</b>", normal)],
        [Paragraph("Số tài khoản thụ hưởng:", normal), Paragraph(f"<b>{p.get('so_tk_fmt') or ''}</b>", normal)],
        [Paragraph("Ngân hàng thụ hưởng:", normal), Paragraph(f"<b>{p.get('ngan_hang') or ''}</b>", normal)],
        [Paragraph("Số tiền chuyển khoản:", normal), Paragraph(f"<b>{float(p.get('tong_ck') or 0):,.0f} đồng</b>", normal)],
        [Paragraph("(Bằng chữ:", normal), Paragraph(f"<i>{p.get('tong_ck_chu') or ''})</i>", normal)],
    ]
    pay_table = Table(payment_rows, colWidths=[42 * mm, 62 * mm])
    qr_cell = ""
    if p.get("qr_url"):
        try:
            with urllib.request.urlopen(p["qr_url"], timeout=5) as resp:
                qr_bytes = io.BytesIO(resp.read())
            qr_cell = Image(qr_bytes, width=27 * mm, height=27 * mm)
        except Exception:
            qr_cell = Paragraph("QR", small)
    story.append(Table([[pay_table, qr_cell]], colWidths=[106 * mm, 28 * mm]))

    story.append(Paragraph("<b>Thông tin chứng từ:</b>", bold_style))
    table_data = [["Loại chứng từ", "Số chứng từ", "Giá trị", "Ngày giờ"]]
    for ct in p.get("chung_tu", []):
        table_data.append([
            ct.get("loai", ""),
            ct.get("so_ct", ""),
            f"{abs(float(ct.get('gia_tri') or 0)):,.0f}",
            ct.get("gio", ""),
        ])
    table_data.append(["TỔNG THANH TOÁN (BK+CỌC+HBTL-HĐ)", "", f"{float(p.get('tong_ck') or 0):,.0f}", ""])
    ct_table = Table(table_data, colWidths=[35 * mm, 36 * mm, 31 * mm, 32 * mm])
    ct_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), base_font),
        ("FONTNAME", (0, 0), (-1, 0), bold),
        ("FONTNAME", (0, -1), (-1, -1), bold),
        ("FONTSIZE", (0, 0), (-1, -1), 7.7),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.yellow),
        ("BACKGROUND", (0, -1), (-1, -1), colors.yellow),
        ("SPAN", (0, -1), (1, -1)),
        ("ALIGN", (0, 1), (0, -2), "CENTER"),
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ("ALIGN", (3, 1), (3, -2), "CENTER"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
    ]))
    story.append(ct_table)
    story.append(Paragraph("Giấy xác nhận thông tin thanh toán có hiệu lực đến lúc khách nhận được tiền vào tài khoản.", normal))
    if p.get("show_payment_time"):
        schedule_html = ["<b>Thời gian thanh toán:</b>"]
        for index, item in enumerate(p.get("payment_schedule") or build_payment_schedule(p.get("tong_ck", 0)), start=1):
            schedule_html.append(
                f"{index}. {item['label']}: {item['percent']}% - tương ứng số tiền "
                f"<b>{int(item['amount']):,} đồng</b>"
            )
        story.append(Paragraph("<br/>".join(schedule_html), normal))
    story.append(Paragraph("* Thông tin liên hệ sau thời hạn thanh toán khách hàng chưa nhận được tiền: <b>0234 3847 588</b>", normal))
    story.append(Paragraph(f"Huế, ngày {p.get('ngay') or ''} tháng {p.get('thang') or ''} năm {p.get('nam') or ''}", right_italic))
    story.append(Spacer(1, 4 * mm))
    story.append(Table([
        [Paragraph("<b>Khách Hàng xác nhận</b><br/><i>(Ký, ghi rõ họ tên)</i>", center_bold), Paragraph("<b>Cửa Hàng xác nhận</b><br/><i>(Ký, ghi rõ họ tên)</i>", center_bold)],
        [Paragraph(f"<b>{p.get('ten_kh') or ''}</b>", center_bold), Paragraph(f"<b>{p.get('nguoi_ki_name') or ''}</b>", center_bold)],
    ], colWidths=[67 * mm, 67 * mm], rowHeights=[18 * mm, 14 * mm]))

    doc.build(story)
    buf.seek(0)
    return buf

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Authentication
# ---------------------------------------------------------------------------

def login_required(f):
    """Decorator: redirect to login page if not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not REQUIRE_LOGIN:
            return f(*args, **kwargs)
        if not session.get("user_id"):
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


@app.before_request
def check_auth():
    """Global auth check — skip for login/static routes."""
    if not REQUIRE_LOGIN:
        return
    allowed = ("login_page", "login_action", "logout_action", "static")
    if request.endpoint in allowed:
        return
    if request.endpoint == "api_print_token" and verify_print_token():
        return
    if request.endpoint == "api_pdf" and verify_pdf_token(request.view_args.get("phieu_id")):
        return
    if not session.get("user_id"):
        return redirect(url_for("login_page"))


@app.route("/login", methods=["GET"])
def login_page():
    if session.get("user_id"):
        return redirect(url_for("index"))
    error = request.args.get("error", "")
    return render_template("login.html", error=error)


@app.route("/login", methods=["POST"])
def login_action():
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    remember = request.form.get("remember")

    user = shared_auth.authenticate(username, password)
    if user:
        session.clear()
        session["user_id"] = user["id"]
        session["user_name"] = user["full_name"] or user["username"]
        session["role"] = user["role"]
        if remember:
            session.permanent = True
            app.permanent_session_lifetime = timedelta(days=30)
        return redirect(url_for("index"))

    return redirect(url_for("login_page", error="Sai tên đăng nhập hoặc mật khẩu"))


@app.route("/logout")
def logout_action():
    session.clear()
    return redirect(url_for("login_page"))


def current_user_id():
    """Return logged-in user's ID, default 1 (admin) for local/no-login mode."""
    return session.get("user_id", 1)


def is_admin():
    """Check if current user has admin role (from shared DB)."""
    if not REQUIRE_LOGIN and not session.get("user_id"):
        return True
    if session.get("role") == "admin":
        return True
    uid = session.get("user_id")
    if uid:
        try:
            user = shared_auth.get_user(uid)
        except Exception:
            return False
        return bool(user and user.get("role") == "admin")
    return False


@app.context_processor
def inject_template_permissions():
    """Expose server-verified permissions to shared navigation templates."""
    admin = is_admin()
    pending_customer_update_count = 0
    if admin:
        store = get_customer_lookup_store()
        if store is not None:
            try:
                pending_customer_update_count = store.pending_candidate_count()
            except CustomerLookupError:
                # Không để lỗi thống kê ảnh hưởng đến toàn bộ giao diện quản trị.
                pass
    return {
        "admin": admin,
        "pending_customer_update_count": pending_customer_update_count,
    }


def _customer_lookup_json(payload, status=200):
    response = jsonify(payload)
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response, status


def _customer_lookup_is_same_origin():
    fetch_site = request.headers.get("Sec-Fetch-Site", "")
    if fetch_site and fetch_site not in ("same-origin", "same-site", "none"):
        return False
    origin = request.headers.get("Origin", "")
    if not origin:
        return True
    try:
        return urllib.parse.urlparse(origin).netloc == request.host
    except Exception:
        return False


def _customer_lookup_admin_required():
    """Return a uniform no-store response for protected customer-data APIs."""
    if not is_admin():
        return _customer_lookup_json(
            {"ok": False, "error": "Bạn không có quyền thực hiện thao tác này."}, 403
        )
    return None


def _record_printed_customer_values(data, phieu_id, user_id, ma_kh, ten_kh, sdt, cccd):
    """Store valid TVV-entered values as encrypted review candidates."""
    store = get_customer_lookup_store()
    if store is None:
        return
    values = {"name": ten_kh, "phone": sdt, "cccd": cccd}
    verified_cccd = _verified_identity_value(ma_kh, "cccd")
    if verified_cccd and remove_all_whitespace(cccd) == verified_cccd:
        values["cccd"] = ""
    try:
        store.record_tvv_values(
            customer_code=ma_kh,
            values=values,
            user_id=user_id,
            phieu_id=phieu_id,
            tvv_code=remove_all_whitespace(data.get("tvv_code", "")),
            tvv_name=str(
                data.get("tvv_name_real", "") or data.get("tvv_name", "")
            ).strip(),
        )
    except Exception:
        # Không để lỗi kho gợi ý làm mất phiếu đã lưu; tuyệt đối không log PII.
        app.logger.warning("Không thể lưu ứng viên cập nhật dữ liệu khách hàng.")


def _set_customer_import_job(job_id, **updates):
    store = get_customer_lookup_store()
    if store is None or not re.fullmatch(r"[A-Za-z0-9_-]{20,64}", job_id):
        return
    job_root = store.db_path.parent / "import-jobs"
    job_root.mkdir(parents=True, exist_ok=True)
    job_path = job_root / f"{job_id}.json"
    with _customer_import_jobs_lock:
        try:
            job = json.loads(job_path.read_text(encoding="utf-8")) if job_path.exists() else {"id": job_id}
        except (OSError, ValueError, TypeError):
            job = {"id": job_id}
        job.update(updates)
        job["updated_at"] = time.time()
        temporary = job_root / f".{job_id}.{secrets.token_hex(6)}.tmp"
        temporary.write_text(
            json.dumps(job, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
        )
        try:
            os.chmod(temporary, 0o600)
        except OSError:
            pass
        os.replace(temporary, job_path)


def _get_customer_import_job(job_id):
    store = get_customer_lookup_store()
    if store is None or not re.fullmatch(r"[A-Za-z0-9_-]{20,64}", str(job_id or "")):
        return None
    job_path = store.db_path.parent / "import-jobs" / f"{job_id}.json"
    try:
        job = json.loads(job_path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return None
    return job if isinstance(job, dict) else None


def _customer_import_lock_path():
    store = get_customer_lookup_store()
    if store is None:
        return None
    root = store.db_path.parent / "import-jobs"
    root.mkdir(parents=True, exist_ok=True)
    try:
        os.chmod(root, 0o700)
    except OSError:
        pass
    return root / "active.lock"


def _acquire_customer_import_job(job_id):
    lock_path = _customer_import_lock_path()
    if lock_path is None:
        return False
    if lock_path.exists() and time.time() - lock_path.stat().st_mtime > 86400:
        lock_path.unlink(missing_ok=True)
    try:
        descriptor = os.open(lock_path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
    except FileExistsError:
        return False
    try:
        os.write(
            descriptor,
            json.dumps({"job_id": job_id, "created_at": time.time()}).encode("utf-8"),
        )
    finally:
        os.close(descriptor)
    return True


def _get_active_customer_import_job():
    lock_path = _customer_import_lock_path()
    if lock_path is None or not lock_path.exists():
        return None
    try:
        lock = json.loads(lock_path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return None
    return _get_customer_import_job(lock.get("job_id"))


def _public_customer_import_job(job):
    allowed = (
        "id", "status", "file_count", "processed_rows", "total_rows",
        "message", "error", "validation", "result", "created_at", "updated_at",
    )
    return {key: job.get(key) for key in allowed if key in job}


def _release_customer_import_job(job_id):
    lock_path = _customer_import_lock_path()
    if lock_path is None or not lock_path.exists():
        return
    try:
        lock = json.loads(lock_path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return
    if lock.get("job_id") == job_id:
        lock_path.unlink(missing_ok=True)


def _run_customer_import_job(job_id, paths, upload_dir):
    store = get_customer_lookup_store()
    try:
        if store is None:
            raise CustomerLookupError("Kho dữ liệu khách hàng chưa sẵn sàng.")
        _set_customer_import_job(
            job_id, status="validating", message="Đang kiểm tra toàn bộ dữ liệu tải lên."
        )
        validation = store.validate_import_files(paths)
        _set_customer_import_job(
            job_id,
            status="backing_up",
            validation=validation,
            total_rows=validation["source_rows"],
            message="Dữ liệu hợp lệ. Đang sao lưu CSDL trước khi cập nhật.",
        )
        backup = store.create_backup("before-web-import")

        def progress(processed):
            _set_customer_import_job(
                job_id,
                status="importing",
                processed_rows=processed,
                message="Đang mã hóa và cập nhật CSDL.",
            )

        _set_customer_import_job(
            job_id,
            status="importing",
            processed_rows=0,
            message="Đang mã hóa và cập nhật CSDL.",
        )
        result = store.import_files(
            paths,
            expected_min=validation["min_customer"],
            expected_max=validation["max_customer"],
            progress=progress,
        )
        summary = store.get_dataset_summary()
        _set_customer_import_job(
            job_id,
            status="completed",
            processed_rows=result["source_rows"],
            result={**result, "backup_name": backup.name, "dataset": summary},
            message="Cập nhật dữ liệu khách hàng đã hoàn tất.",
        )
    except CustomerLookupError as exc:
        _set_customer_import_job(
            job_id, status="failed", error=str(exc), message="Cập nhật không thành công."
        )
    except Exception:
        app.logger.exception("Lỗi nền khi cập nhật CSDL khách hàng; không ghi dữ liệu PII vào log.")
        _set_customer_import_job(
            job_id,
            status="failed",
            error="Có lỗi nội bộ khi cập nhật. CSDL đã được rollback.",
            message="Cập nhật không thành công.",
        )
    finally:
        shutil.rmtree(upload_dir, ignore_errors=True)
        _release_customer_import_job(job_id)


def _verify_customer_lookup_turnstile(token):
    if (
        not token
        or len(token) > 2048
        or not CUSTOMER_LOOKUP_TURNSTILE_SECRET
        or not CUSTOMER_LOOKUP_TURNSTILE_SITEKEY
    ):
        return False
    body = urllib.parse.urlencode(
        {
            "secret": CUSTOMER_LOOKUP_TURNSTILE_SECRET,
            "response": token,
            "remoteip": request.remote_addr or "",
        }
    ).encode("utf-8")
    turnstile_request = urllib.request.Request(
        "https://challenges.cloudflare.com/turnstile/v0/siteverify",
        data=body,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(turnstile_request, timeout=8) as response:
            result = json.loads(response.read().decode("utf-8"))
    except Exception:
        return False
    if not result.get("success"):
        return False
    action = result.get("action")
    if action and action != "customer_lookup":
        return False
    if (
        CUSTOMER_LOOKUP_TURNSTILE_HOSTNAME
        and result.get("hostname") != CUSTOMER_LOOKUP_TURNSTILE_HOSTNAME
    ):
        return False
    return True


def get_accessible_phieu(db, phieu_id):
    """Return a phieu row the current user may read."""
    if is_admin():
        return db.execute("SELECT * FROM phieu WHERE id = ?", (phieu_id,)).fetchone()
    return db.execute(
        "SELECT * FROM phieu WHERE id = ? AND user_id = ?",
        (phieu_id, current_user_id()),
    ).fetchone()


def get_owned_phieu(db, phieu_id):
    """Return a phieu row owned by the current user."""
    return db.execute(
        "SELECT * FROM phieu WHERE id = ? AND user_id = ?",
        (phieu_id, current_user_id()),
    ).fetchone()


@app.route("/")
def index():
    """Main page with input form."""
    settings = get_settings()
    bk_prefix = settings.get("bk_prefix", "4403")
    return render_template(
        "index.html",
        staff=STAFF,
        bank_list=sorted(BANK_BINS.keys()),
        settings=settings,
        bk_prefix=bk_prefix,
        admin=is_admin(),
        customer_lookup_enabled=(get_customer_lookup_store() is not None or get_employee_lookup_store() is not None),
        customer_lookup_turnstile_sitekey=CUSTOMER_LOOKUP_TURNSTILE_SITEKEY,
    )


@app.route("/api/customer-suggestion", methods=["POST"])
def api_customer_suggestion():
    """Trả tối đa một gợi ý cho đúng một trường của đúng một mã KH."""
    if request.content_length is not None and request.content_length > 4096:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)

    store = get_customer_lookup_store()
    if store is None or normalize_customer_code is None or suggestion_for_field is None:
        return _customer_lookup_json(
            {"ok": False, "error": "Tính năng gợi ý chưa sẵn sàng."}, 503
        )

    data = request.get_json(silent=True) or {}
    field = data.get("field")
    if field not in ("name", "phone", "cccd"):
        return _customer_lookup_json({"ok": False, "error": "Trường tra cứu không hợp lệ."}, 400)
    canonical = normalize_customer_code(data.get("customer_code"))
    employee_code = normalize_employee_code(data.get("customer_code")) if normalize_employee_code else None
    is_employee_code = canonical is None and employee_code is not None
    canonical = canonical or employee_code
    if canonical is None:
        return _customer_lookup_json({"ok": True, "suggestions": [], "suggestion": None})

    lookup_session_id = session.get("customer_lookup_session_id")
    if not lookup_session_id:
        lookup_session_id = secrets.token_urlsafe(24)
        session["customer_lookup_session_id"] = lookup_session_id
    principal_id = f"user:{current_user_id()}|ip:{request.remote_addr or ''}"
    key = store.lookup_key(canonical)

    try:
        assessment = store.assess_risk(lookup_session_id, principal_id, key)
        captcha_passed = False
        if assessment.requires_captcha:
            token = str(data.get("turnstile_token") or "")
            if not CUSTOMER_LOOKUP_TURNSTILE_SITEKEY or not CUSTOMER_LOOKUP_TURNSTILE_SECRET:
                store.record_event(
                    session_id=lookup_session_id,
                    principal_id=principal_id,
                    lookup_key=key,
                    requested_field=field,
                    outcome="captcha_unavailable",
                    lookup_performed=False,
                )
                return _customer_lookup_json(
                    {"ok": False, "error": "CAPTCHA local chưa được cấu hình."}, 503
                )
            if not token or not _verify_customer_lookup_turnstile(token):
                store.record_event(
                    session_id=lookup_session_id,
                    principal_id=principal_id,
                    lookup_key=key,
                    requested_field=field,
                    outcome="captcha_required" if not token else "captcha_failed",
                    lookup_performed=False,
                )
                return _customer_lookup_json(
                    {"ok": False, "captcha_required": True}, 403
                )
            captcha_passed = True

        record = store.get_record(canonical) if not is_employee_code else None
        employee_store = get_employee_lookup_store(create=False) if is_employee_code else None
        suggestions = (employee_store.get_suggestions(canonical, field) if employee_store else []) if is_employee_code else store.get_suggestions(canonical, field)
        identity_record = None
        identity_store = get_customer_identity_store(create=False)
        if not is_employee_code and identity_store is not None and field in ("name", "cccd"):
            identity_record = identity_store.get_record(canonical)
        identity_value = ""
        if identity_record and field == "cccd":
            raw_identity = str(identity_record.get("identity_value") or "").strip()
            # Hiện tại biểu mẫu/eOffice chỉ nhận CCCD 12 số. Hộ chiếu vẫn được
            # lưu mã hóa nhưng chưa đưa ra giao diện cho đến khi đổi validation.
            if re.fullmatch(r"[0-9]{12}", raw_identity):
                identity_value = raw_identity
        elif identity_record and field == "name":
            identity_value = str(identity_record.get("verified_name") or "").strip()
            if not identity_value and record is None:
                identity_value = str(identity_record.get("source_name") or "").strip()
        if identity_value:
            suggestions = [
                {"value": identity_value, "source": "verified_bk"},
                *[item for item in suggestions if item.get("value") != identity_value],
            ]
        suggestion = suggestions[0]["value"] if suggestions else None
        store.record_event(
            session_id=lookup_session_id,
            principal_id=principal_id,
            lookup_key=key,
            requested_field=field,
            outcome="suggestion" if suggestion is not None else "no_suggestion",
            lookup_performed=True,
            record_found=record is not None or identity_record is not None or bool(employee_store and employee_store.get_record(canonical)),
            suggestion_shown=bool(suggestions),
            captcha_passed=captcha_passed,
        )
        public_suggestions = [{"value": item["value"]} for item in suggestions]
        return _customer_lookup_json(
            {"ok": True, "suggestions": public_suggestions, "suggestion": suggestion}
        )
    except CustomerLookupError:
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể tra cứu dữ liệu lúc này."}, 503
        )


@app.route("/api/customer-local-profile", methods=["POST"])
def api_customer_local_profile():
    """Trả nhanh Tên KH, SĐT và CCCD từ CSDL local trước khi gọi SAP."""
    if request.content_length is not None and request.content_length > 4096:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)

    store = get_customer_lookup_store()
    if store is None or normalize_customer_code is None:
        return _customer_lookup_json(
            {"ok": False, "error": "Tính năng tra cứu khách hàng chưa sẵn sàng."}, 503
        )

    data = request.get_json(silent=True) or {}
    canonical = normalize_customer_code(data.get("customer_code"))
    employee_code = normalize_employee_code(data.get("customer_code")) if normalize_employee_code else None
    is_employee_code = canonical is None and employee_code is not None
    canonical = canonical or employee_code
    if canonical is None:
        return _customer_lookup_json({"ok": True, "profile": {}})

    lookup_session_id = session.get("customer_lookup_session_id")
    if not lookup_session_id:
        lookup_session_id = secrets.token_urlsafe(24)
        session["customer_lookup_session_id"] = lookup_session_id
    principal_id = f"user:{current_user_id()}|ip:{request.remote_addr or ''}"
    key = store.lookup_key(canonical)

    try:
        assessment = store.assess_risk(lookup_session_id, principal_id, key)
        captcha_passed = False
        if assessment.requires_captcha:
            token = str(data.get("turnstile_token") or "")
            if not CUSTOMER_LOOKUP_TURNSTILE_SITEKEY or not CUSTOMER_LOOKUP_TURNSTILE_SECRET:
                store.record_event(
                    session_id=lookup_session_id,
                    principal_id=principal_id,
                    lookup_key=key,
                    requested_field="profile",
                    outcome="captcha_unavailable",
                    lookup_performed=False,
                )
                return _customer_lookup_json(
                    {"ok": False, "error": "CAPTCHA local chưa được cấu hình."}, 503
                )
            if not token or not _verify_customer_lookup_turnstile(token):
                store.record_event(
                    session_id=lookup_session_id,
                    principal_id=principal_id,
                    lookup_key=key,
                    requested_field="profile",
                    outcome="captcha_required" if not token else "captcha_failed",
                    lookup_performed=False,
                )
                return _customer_lookup_json(
                    {"ok": False, "captcha_required": True}, 403
                )
            captcha_passed = True

        profile = {"name": "", "phone": "", "cccd": ""}
        record = store.get_record(canonical) if not is_employee_code else None
        employee_store = get_employee_lookup_store(create=False) if is_employee_code else None
        employee_record = employee_store.get_record(canonical) if employee_store else None
        identity_record = None
        identity_store = get_customer_identity_store(create=False)
        if not is_employee_code and identity_store is not None:
            identity_record = identity_store.get_record(canonical)

        for field, public_key in (("name", "name"), ("phone", "phone"), ("cccd", "cccd")):
            if is_employee_code:
                suggestions = employee_store.get_suggestions(canonical, field) if employee_store else []
            else:
                suggestions = store.get_suggestions(canonical, field)
            identity_value = ""
            if identity_record and field == "cccd":
                raw_identity = str(identity_record.get("identity_value") or "").strip()
                if re.fullmatch(r"[0-9]{12}", raw_identity):
                    identity_value = raw_identity
            elif identity_record and field == "name":
                identity_value = str(identity_record.get("verified_name") or "").strip()
                if not identity_value and record is None:
                    identity_value = str(identity_record.get("source_name") or "").strip()
            if identity_value:
                profile[public_key] = identity_value
            elif suggestions:
                profile[public_key] = suggestions[0]["value"]

        found = record is not None or identity_record is not None or employee_record is not None
        shown = any(bool(value) for value in profile.values())
        store.record_event(
            session_id=lookup_session_id,
            principal_id=principal_id,
            lookup_key=key,
            requested_field="profile",
            outcome="suggestion" if shown else "no_suggestion",
            lookup_performed=True,
            record_found=found,
            suggestion_shown=shown,
            captcha_passed=captcha_passed,
        )
        return _customer_lookup_json({"ok": True, "profile": profile})
    except CustomerLookupError:
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể tra cứu dữ liệu khách hàng lúc này."}, 503
        )


@app.route("/api/billing-suggestions", methods=["POST"])
def api_billing_suggestions():
    """Return recent ERP billing suggestions for one customer code."""
    if request.content_length is not None and request.content_length > 4096:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)

    data = request.get_json(silent=True) or {}
    customer_code = data.get("customer_code")
    if not erp_billing.normalize_customer_code(customer_code):
        return _customer_lookup_json({"ok": True, "suggestions": []})

    try:
        suggestions = erp_billing.billing_suggestions(
            customer_code=customer_code,
            target_date=data.get("billing_date"),
            lookback_days=data.get("lookback_days", erp_billing.DEFAULT_LOOKBACK_DAYS),
            limit=data.get("limit", erp_billing.MAX_SUGGESTIONS),
        )
    except Exception:
        app.logger.exception("Không thể lấy gợi ý hóa đơn ERP.")
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể lấy gợi ý hóa đơn lúc này."}, 503
        )
    return _customer_lookup_json({"ok": True, "suggestions": suggestions})


@app.route("/api/purchase-order-suggestions", methods=["POST"])
def api_purchase_order_suggestions():
    """Return recent ERP buyback purchase-order suggestions for one customer code."""
    if request.content_length is not None and request.content_length > 4096:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)

    data = request.get_json(silent=True) or {}
    customer_code = data.get("customer_code")
    if not erp_purchase_orders.normalize_customer_code(customer_code):
        return _customer_lookup_json({"ok": True, "suggestions": []})

    try:
        suggestions = erp_purchase_orders.purchase_order_suggestions(
            customer_code=customer_code,
            target_date=data.get("purchase_order_date"),
            lookback_days=data.get("lookback_days", erp_purchase_orders.DEFAULT_LOOKBACK_DAYS),
            limit=data.get("limit", erp_purchase_orders.MAX_SUGGESTIONS),
        )
    except Exception:
        app.logger.exception("Không thể lấy gợi ý bảng kê ERP.")
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể lấy gợi ý bảng kê lúc này."}, 503
        )
    return _customer_lookup_json({"ok": True, "suggestions": suggestions})


@app.route("/api/purchase-order-customer-profile", methods=["POST"])
def api_purchase_order_customer_profile():
    """Return customer name, phone, and address from the nearest recent ERP buyback PO."""
    if request.content_length is not None and request.content_length > 4096:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)

    data = request.get_json(silent=True) or {}
    customer_code = data.get("customer_code")
    if not erp_purchase_orders.normalize_customer_code(customer_code):
        return _customer_lookup_json({"ok": True, "profile": {}})

    try:
        profile = erp_purchase_orders.purchase_order_customer_profile(
            customer_code=customer_code,
            target_date=data.get("purchase_order_date"),
            lookback_days=data.get("lookback_days", erp_purchase_orders.DEFAULT_LOOKBACK_DAYS),
        )
    except Exception:
        app.logger.exception("Không thể lấy hồ sơ khách hàng từ bảng kê ERP.")
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể lấy thông tin khách hàng từ ERP lúc này."}, 503
        )
    return _customer_lookup_json({"ok": True, "profile": profile})


@app.route("/api/erp-business-partner-profile", methods=["POST"])
def api_erp_business_partner_profile():
    """Return customer profile directly from ERP Business Partner when available."""
    if request.content_length is not None and request.content_length > 4096:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)

    data = request.get_json(silent=True) or {}
    customer_code = data.get("customer_code")
    if not erp_business_partner.normalize_customer_code(customer_code):
        return _customer_lookup_json({"ok": True, "profile": {}})

    try:
        profile = erp_business_partner.business_partner_profile(customer_code)
    except Exception:
        app.logger.exception("Không thể lấy hồ sơ Business Partner ERP.")
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể lấy thông tin Business Partner từ ERP lúc này."}, 503
        )
    return _customer_lookup_json({"ok": True, "profile": profile})


@app.route("/history")
def history_page():
    """History page."""
    return render_template("history.html")


@app.route("/bieu-mau")
def bieu_mau_page():
    """Biểu Mẫu — merged page: BB Hủy BK, F1, F2."""
    settings = get_settings()
    bk_prefix = settings.get("bk_prefix", "4403")
    return render_template("bieu_mau.html", settings=settings, bk_prefix=bk_prefix)


@app.route("/bb-huy")
def bb_huy_page():
    """Redirect old BB Hủy page to Biểu Mẫu."""
    return redirect(url_for('bieu_mau_page'))


@app.route("/bb-huy/print")
def bb_huy_print():
    """BB Hủy Bảng Kê — printable A5 landscape."""
    settings = get_settings()
    so_bk = request.args.get("so_bk", "")
    tvv_name = request.args.get("tvv", "")
    ly_do = request.args.get("ly_do", "")
    ten_kh = request.args.get("ten_kh", "").strip()
    kt_name = request.args.get("kt", settings.get("kt1_name", ""))
    cht_name = settings.get("cht_name", "")
    plant = settings.get("plant", "1305")

    now = datetime.now()
    ngay = request.args.get("ngay", f"{now.day:02d}")
    thang = request.args.get("thang", f"{now.month:02d}")
    nam = request.args.get("nam", str(now.year))
    ngay_str = f"Hôm nay, ngày {ngay} tháng {thang} năm {nam}, tại Cửa Hàng PNJ NEXT 27 Hà Nội - Huế,"

    return render_template("bb_huy_print.html",
        so_bk=so_bk, tvv_name=tvv_name, ly_do=ly_do,
        ten_kh=ten_kh,
        kt_name=kt_name, cht_name=cht_name,
        ngay_str=ngay_str, plant=plant)


@app.route("/doi-thongtin")
def doi_thongtin_page():
    """Redirect old Đổi TT KH page to Biểu Mẫu."""
    return redirect(url_for('bieu_mau_page'))


@app.route("/doi-thongtin/print-f1")
def doi_thongtin_print_f1():
    """F1 — VB đồng ý XLDL cá nhân — printable A4."""
    now = datetime.now()
    return render_template("doi_thongtin_print_f1.html",
        ten_cu=request.args.get("ten_cu", ""),
        ten_moi=request.args.get("ten_moi", ""),
        ns_cu=request.args.get("ns_cu", ""),
        ns_moi=request.args.get("ns_moi", ""),
        sdt_cu=request.args.get("sdt_cu", ""),
        sdt_moi=request.args.get("sdt_moi", ""),
        ngay=request.args.get("ngay", now.strftime("%d")),
        thang=request.args.get("thang", now.strftime("%m")),
        nam=request.args.get("nam", now.strftime("%Y")))


@app.route("/doi-thongtin/print-f2")
def doi_thongtin_print_f2():
    """F2 — Đề nghị khóa dữ liệu — printable A4."""
    now = datetime.now()
    return render_template("doi_thongtin_print_f2.html",
        ho_ten=request.args.get("ho_ten", ""),
        sdt=request.args.get("sdt", ""),
        cccd=request.args.get("cccd", ""),
        ma_kh=request.args.get("ma_kh", ""),
        ngay=request.args.get("ngay", now.strftime("%d")),
        thang=request.args.get("thang", now.strftime("%m")),
        nam=request.args.get("nam", now.strftime("%Y")))


@app.route("/cao-hml/print")
def cao_hml_print():
    """CAO — Mẫu kiểm tra HML printable HTML."""
    product_codes_raw = request.args.get("product_codes", "")
    product_codes = [remove_all_whitespace(line).upper() for line in product_codes_raw.splitlines() if line.strip()]
    try:
        blank_rows = int(request.args.get("blank_rows", "2"))
    except ValueError:
        blank_rows = 2
    blank_rows = max(0, min(blank_rows, 20))
    rows = product_codes + [""] * blank_rows
    if not rows:
        rows = ["", ""]
    return render_template("cao_hml_print.html",
        store=request.args.get("store", "CAO 27 HÀ NỘI - HUẾ").strip().upper(),
        plan=request.args.get("plan", "2122").strip(),
        rows=rows)


@app.route("/eoffice")
def eoffice_index():
    """Trang chuẩn bị QT82 chỉ dành cho ADMIN."""
    if not is_admin():
        return "Bạn không có quyền truy cập trang này.", 403
    db = get_db()
    mode = str(request.args.get("mode") or "phieu").strip().lower()
    if mode in ("dnck", "khac", "other"):
        row = db.execute("SELECT id FROM dnck ORDER BY datetime(created_at) DESC, id DESC LIMIT 1").fetchone()
        if row:
            return redirect(url_for("eoffice_dnck_page", dnck_id=row["id"]))
    else:
        row = db.execute("SELECT id FROM phieu ORDER BY datetime(created_at) DESC, id DESC LIMIT 1").fetchone()
        if row:
            return redirect(url_for("eoffice_page", phieu_id=row["id"]))
    eoffice_mode = "dnck" if mode in ("dnck", "khac", "other") else "phieu"
    response = app.make_response(render_template(
        "eoffice.html",
        phieu=None,
        qt82_payload=None,
        eoffice_mode=eoffice_mode,
    ))
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


@app.route("/api/qt82-extension")
def api_qt82_extension():
    """Tải gói extension QT82; chỉ ADMIN đã đăng nhập mới được phép."""
    if not is_admin():
        response = app.make_response(("Bạn không có quyền tải tiện ích này.", 403))
        response.headers["Cache-Control"] = "no-store, max-age=0"
        response.headers["Pragma"] = "no-cache"
        return response

    try:
        archive, version = build_qt82_extension_archive()
    except (OSError, ValueError, json.JSONDecodeError, zipfile.BadZipFile):
        app.logger.exception("Không thể đóng gói extension QT82")
        response = app.make_response(("Không thể đóng gói tiện ích Chrome.", 500))
        response.headers["Cache-Control"] = "no-store, max-age=0"
        response.headers["Pragma"] = "no-cache"
        return response

    response = send_file(
        archive,
        as_attachment=True,
        download_name=f"PNJ-QT82-Draft-Helper-v{version}.zip",
        mimetype="application/zip",
        max_age=0,
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.route("/eoffice/<int:phieu_id>")
def eoffice_page(phieu_id):
    """Trang chuẩn bị bản nháp QT82 của một phiếu, chỉ dành cho ADMIN."""
    if not is_admin():
        return "Bạn không có quyền truy cập trang này.", 403
    db = get_db()
    row = get_accessible_phieu(db, phieu_id)
    if not row:
        return "Không tìm thấy phiếu.", 404

    d = row_to_dict(row)
    d["source"] = "phieu"
    try:
        d["chung_tu"] = json.loads(d["chung_tu_json"]) if d["chung_tu_json"] else []
    except (json.JSONDecodeError, TypeError):
        d["chung_tu"] = []

    settings = get_settings()
    plant = d.get("plant", settings.get("plant", "1305"))

    # Build eOffice fields
    # Nội dung: "1305_CK BK 4403... ngày 2026-04-09 cho HỒ THỊ MỸ VÂN - 1.234.567 VND"
    try:
        dt = datetime.strptime(d["created_at"], "%Y-%m-%d %H:%M:%S")
        ngay_str = dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        ngay_str = d["created_at"][:10] if d["created_at"] else ""

    # Số BK: ưu tiên từ form input (44xx), fallback từ chứng từ
    so_bk = d.get("so_bk", "")

    d["eo_ma_kh"] = d.get("ma_kh", "")
    d["eo_noi_dung"] = build_noi_dung(plant, so_bk, ngay_str, d.get("ten_kh", ""), d.get("tong_ck"))
    d["eo_ten_tk"] = d.get("ten_tk", "")
    d["eo_so_tk"] = normalize_account_number(d.get("so_tk", ""))
    d["eo_ma_nh"] = find_eoffice_bank_code(d.get("ngan_hang", ""))
    d["eo_cccd"] = re.sub(r"\D", "", d.get("cccd", ""))
    # Tên file: "1305_BK HO THI MY VAN"
    def remove_diacritics(s):
        import unicodedata
        s = s.replace("đ", "d").replace("Đ", "D")
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    d["eo_ten_file"] = f"{plant}_BK {remove_diacritics(d.get('ten_kh', ''))}"
    qt82_payload = build_qt82_payload(d, settings)
    d["eo_so_ct_sap"] = qt82_payload["sapDocument"]

    response = app.make_response(
        render_template("eoffice.html", phieu=d, qt82_payload=qt82_payload, eoffice_mode="phieu")
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.route("/eoffice/dnck/<int:dnck_id>")
def eoffice_dnck_page(dnck_id):
    """Trang chuẩn bị bản nháp QT82 của DNCK, chỉ dành cho ADMIN."""
    if not is_admin():
        return "Bạn không có quyền truy cập trang này.", 403
    row = get_db().execute("SELECT * FROM dnck WHERE id = ?", (dnck_id,)).fetchone()
    if not row:
        return "Không tìm thấy đề nghị CK.", 404

    d = dnck_row_to_dict(row)
    d["source"] = "dnck"
    d["eo_mode_label"] = "ĐNCK thanh toán khác"
    d["eo_mode_note"] = "Dữ liệu từ chức năng tạo đề nghị CK khác, không in phiếu XNCK."
    d["ten_kh"] = d.get("object_name", "")
    d["ma_kh"] = d.get("object_code", "")
    d["tong_ck"] = d.get("amount", 0)
    d["eo_ma_kh"] = d.get("object_code", "")
    d["eo_noi_dung"] = d.get("request_content", "")
    d["eo_ten_tk"] = d.get("account_name", "")
    d["eo_so_tk"] = normalize_account_number(d.get("account_number", ""))
    d["eo_ma_nh"] = find_eoffice_bank_code(d.get("bank", ""))
    d["eo_cccd"] = d.get("identity_value", "")
    d["eo_so_ct_sap"] = remove_all_whitespace(d.get("sap_document", "")) or "1234"
    qt82_payload = build_dnck_qt82_payload(d, get_settings())

    response = app.make_response(
        render_template("eoffice.html", phieu=d, qt82_payload=qt82_payload, eoffice_mode="dnck")
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.route("/dnck")
def dnck_page():
    """Trang tạo đề nghị CK khác, chỉ dành cho ADMIN."""
    if not is_admin():
        return "Bạn không có quyền truy cập trang này.", 403
    settings = get_settings()
    response = app.make_response(render_template(
        "dnck.html",
        settings=settings,
        bank_list=sorted(BANK_BINS.keys()),
        dnck=None,
        qt82_payload=None,
    ))
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    return response


@app.route("/dnck/<int:dnck_id>")
def dnck_detail_page(dnck_id):
    """Xem lại một DNCK đã lưu và chuẩn bị mở QT82."""
    if not is_admin():
        return "Bạn không có quyền truy cập trang này.", 403
    row = get_db().execute("SELECT * FROM dnck WHERE id = ?", (dnck_id,)).fetchone()
    if not row:
        return "Không tìm thấy đề nghị CK.", 404
    dnck = dnck_row_to_dict(row)
    settings = get_settings()
    qt82_payload = build_dnck_qt82_payload(dnck, settings)
    response = app.make_response(render_template(
        "dnck.html",
        settings=settings,
        bank_list=sorted(BANK_BINS.keys()),
        dnck=dnck,
        qt82_payload=qt82_payload,
    ))
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.route("/settings")
def settings_page():
    """Settings page."""
    if not is_admin():
        return "Bạn không có quyền truy cập trang này.", 403
    settings = get_settings()
    customer_import_csrf = session.get("customer_import_csrf") or secrets.token_urlsafe(32)
    session["customer_import_csrf"] = customer_import_csrf
    response = app.make_response(
        render_template(
            "settings.html",
            settings=settings,
            admin=True,
            customer_import_csrf=customer_import_csrf,
            default_qt82_form_url=DEFAULT_QT82_FORM_URL,
        )
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.route("/customer-updates")
def customer_updates_page():
    """ADMIN-only report for reviewing encrypted customer-data candidates."""
    if not is_admin():
        return "Bạn không có quyền truy cập báo cáo này.", 403
    return render_template("customer_updates.html")


def _customer_update_user_label(user_id):
    if not user_id:
        return ""
    try:
        user = shared_auth.get_user(int(user_id))
    except Exception:
        user = None
    if not user:
        return f"Tài khoản #{user_id}"
    return user.get("full_name") or user.get("username") or f"Tài khoản #{user_id}"


def _verified_identity_value(customer_code, field):
    """Return verified legacy-BK value for customer updates when available."""
    if field not in ("name", "cccd") or normalize_customer_code is None:
        return ""
    canonical = normalize_customer_code(customer_code)
    if canonical is None:
        return ""
    identity_store = get_customer_identity_store(create=False)
    if identity_store is None:
        return ""
    record = identity_store.get_record(canonical)
    if not record:
        return ""
    if field == "cccd":
        raw_identity = str(record.get("identity_value") or "").strip()
        return raw_identity if re.fullmatch(r"[0-9]{12}", raw_identity) else ""
    value = str(record.get("verified_name") or "").strip()
    return value or str(record.get("source_name") or "").strip()


@app.route("/api/customer-updates", methods=["GET"])
def api_customer_updates():
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    store = get_customer_lookup_store()
    if store is None:
        return _customer_lookup_json(
            {"ok": False, "error": "Kho dữ liệu khách hàng chưa sẵn sàng."}, 503
        )
    status = request.args.get("status", "pending")
    try:
        page = max(1, int(request.args.get("page", "1")))
        report = store.list_candidate_report(status=status, page=page, page_size=50)
    except (TypeError, ValueError, CustomerLookupError):
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể đọc báo cáo lúc này."}, 503
        )
    user_labels = {}
    for item in report["items"]:
        for key in ("first_user_id", "last_user_id", "reviewed_by"):
            user_id = item.get(key)
            if user_id and user_id not in user_labels:
                user_labels[user_id] = _customer_update_user_label(user_id)
        item["first_user_label"] = user_labels.get(item.get("first_user_id"), "")
        item["last_user_label"] = user_labels.get(item.get("last_user_id"), "")
        item["reviewed_by_label"] = user_labels.get(item.get("reviewed_by"), "")
        if item.get("field") in ("name", "cccd") and not item.get("original_value"):
            verified_value = _verified_identity_value(item.get("customer_code"), item.get("field"))
            if verified_value:
                item["original_value"] = verified_value
    return _customer_lookup_json({"ok": True, "data": report})


@app.route("/api/customer-updates/<int:candidate_id>/review", methods=["POST"])
def api_review_customer_update(candidate_id):
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if request.content_length is not None and request.content_length > 2048:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    action = (request.get_json(silent=True) or {}).get("action")
    if action not in ("approve", "reject"):
        return _customer_lookup_json({"ok": False, "error": "Thao tác không hợp lệ."}, 400)
    store = get_customer_lookup_store()
    if store is None:
        return _customer_lookup_json(
            {"ok": False, "error": "Kho dữ liệu khách hàng chưa sẵn sàng."}, 503
        )
    try:
        changed = store.review_candidate(candidate_id, action, current_user_id())
    except CustomerLookupError:
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể cập nhật yêu cầu lúc này."}, 503
        )
    if not changed:
        return _customer_lookup_json(
            {"ok": False, "error": "Yêu cầu không còn ở trạng thái chờ duyệt."}, 409
        )
    return _customer_lookup_json({"ok": True})


@app.route("/api/customer-import/summary", methods=["GET"])
def api_customer_import_summary():
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    store = get_customer_lookup_store()
    if store is None:
        return _customer_lookup_json(
            {"ok": False, "error": "Kho dữ liệu khách hàng chưa sẵn sàng."}, 503
        )
    try:
        summary = store.get_dataset_summary()
    except CustomerLookupError:
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể đọc thống kê CSDL lúc này."}, 503
        )
    active_job = _get_active_customer_import_job()
    public_job = _public_customer_import_job(active_job) if active_job else None
    return _customer_lookup_json({"ok": True, "data": summary, "active_job": public_job})


@app.route("/api/customer-import", methods=["POST"])
def api_customer_import_upload():
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    csrf = request.headers.get("X-CSRF-Token", "")
    expected_csrf = session.get("customer_import_csrf", "")
    if not csrf or not expected_csrf or not secrets.compare_digest(csrf, expected_csrf):
        return _customer_lookup_json({"ok": False, "error": "Phiên xác nhận không hợp lệ."}, 403)
    if request.content_length is not None and request.content_length > CUSTOMER_IMPORT_MAX_BYTES:
        return _customer_lookup_json({"ok": False, "error": "Tổng dung lượng tải lên vượt quá 500 MB."}, 413)
    if request.form.get("confirmed") != "yes":
        return _customer_lookup_json({"ok": False, "error": "Bạn chưa xác nhận cập nhật CSDL."}, 400)

    files = [item for item in request.files.getlist("files") if item and item.filename]
    if not 1 <= len(files) <= CUSTOMER_IMPORT_MAX_FILES:
        return _customer_lookup_json(
            {"ok": False, "error": "Mỗi lần chỉ được tải từ 1 đến 10 file."}, 400
        )
    store = get_customer_lookup_store()
    if store is None:
        return _customer_lookup_json(
            {"ok": False, "error": "Kho dữ liệu khách hàng chưa sẵn sàng."}, 503
        )

    job_id = secrets.token_urlsafe(18)
    now = time.time()
    if not _acquire_customer_import_job(job_id):
        return _customer_lookup_json(
            {"ok": False, "error": "Đang có một lần cập nhật khác chạy."}, 409
        )
    _set_customer_import_job(
        job_id,
        owner_user_id=current_user_id(),
        status="uploading",
        file_count=len(files),
        processed_rows=0,
        message="Đang nhận file tải lên.",
        created_at=now,
    )

    upload_root = store.db_path.parent / "pending-imports"
    upload_dir = upload_root / job_id
    paths = []
    try:
        upload_root.mkdir(parents=True, exist_ok=True)
        upload_dir.mkdir(parents=False, exist_ok=False)
        for directory in upload_root.iterdir():
            if directory.is_dir() and directory != upload_dir and now - directory.stat().st_mtime > 86400:
                shutil.rmtree(directory, ignore_errors=True)
        for directory in (upload_root, upload_dir):
            try:
                os.chmod(directory, 0o700)
            except OSError:
                pass

        total_size = 0
        for index, uploaded in enumerate(files, 1):
            target = upload_dir / f"upload-{index:02d}.tsv"
            uploaded.save(target)
            size = target.stat().st_size
            if size <= 0:
                raise CustomerLookupError("Có file tải lên bị trống.")
            total_size += size
            if total_size > CUSTOMER_IMPORT_MAX_BYTES:
                raise CustomerLookupError("Tổng dung lượng tải lên vượt quá 500 MB.")
            try:
                os.chmod(target, 0o600)
            except OSError:
                pass
            paths.append(target)

        worker = threading.Thread(
            target=_run_customer_import_job,
            args=(job_id, paths, upload_dir),
            daemon=True,
            name=f"customer-import-{job_id[:8]}",
        )
        worker.start()
        job = _get_customer_import_job(job_id)
        return _customer_lookup_json(
            {"ok": True, "job": _public_customer_import_job(job)}, 202
        )
    except CustomerLookupError as exc:
        shutil.rmtree(upload_dir, ignore_errors=True)
        _set_customer_import_job(job_id, status="failed", error=str(exc))
        _release_customer_import_job(job_id)
        return _customer_lookup_json({"ok": False, "error": str(exc)}, 400)
    except Exception:
        shutil.rmtree(upload_dir, ignore_errors=True)
        _set_customer_import_job(
            job_id, status="failed", error="Không thể lưu file tải lên an toàn."
        )
        _release_customer_import_job(job_id)
        return _customer_lookup_json(
            {"ok": False, "error": "Không thể lưu file tải lên an toàn."}, 500
        )


@app.route("/api/customer-import/<job_id>", methods=["GET"])
def api_customer_import_job(job_id):
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    job = _get_customer_import_job(job_id)
    if not job:
        return _customer_lookup_json({"ok": False, "error": "Không tìm thấy tiến trình."}, 404)
    data = _public_customer_import_job(job)
    return _customer_lookup_json({"ok": True, "job": data})


def _save_identity_upload(uploaded, store):
    """Lưu file tạm quyền 0600 và luôn để caller xóa trong finally."""
    temporary_dir = store.db_path.parent / "identity-upload-tmp"
    temporary_dir.mkdir(parents=True, exist_ok=True)
    try:
        os.chmod(temporary_dir, 0o700)
    except OSError:
        pass
    descriptor, raw_path = tempfile.mkstemp(
        prefix="identity-", suffix=".xlsx", dir=temporary_dir
    )
    os.close(descriptor)
    path = Path(raw_path)
    try:
        uploaded.save(path)
        size = path.stat().st_size
        if size <= 0 or size > CUSTOMER_IDENTITY_IMPORT_MAX_BYTES:
            raise CustomerLookupError("File CCCD bị trống hoặc vượt quá 150 MB.")
        try:
            os.chmod(path, 0o600)
        except OSError:
            pass
        return path
    except Exception:
        path.unlink(missing_ok=True)
        raise


@app.route("/api/customer-identity-import/summary", methods=["GET"])
def api_customer_identity_import_summary():
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    store = get_customer_identity_store(create=True)
    if store is None:
        return _customer_lookup_json(
            {"ok": False, "error": "Kho CCCD chưa sẵn sàng."}, 503
        )
    employee_summary = None
    employee_store = get_employee_lookup_store(create=False)
    if employee_store is not None:
        try:
            employee_summary = employee_store.get_summary()
        except CustomerLookupError:
            employee_summary = None
    return _customer_lookup_json(
        {"ok": True, "data": store.get_summary(), "employee": employee_summary}
    )


@app.route("/api/customer-identity-import/preview", methods=["POST"])
def api_customer_identity_import_preview():
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    csrf = request.headers.get("X-CSRF-Token", "")
    expected_csrf = session.get("customer_import_csrf", "")
    if not csrf or not expected_csrf or not secrets.compare_digest(csrf, expected_csrf):
        return _customer_lookup_json({"ok": False, "error": "Phiên xác nhận không hợp lệ."}, 403)
    if request.content_length is not None and request.content_length > CUSTOMER_IDENTITY_IMPORT_MAX_BYTES:
        return _customer_lookup_json({"ok": False, "error": "File CCCD vượt quá 150 MB."}, 413)
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return _customer_lookup_json({"ok": False, "error": "Bạn chưa chọn file XLSX."}, 400)
    store = get_customer_identity_store(create=True)
    if store is None:
        return _customer_lookup_json({"ok": False, "error": "Kho CCCD chưa sẵn sàng."}, 503)
    path = None
    try:
        path = _save_identity_upload(uploaded, store)
        preview = store.preview_file(path)
        records, _source = select_identity_records(path)
        employee_records = [
            item for item in records
            if normalize_employee_code is not None and normalize_employee_code(item.get("vendor"))
        ]
        preview["employee_with_identity"] = len(employee_records)
        session["customer_identity_preview"] = {
            "sha256": preview["source_sha256"],
            "mode": preview["mode"],
            "user_id": current_user_id(),
            "expires_at": time.time() + 15 * 60,
        }
        public_preview = {
            key: value for key, value in preview.items() if key != "source_sha256"
        }
        return _customer_lookup_json({"ok": True, "data": public_preview})
    except CustomerLookupError as exc:
        return _customer_lookup_json({"ok": False, "error": str(exc)}, 400)
    except Exception:
        app.logger.exception("Lỗi kiểm tra file CCCD; không ghi dữ liệu PII vào log.")
        return _customer_lookup_json({"ok": False, "error": "Không thể kiểm tra file CCCD."}, 500)
    finally:
        if path is not None:
            path.unlink(missing_ok=True)


@app.route("/api/customer-identity-import/apply", methods=["POST"])
def api_customer_identity_import_apply():
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    csrf = request.headers.get("X-CSRF-Token", "")
    expected_csrf = session.get("customer_import_csrf", "")
    if not csrf or not expected_csrf or not secrets.compare_digest(csrf, expected_csrf):
        return _customer_lookup_json({"ok": False, "error": "Phiên xác nhận không hợp lệ."}, 403)
    preview = session.get("customer_identity_preview") or {}
    if (
        request.form.get("confirmed") != "yes"
        or preview.get("user_id") != current_user_id()
        or float(preview.get("expires_at") or 0) < time.time()
        or preview.get("mode") not in ("initial", "periodic")
        or not re.fullmatch(r"[0-9a-f]{64}", str(preview.get("sha256") or ""))
    ):
        return _customer_lookup_json(
            {"ok": False, "error": "Bản kiểm tra đã hết hạn; vui lòng kiểm tra lại file."}, 409
        )
    uploaded = request.files.get("file")
    if not uploaded or not uploaded.filename:
        return _customer_lookup_json({"ok": False, "error": "Bạn chưa gửi lại file đã kiểm tra."}, 400)
    store = get_customer_identity_store(create=True)
    path = None
    try:
        path = _save_identity_upload(uploaded, store)
        records, _source = select_identity_records(path)
        employee_records = [
            item for item in records
            if normalize_employee_code is not None and normalize_employee_code(item.get("vendor"))
        ]
        employee_result = None
        if employee_records:
            employee_store = get_employee_lookup_store(create=True)
            if employee_store is None:
                raise CustomerLookupError("Kho mã NV chưa sẵn sàng.")
            employee_result = employee_store.import_identity_records(
                employee_records, preview["sha256"]
            )
        result = store.import_file(path, preview["sha256"], preview["mode"])
        result["employee_with_identity"] = len(employee_records)
        result["employee_result"] = employee_result
        session.pop("customer_identity_preview", None)
        return _customer_lookup_json({"ok": True, "data": result})
    except CustomerLookupError as exc:
        return _customer_lookup_json({"ok": False, "error": str(exc)}, 400)
    except Exception:
        app.logger.exception("Lỗi cập nhật CSDL CCCD; không ghi dữ liệu PII vào log.")
        return _customer_lookup_json(
            {"ok": False, "error": "Cập nhật CCCD không thành công; CSDL đã rollback."}, 500
        )
    finally:
        if path is not None:
            path.unlink(missing_ok=True)


@app.route("/api/settings", methods=["GET"])
def api_get_settings():
    """Get all settings."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    return _customer_lookup_json({"ok": True, "data": get_settings()})


@app.route("/api/settings", methods=["POST"])
def api_save_settings():
    """Save settings."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if request.content_length is not None and request.content_length > 65536:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return _customer_lookup_json({"ok": False, "error": "Dữ liệu cài đặt không hợp lệ."}, 400)
    if "qt82_form_url" in data:
        qt82_form_url = normalize_qt82_form_url(data.get("qt82_form_url"))
        if not qt82_form_url:
            return _customer_lookup_json(
                {
                    "ok": False,
                    "error": "URL QT82 phải dùng HTTPS, đúng miền eoffice.pnj.com.vn và đường dẫn /workflow/.",
                },
                400,
            )
        data["qt82_form_url"] = qt82_form_url
    if "billing_invoice_days" in data:
        try:
            billing_invoice_days = int(data.get("billing_invoice_days"))
        except (TypeError, ValueError):
            return _customer_lookup_json({"ok": False, "error": "Sá»‘ ngÃ y Ä‘á»c hoÃ¡ Ä‘Æ¡n khÃ´ng há»£p lá»‡."}, 400)
        data["billing_invoice_days"] = str(max(1, min(billing_invoice_days, 31)))
    for flag_key in ("use_bk_ref_default", "show_payment_dates_default"):
        if flag_key in data:
            data[flag_key] = str(settings_flag(data, flag_key, "0"))
    db = get_db()
    for k, v in data.items():
        db.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (k, str(v)))
    db.commit()
    return _customer_lookup_json({"ok": True})


@app.route("/api/da-trinh/<int:phieu_id>", methods=["POST"])
def api_da_trinh(phieu_id):
    """Toggle da_trinh status."""
    data = request.get_json(force=True)
    db = get_db()
    # User thường chỉ thao tác phiếu của mình; admin được thao tác mọi phiếu.
    row = get_accessible_phieu(db, phieu_id)
    if not row:
        return jsonify({"ok": False, "error": "Kh\u00f4ng t\u00ecm th\u1ea5y phi\u1ebfu"}), 404
    db.execute("UPDATE phieu SET da_trinh = ? WHERE id = ?",
               (1 if data.get("da_trinh") else 0, phieu_id))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/phieu/<int:phieu_id>/sap-document", methods=["POST"])
def api_update_sap_document(phieu_id):
    """Allow admins to override the SAP document used for QT82 without changing source documents."""
    if not is_admin():
        return jsonify({"ok": False, "error": "Bạn không có quyền sửa số chứng từ SAP."}), 403
    if request.content_length is not None and request.content_length > 2048:
        return jsonify({"ok": False, "error": "Yêu cầu quá lớn."}), 413
    if not request.is_json or not _customer_lookup_is_same_origin():
        return jsonify({"ok": False, "error": "Yêu cầu không hợp lệ."}), 400

    raw_value = str((request.get_json(silent=True) or {}).get("sap_document", "") or "")
    value = re.sub(r"\s+", "", raw_value).strip(" ,")
    if len(value) > 200 or (value and not re.fullmatch(r"[0-9A-Za-z,._/-]+", value)):
        return jsonify({"ok": False, "error": "Số chứng từ SAP chỉ được gồm chữ/số và dấu , . _ / -."}), 400

    db = get_db()
    row = get_accessible_phieu(db, phieu_id)
    if not row:
        return jsonify({"ok": False, "error": "Không tìm thấy phiếu."}), 404
    db.execute("UPDATE phieu SET sap_document_override = ? WHERE id = ?", (value, phieu_id))
    db.commit()

    refreshed = row_to_dict(db.execute("SELECT * FROM phieu WHERE id = ?", (phieu_id,)).fetchone())
    try:
        refreshed["chung_tu"] = json.loads(refreshed["chung_tu_json"]) if refreshed["chung_tu_json"] else []
    except (json.JSONDecodeError, TypeError):
        refreshed["chung_tu"] = []
    payload = build_qt82_payload(refreshed, get_settings())
    return jsonify({"ok": True, "sap_document": payload["sapDocument"], "sap_placeholder": payload["sapPlaceholder"]})


@app.route("/api/save", methods=["POST"])
def api_save():
    """Save a phieu to the database. Returns the new ID."""
    data = request.get_json(force=True)

    # Parse chung_tu from SAP paste or from already-parsed JSON
    chung_tu_list = data.get("chung_tu", [])
    if isinstance(chung_tu_list, str):
        # Might be raw SAP paste text
        chung_tu_list = parse_sap_paste(chung_tu_list)
    chung_tu_list = sanitize_chung_tu_list(chung_tu_list)

    created_at = build_created_at_from_form(data, chung_tu_list)

    chung_tu_json = json.dumps(chung_tu_list, ensure_ascii=False)
    tong_ck = data.get("tong_ck") or calc_tong_ck(chung_tu_list)
    tong_ck = float(tong_ck)

    # Tính ngày TT từ giờ BK gần nhất + settings thoi_gian_ck (default 48h)
    settings = get_settings()
    ck_hours = int(settings.get("thoi_gian_ck", "48"))
    ngay_tt = data.get("ngay_tt", "")
    if not ngay_tt:
        # Tìm giờ BK gần nhất trong chung_tu
        bk_times = []
        for ct in chung_tu_list:
            if ct.get("loai") in ("Bảng kê", "Biên nhận cọc") and ct.get("gio"):
                try:
                    t = datetime.strptime(ct["gio"], "%d/%m/%Y %H:%M")
                    bk_times.append(t)
                except ValueError:
                    pass
        if bk_times:
            latest_bk = max(bk_times)
            ngay_tt_dt = latest_bk + timedelta(hours=ck_hours)
            ngay_tt = ngay_tt_dt.strftime("%Y-%m-%d %H:%M")
        else:
            ngay_tt = calc_ngay_tt(created_at)

    plant = remove_all_whitespace(data.get("plant", settings.get("plant", "1305")))
    ngan_hang = str(data.get("ngan_hang", "")).strip()
    so_tk = normalize_account_number(data.get("so_tk"))
    ten_kh = str(data.get("ten_kh", "")).strip()
    ma_kh = remove_all_whitespace(data.get("ma_kh", ""))
    sdt = remove_all_whitespace(data.get("sdt", ""))
    cccd = remove_all_whitespace(data.get("cccd", ""))
    dia_chi = str(data.get("dia_chi", "") or "").strip()
    so_bk = remove_all_whitespace(data.get("so_bk", ""))
    use_bk_ref = settings_flag(
        data, "use_bk_ref", settings.get("use_bk_ref_default", "0")
    )
    show_payment_dates = settings_flag(
        data, "show_payment_dates", settings.get("show_payment_dates_default", "1")
    )

    # Build QR URL (only BIN + account, no amount)
    qr_url = build_qr_url(ngan_hang, so_tk)

    # eOffice noi_dung
    ngay_str = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S").strftime("%d/%m/%Y")
    noi_dung = build_noi_dung(plant, so_bk, ngay_str, ten_kh, tong_ck)

    nguoi_ki = data.get("nguoi_ki", "tvv")

    db = get_db()
    user_id = current_user_id()
    requested_status = data.get("status") if data.get("status") in ("draft", "printed") else "draft"
    target_phieu_id = None
    if not data.get("force_create"):
        try:
            target_phieu_id = int(data.get("phieu_id") or data.get("id") or 0) or None
        except (TypeError, ValueError):
            target_phieu_id = None
    if target_phieu_id:
        existing = get_accessible_phieu(db, target_phieu_id)
        if not existing:
            return jsonify({"ok": False, "error": "Không tìm thấy phiếu cũ hoặc bạn không có quyền cập nhật."}), 404
        owner_user_id = int(existing["user_id"] or user_id)
        db.execute("""
            UPDATE phieu
            SET created_at = ?,
                ma_kh = ?,
                ten_kh = ?,
                sdt = ?,
                cccd = ?,
                dia_chi = ?,
                so_tk = ?,
                ten_tk = ?,
                ngan_hang = ?,
                so_bk = ?,
                tvv_code = ?,
                tvv_name = ?,
                cht_name = ?,
                plant = ?,
                chung_tu_json = ?,
                tong_ck = ?,
                ngay_tt = ?,
                status = ?,
                qr_url = ?,
                noi_dung = ?,
                nguoi_ki = ?,
                use_bk_ref = ?,
                show_payment_dates = ?
            WHERE id = ?
        """, (
            created_at,
            ma_kh,
            ten_kh,
            sdt,
            cccd,
            dia_chi,
            so_tk,
            str(data.get("ten_tk", "")).strip(),
            ngan_hang,
            so_bk,
            remove_all_whitespace(data.get("tvv_code", "")),
            str(data.get("tvv_name_real", "") or data.get("tvv_name", "")).strip(),
            str(data.get("cht_name", STAFF["cua_hang_truong"])).strip(),
            plant,
            chung_tu_json,
            tong_ck,
            ngay_tt,
            requested_status,
            qr_url,
            noi_dung,
            nguoi_ki,
            use_bk_ref,
            show_payment_dates,
            target_phieu_id,
        ))
        db.commit()
        if requested_status == "printed":
            _record_printed_customer_values(
                data, target_phieu_id, owner_user_id, ma_kh, ten_kh, sdt, cccd
            )
        return jsonify({
            "ok": True,
            "id": target_phieu_id,
            "updated": True,
            "pdf_token": create_pdf_token(target_phieu_id, owner_user_id),
            "tong_ck": tong_ck,
            "tong_ck_chu": so_thanh_chu(tong_ck),
            "ngay_tt": ngay_tt,
            "qr_url": qr_url,
            "noi_dung": noi_dung,
        })
    duplicate = None if requested_status == "draft" else find_recent_duplicate_phieu(db, user_id, data, ten_kh, so_bk, tong_ck)
    if duplicate:
        if requested_status == "printed":
            db.execute(
                "UPDATE phieu SET status = 'printed' WHERE id = ? AND user_id = ?",
                (duplicate["id"], user_id),
            )
            db.commit()
            _record_printed_customer_values(
                data, duplicate["id"], user_id, ma_kh, ten_kh, sdt, cccd
            )
        return jsonify({
            "ok": True,
            "id": duplicate["id"],
            "pdf_token": create_pdf_token(duplicate["id"], user_id),
            "duplicate": True,
            "message": "Phiếu này đã được lưu trước đó, hệ thống mở lại bản đã lưu.",
        })

    cursor = db.execute("""
        INSERT INTO phieu
            (created_at, ma_kh, ten_kh, sdt, cccd, dia_chi,
             so_tk, ten_tk, ngan_hang, so_bk,
             tvv_code, tvv_name, cht_name, plant,
             chung_tu_json, tong_ck, ngay_tt, status, qr_url, noi_dung, nguoi_ki,
             user_id, use_bk_ref, show_payment_dates)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        created_at,
        ma_kh,
        ten_kh,
        sdt,
        cccd,
        dia_chi,
        so_tk,
        str(data.get("ten_tk", "")).strip(),
        ngan_hang,
        so_bk,
        remove_all_whitespace(data.get("tvv_code", "")),
        str(data.get("tvv_name_real", "") or data.get("tvv_name", "")).strip(),
        str(data.get("cht_name", STAFF["cua_hang_truong"])).strip(),
        plant,
        chung_tu_json,
        tong_ck,
        ngay_tt,
        requested_status,
        qr_url,
        noi_dung,
        nguoi_ki,
        user_id,
        use_bk_ref,
        show_payment_dates,
    ))
    db.commit()
    new_id = cursor.lastrowid

    if requested_status == "printed":
        _record_printed_customer_values(
            data, new_id, user_id, ma_kh, ten_kh, sdt, cccd
        )

    return jsonify({
        "ok": True,
        "id": new_id,
        "pdf_token": create_pdf_token(new_id, user_id),
        "tong_ck": tong_ck,
        "tong_ck_chu": so_thanh_chu(tong_ck),
        "ngay_tt": ngay_tt,
        "qr_url": qr_url,
        "noi_dung": noi_dung,
    })


@app.route("/api/dnck", methods=["POST"])
def api_dnck_save():
    """Lưu đề nghị CK khác để chuẩn bị QT82; chỉ ADMIN."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if request.content_length is not None and request.content_length > 262144:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    data = request.get_json(silent=True) or {}

    prepared = prepare_dnck_record(data)
    if prepared["errors"]:
        return _customer_lookup_json({"ok": False, "error": " ".join(prepared["errors"])}, 400)

    db = get_db()
    cursor = db.execute("""
        INSERT INTO dnck
            (created_at, object_type, object_code, object_name, identity_value, phone,
             account_number, account_name, bank, purpose, approval_level, expense_type,
             cost_group, request_content, sap_document, amount, payment_tag, approver_option,
             detail_json, hashtags_json, reference_note, reference_links_json,
             cost_limit_ref, da_trinh, user_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?)
    """, (
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        prepared["object_type"],
        prepared["object_code"],
        prepared["object_name"],
        prepared["identity_value"],
        prepared["phone"],
        prepared["account_number"],
        prepared["account_name"],
        prepared["bank"],
        prepared["purpose"],
        prepared["approval_level"],
        prepared["expense_type"],
        prepared["cost_group"],
        prepared["request_content"],
        prepared["sap_document"],
        prepared["amount"],
        prepared["payment_tag"],
        prepared["approver_option"],
        json.dumps(prepared["detail"], ensure_ascii=False),
        json.dumps(prepared["hashtags"], ensure_ascii=False),
        prepared["reference_note"],
        json.dumps(prepared["reference_links"], ensure_ascii=False),
        prepared["cost_limit_ref"],
        current_user_id(),
    ))
    db.commit()
    dnck_id = cursor.lastrowid
    row = db.execute("SELECT * FROM dnck WHERE id = ?", (dnck_id,)).fetchone()
    payload = build_dnck_qt82_payload(dnck_row_to_dict(row), get_settings())
    return _customer_lookup_json({"ok": True, "id": dnck_id, "qt82": payload})


def prepare_dnck_record(data):
    """Chuẩn hóa dữ liệu DNCK từ request để dùng chung cho tạo mới/cập nhật."""
    object_code = remove_all_whitespace(data.get("object_code", "")).upper()
    purpose = str(data.get("purpose") or "").strip()
    object_type = infer_dnck_object_type(object_code, purpose or data.get("object_type"))
    object_name = str(data.get("object_name") or "").strip()
    identity_value = remove_all_whitespace(data.get("identity_value", ""))
    phone = ""
    account_number = normalize_account_number(data.get("account_number", ""))
    account_name = str(data.get("account_name") or "").strip()
    bank = str(data.get("bank") or "").strip()
    amount = int(float(data.get("amount") or 0))
    sap_document = remove_all_whitespace(data.get("sap_document", ""))
    payment_tag = "Thanh Toán Khác"
    approval_level = str(data.get("approval_level") or "Cấp cửa hàng").strip()
    expense_type = str(data.get("expense_type") or data.get("cost_group") or "Khác").strip()
    cost_group = str(data.get("cost_group") or "Khác").strip()
    if cost_group not in ("Tiếp khách", "Công tác", "Khác", "Hàng hóa(ML)"):
        cost_group = "Khác"
    purpose = purpose or default_dnck_purpose(object_type)
    approver_option = str(data.get("approver_option") or "none").strip()
    if approver_option not in DNCK_APPROVER_OPTIONS:
        approver_option = "none"
    request_content = str(data.get("request_content") or "").strip()
    hashtags = normalize_dnck_hashtags(data.get("hashtags") or data.get("hashtags_text") or "")
    reference_note = str(data.get("reference_note") or "").strip()
    cost_limit_ref = str(data.get("cost_limit_ref") or "").strip()
    reference_links_raw = data.get("reference_links")
    if isinstance(reference_links_raw, list):
        reference_links = [
            str(item or "").strip()
            for item in reference_links_raw
            if str(item or "").strip()
        ][:20]
    else:
        reference_links = [
            line.strip()
            for line in str(reference_links_raw or "").splitlines()
            if line.strip()
        ][:20]
    detail = sanitize_dnck_detail(
        data.get("detail") if isinstance(data.get("detail"), list) else [],
        payment_tag,
        amount,
        sap_document,
    )
    if detail:
        amount = sum(int(item.get("amount") or 0) for item in detail)

    errors = []
    if not object_code:
        errors.append("Thiếu mã đối tượng.")
    if not object_name:
        errors.append("Thiếu tên đối tượng.")
    if not account_number:
        errors.append("Thiếu số tài khoản.")
    if not account_name:
        errors.append("Thiếu tên tài khoản.")
    if not bank:
        errors.append("Thiếu ngân hàng.")
    if amount == 0:
        errors.append("Số tiền phải khác 0.")
    if not request_content:
        errors.append("Thiếu nội dung đề nghị thanh toán.")
    if errors:
        return {"errors": errors}
    return {
        "errors": [],
        "object_code": object_code,
        "purpose": purpose,
        "object_type": object_type,
        "object_name": object_name,
        "identity_value": identity_value,
        "phone": phone,
        "account_number": account_number,
        "account_name": account_name,
        "bank": bank,
        "amount": amount,
        "sap_document": sap_document,
        "payment_tag": payment_tag,
        "approval_level": approval_level,
        "expense_type": expense_type,
        "cost_group": cost_group,
        "approver_option": approver_option,
        "request_content": request_content,
        "hashtags": hashtags,
        "reference_note": reference_note,
        "reference_links": reference_links,
        "cost_limit_ref": cost_limit_ref,
        "detail": detail,
    }


@app.route("/api/dnck/<int:dnck_id>", methods=["PUT"])
def api_dnck_update(dnck_id):
    """Cập nhật DNCK đã lưu, chỉ ADMIN."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if request.content_length is not None and request.content_length > 262144:
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu quá lớn."}, 413)
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    db = get_db()
    row = db.execute("SELECT id FROM dnck WHERE id = ?", (dnck_id,)).fetchone()
    if not row:
        return _customer_lookup_json({"ok": False, "error": "Không tìm thấy đề nghị CK."}, 404)
    prepared = prepare_dnck_record(request.get_json(silent=True) or {})
    if prepared["errors"]:
        return _customer_lookup_json({"ok": False, "error": " ".join(prepared["errors"])}, 400)
    db.execute("""
        UPDATE dnck
        SET object_type = ?, object_code = ?, object_name = ?, identity_value = ?, phone = ?,
            account_number = ?, account_name = ?, bank = ?, purpose = ?, approval_level = ?,
            expense_type = ?, cost_group = ?, request_content = ?, sap_document = ?,
            amount = ?, payment_tag = ?, approver_option = ?, detail_json = ?,
            hashtags_json = ?, reference_note = ?, reference_links_json = ?,
            cost_limit_ref = ?
        WHERE id = ?
    """, (
        prepared["object_type"],
        prepared["object_code"],
        prepared["object_name"],
        prepared["identity_value"],
        prepared["phone"],
        prepared["account_number"],
        prepared["account_name"],
        prepared["bank"],
        prepared["purpose"],
        prepared["approval_level"],
        prepared["expense_type"],
        prepared["cost_group"],
        prepared["request_content"],
        prepared["sap_document"],
        prepared["amount"],
        prepared["payment_tag"],
        prepared["approver_option"],
        json.dumps(prepared["detail"], ensure_ascii=False),
        json.dumps(prepared["hashtags"], ensure_ascii=False),
        prepared["reference_note"],
        json.dumps(prepared["reference_links"], ensure_ascii=False),
        prepared["cost_limit_ref"],
        dnck_id,
    ))
    db.commit()
    row = db.execute("SELECT * FROM dnck WHERE id = ?", (dnck_id,)).fetchone()
    payload = build_dnck_qt82_payload(dnck_row_to_dict(row), get_settings())
    return _customer_lookup_json({"ok": True, "id": dnck_id, "qt82": payload})


@app.route("/api/dnck/<int:dnck_id>/copy", methods=["POST"])
def api_dnck_copy(dnck_id):
    """Sao chép DNCK thành bản mới để chỉnh tiếp."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    db = get_db()
    row = db.execute("SELECT * FROM dnck WHERE id = ?", (dnck_id,)).fetchone()
    if not row:
        return _customer_lookup_json({"ok": False, "error": "Không tìm thấy đề nghị CK."}, 404)
    d = dnck_row_to_dict(row)
    cursor = db.execute("""
        INSERT INTO dnck
            (created_at, object_type, object_code, object_name, identity_value, phone,
             account_number, account_name, bank, purpose, approval_level, expense_type,
             cost_group, request_content, sap_document, amount, payment_tag, approver_option,
             detail_json, hashtags_json, reference_note, reference_links_json,
             cost_limit_ref, da_trinh, user_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?)
    """, (
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        d.get("object_type") or "customer",
        d.get("object_code", ""),
        d.get("object_name", ""),
        d.get("identity_value", ""),
        d.get("phone", ""),
        d.get("account_number", ""),
        d.get("account_name", ""),
        d.get("bank", ""),
        d.get("purpose", ""),
        d.get("approval_level", ""),
        d.get("expense_type", ""),
        d.get("cost_group", ""),
        d.get("request_content", ""),
        d.get("sap_document", ""),
        d.get("amount", 0),
        d.get("payment_tag", "Thanh Toán Khác"),
        d.get("approver_option", "none"),
        json.dumps(d.get("detail") or [], ensure_ascii=False),
        json.dumps(d.get("hashtags") or [], ensure_ascii=False),
        d.get("reference_note", ""),
        json.dumps(d.get("reference_links") or [], ensure_ascii=False),
        d.get("cost_limit_ref", ""),
        current_user_id(),
    ))
    db.commit()
    return _customer_lookup_json({"ok": True, "id": cursor.lastrowid})


@app.route("/api/dnck/<int:dnck_id>", methods=["DELETE"])
def api_dnck_delete(dnck_id):
    """Xóa DNCK, chỉ ADMIN."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    db = get_db()
    row = db.execute("SELECT id FROM dnck WHERE id = ?", (dnck_id,)).fetchone()
    if not row:
        return _customer_lookup_json({"ok": False, "error": "Không tìm thấy đề nghị CK."}, 404)
    db.execute("DELETE FROM dnck WHERE id = ?", (dnck_id,))
    db.commit()
    return _customer_lookup_json({"ok": True})


@app.route("/api/dnck/history")
def api_dnck_history():
    """Lịch sử DNCK riêng cho ADMIN."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    rows = get_db().execute("SELECT * FROM dnck ORDER BY id DESC LIMIT 200").fetchall()
    data = []
    for row in rows:
        d = dnck_row_to_dict(row)
        d.pop("detail_json", None)
        data.append(d)
    return _customer_lookup_json({"ok": True, "data": data, "admin": True})


@app.route("/api/dnck/object-lookup")
def api_dnck_object_lookup():
    """Tra thông tin đối tượng DNCK theo mã NV/KH/NCC; hiện seed demo cho mã NV."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    code = remove_all_whitespace(request.args.get("code", "")).upper()
    if not code:
        return _customer_lookup_json({"ok": False, "error": "Thiếu mã đối tượng."}, 400)
    rows = get_db().execute("""
        SELECT *
        FROM dnck_object_lookup
        WHERE UPPER(object_code) = ?
        ORDER BY is_primary DESC, id ASC
    """, (code,)).fetchall()
    if not rows:
        return _customer_lookup_json({"ok": True, "found": False, "data": None})
    accounts = []
    for row in rows:
        d = row_to_dict(row)
        accounts.append({
            "account_number": normalize_account_number(d.get("account_number", "")),
            "bank": d.get("bank", ""),
            "bank_eoffice_code": d.get("bank_eoffice_code", ""),
            "is_primary": bool(d.get("is_primary")),
            "source": d.get("source", ""),
        })
    primary = accounts[0]
    first = row_to_dict(rows[0])
    return _customer_lookup_json({
        "ok": True,
        "found": True,
        "data": {
            "object_code": first.get("object_code", ""),
            "object_name": first.get("object_name", ""),
            "identity_value": first.get("identity_value", ""),
            "account_number": primary["account_number"],
            "bank": primary["bank"],
            "bank_eoffice_code": primary["bank_eoffice_code"],
            "account_name": first.get("object_name", ""),
            "accounts": accounts,
        },
    })


def dnck_object_row_to_dict(row):
    d = row_to_dict(row)
    return {
        "id": d.get("id"),
        "object_code": d.get("object_code", ""),
        "object_name": d.get("object_name", ""),
        "account_number": normalize_account_number(d.get("account_number", "")),
        "bank": d.get("bank", ""),
        "identity_value": d.get("identity_value", ""),
        "bank_eoffice_code": d.get("bank_eoffice_code", ""),
        "is_primary": bool(d.get("is_primary")),
        "source": d.get("source", ""),
    }


@app.route("/api/dnck/objects")
def api_dnck_objects_list():
    """Danh sách dữ liệu đối tượng DNCK, chỉ ADMIN."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    q = str(request.args.get("q") or "").strip()
    db = get_db()
    if q:
        like = f"%{q.upper()}%"
        rows = db.execute("""
            SELECT *
            FROM dnck_object_lookup
            WHERE UPPER(object_code) LIKE ?
               OR UPPER(object_name) LIKE ?
               OR account_number LIKE ?
               OR identity_value LIKE ?
            ORDER BY object_code, is_primary DESC, id
            LIMIT 300
        """, (like, like, f"%{q}%", f"%{q}%")).fetchall()
    else:
        rows = db.execute("""
            SELECT *
            FROM dnck_object_lookup
            ORDER BY object_code, is_primary DESC, id
            LIMIT 300
        """).fetchall()
    return _customer_lookup_json({"ok": True, "data": [dnck_object_row_to_dict(row) for row in rows]})


@app.route("/api/dnck/objects", methods=["POST"])
def api_dnck_objects_save():
    """Thêm dữ liệu đối tượng DNCK."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    data = request.get_json(silent=True) or {}
    object_code = remove_all_whitespace(data.get("object_code", "")).upper()
    object_name = str(data.get("object_name") or "").strip().upper()
    account_number = normalize_account_number(data.get("account_number", ""))
    bank = str(data.get("bank") or "").strip()
    identity_value = remove_all_whitespace(data.get("identity_value", ""))
    bank_eoffice_code = remove_all_whitespace(data.get("bank_eoffice_code", "")) or find_eoffice_bank_code(bank)
    is_primary = 1 if data.get("is_primary", True) else 0
    errors = []
    if not object_code:
        errors.append("Thiếu mã đối tượng.")
    if not object_name:
        errors.append("Thiếu tên đối tượng.")
    if not account_number:
        errors.append("Thiếu số tài khoản.")
    if not bank:
        errors.append("Thiếu ngân hàng.")
    if errors:
        return _customer_lookup_json({"ok": False, "error": " ".join(errors)}, 400)
    db = get_db()
    if is_primary:
        db.execute("UPDATE dnck_object_lookup SET is_primary = 0 WHERE object_code = ?", (object_code,))
    cursor = db.execute("""
        INSERT INTO dnck_object_lookup
            (object_code, object_name, account_number, bank, identity_value,
             bank_eoffice_code, is_primary, source)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'manual')
        ON CONFLICT(object_code, account_number) DO UPDATE SET
            object_name = excluded.object_name,
            bank = excluded.bank,
            identity_value = excluded.identity_value,
            bank_eoffice_code = excluded.bank_eoffice_code,
            is_primary = excluded.is_primary,
            source = 'manual'
    """, (object_code, object_name, account_number, bank, identity_value, bank_eoffice_code, is_primary))
    db.commit()
    row = db.execute(
        "SELECT * FROM dnck_object_lookup WHERE object_code = ? AND account_number = ?",
        (object_code, account_number),
    ).fetchone()
    return _customer_lookup_json({"ok": True, "data": dnck_object_row_to_dict(row), "id": row["id"] if row else cursor.lastrowid})


@app.route("/api/dnck/objects/<int:object_id>", methods=["PUT"])
def api_dnck_objects_update(object_id):
    """Sửa một dòng dữ liệu đối tượng DNCK."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    data = request.get_json(silent=True) or {}
    db = get_db()
    row = db.execute("SELECT * FROM dnck_object_lookup WHERE id = ?", (object_id,)).fetchone()
    if not row:
        return _customer_lookup_json({"ok": False, "error": "Không tìm thấy dữ liệu đối tượng."}, 404)
    object_code = remove_all_whitespace(data.get("object_code", row["object_code"])).upper()
    object_name = str(data.get("object_name", row["object_name"]) or "").strip().upper()
    account_number = normalize_account_number(data.get("account_number", row["account_number"]))
    bank = str(data.get("bank", row["bank"]) or "").strip()
    identity_value = remove_all_whitespace(data.get("identity_value", row["identity_value"]))
    bank_eoffice_code = remove_all_whitespace(data.get("bank_eoffice_code", row["bank_eoffice_code"])) or find_eoffice_bank_code(bank)
    is_primary = 1 if data.get("is_primary", bool(row["is_primary"])) else 0
    if not object_code or not object_name or not account_number or not bank:
        return _customer_lookup_json({"ok": False, "error": "Thiếu mã, tên, STK hoặc ngân hàng."}, 400)
    if is_primary:
        db.execute("UPDATE dnck_object_lookup SET is_primary = 0 WHERE object_code = ? AND id <> ?", (object_code, object_id))
    db.execute("""
        UPDATE dnck_object_lookup
        SET object_code = ?, object_name = ?, account_number = ?, bank = ?,
            identity_value = ?, bank_eoffice_code = ?, is_primary = ?, source = 'manual'
        WHERE id = ?
    """, (object_code, object_name, account_number, bank, identity_value, bank_eoffice_code, is_primary, object_id))
    db.commit()
    row = db.execute("SELECT * FROM dnck_object_lookup WHERE id = ?", (object_id,)).fetchone()
    return _customer_lookup_json({"ok": True, "data": dnck_object_row_to_dict(row)})


@app.route("/api/dnck/objects/<int:object_id>", methods=["DELETE"])
def api_dnck_objects_delete(object_id):
    """Xóa một dòng dữ liệu đối tượng DNCK."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    db = get_db()
    row = db.execute("SELECT id FROM dnck_object_lookup WHERE id = ?", (object_id,)).fetchone()
    if not row:
        return _customer_lookup_json({"ok": False, "error": "Không tìm thấy dữ liệu đối tượng."}, 404)
    db.execute("DELETE FROM dnck_object_lookup WHERE id = ?", (object_id,))
    db.commit()
    return _customer_lookup_json({"ok": True})


@app.route("/api/dnck/da-trinh/<int:dnck_id>", methods=["POST"])
def api_dnck_da_trinh(dnck_id):
    """Cập nhật trạng thái đã trình cho DNCK."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if not request.is_json or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    value = 1 if (request.get_json(silent=True) or {}).get("da_trinh") else 0
    db = get_db()
    row = db.execute("SELECT id FROM dnck WHERE id = ?", (dnck_id,)).fetchone()
    if not row:
        return _customer_lookup_json({"ok": False, "error": "Không tìm thấy đề nghị CK."}, 404)
    db.execute("UPDATE dnck SET da_trinh = ? WHERE id = ?", (value, dnck_id))
    db.commit()
    return _customer_lookup_json({"ok": True})


@app.route("/api/phieu/<int:phieu_id>")
def api_get_phieu(phieu_id):
    """Get a single phieu as JSON."""
    db = get_db()
    row = get_accessible_phieu(db, phieu_id)
    if not row:
        return jsonify({"ok": False, "error": "Kh\u00f4ng t\u00ecm th\u1ea5y phi\u1ebfu"}), 404

    d = row_to_dict(row)
    d["tong_ck_chu"] = so_thanh_chu(d["tong_ck"])
    # Parse chung_tu_json back to list
    try:
        d["chung_tu"] = json.loads(d["chung_tu_json"]) if d["chung_tu_json"] else []
    except (json.JSONDecodeError, TypeError):
        d["chung_tu"] = []
    return jsonify({"ok": True, "phieu": d})


@app.route("/api/print/<int:phieu_id>")
def api_print(phieu_id):
    """Return printable HTML for a specific phieu."""
    db = get_db()
    row = get_accessible_phieu(db, phieu_id)
    if not row:
        return "Kh\u00f4ng t\u00ecm th\u1ea5y phi\u1ebfu", 404

    settings = get_settings()
    d = prepare_phieu_for_output(row, settings)
    d["pdf_token"] = create_pdf_token(phieu_id, row["user_id"])
    return render_template("print.html", p=d, staff=STAFF)


@app.route("/p/<token>")
def api_print_token(token):
    """Return printable HTML from a short-lived token so browser footer does not expose row IDs."""
    token_payload = verify_print_token()
    if not token_payload:
        return "Link in phiếu đã hết hạn hoặc không hợp lệ.", 404
    db = get_db()
    row = db.execute(
        "SELECT * FROM phieu WHERE id = ? AND user_id = ?",
        (int(token_payload["phieu_id"]), int(token_payload["user_id"])),
    ).fetchone()
    if not row:
        return "Không tìm thấy phiếu", 404

    settings = get_settings()
    d = prepare_phieu_for_output(row, settings)
    d["pdf_token"] = token
    response = app.make_response(render_template("print.html", p=d, staff=STAFF))
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response



@app.route("/api/pdf/<int:phieu_id>")
def api_pdf(phieu_id):
    """Download a PDF copy for in-app browsers where window.print() is blocked."""
    db = get_db()
    token_payload = verify_pdf_token(phieu_id)
    if token_payload:
        row = db.execute("SELECT * FROM phieu WHERE id = ? AND user_id = ?",
                         (phieu_id, int(token_payload["user_id"]))).fetchone()
    else:
        row = get_accessible_phieu(db, phieu_id)
    if not row:
        row = get_accessible_phieu(db, phieu_id) if is_admin() else None
    if not row:
        return "Kh\u00f4ng t\u00ecm th\u1ea5y phi\u1ebfu", 404

    p = prepare_phieu_for_output(row, get_settings())
    try:
        html = render_template("print.html", p=p, staff=STAFF)
        pdf = make_pdf_from_print_html(html)
    except RuntimeError:
        try:
            pdf = make_phieu_pdf(p)
        except RuntimeError:
            return "Server ch\u01b0a c\u00e0i th\u01b0 vi\u1ec7n xu\u1ea5t PDF. Vui l\u00f2ng b\u00e1o admin c\u00e0i reportlab.", 500

    filename = f"{p.get('file_title') or ('CK ' + ascii_filename_part(p.get('ten_kh'), 30) + ' ' + str(p.get('id')))}.pdf"
    return send_file(pdf, mimetype="application/pdf", as_attachment=True, download_name=filename)


def get_payment_planning_row(phieu_id):
    db = get_db()
    token_payload = verify_pdf_token(phieu_id)
    if token_payload:
        return db.execute(
            "SELECT * FROM phieu WHERE id = ? AND user_id = ?",
            (phieu_id, int(token_payload["user_id"])),
        ).fetchone()
    return get_accessible_phieu(db, phieu_id)


@app.route("/api/payment-planning/<int:phieu_id>")
def api_payment_planning_print(phieu_id):
    """Return printable Payment Planning HTML for a specific phieu."""
    row = get_payment_planning_row(phieu_id)
    if not row:
        return "Không tìm thấy phiếu", 404
    p = prepare_payment_planning_for_output(row, get_settings())
    p["pdf_token"] = request.args.get("token", "")
    return render_template("payment_planning_print.html", p=p, staff=STAFF)


@app.route("/api/payment-planning-pdf/<int:phieu_id>")
def api_payment_planning_pdf(phieu_id):
    """Download Payment Planning PDF rendered from the printable HTML template."""
    row = get_payment_planning_row(phieu_id)
    if not row:
        return "Không tìm thấy phiếu", 404
    p = prepare_payment_planning_for_output(row, get_settings())
    p["pdf_token"] = request.args.get("token", "")
    try:
        html = render_template("payment_planning_print.html", p=p, staff=STAFF)
        pdf = make_pdf_from_print_html(html)
    except RuntimeError:
        return "Server chưa cài thư viện xuất PDF. Vui lòng báo admin kiểm tra renderer.", 500
    filename = f"{p.get('planning_file_title')}.pdf"
    return send_file(pdf, mimetype="application/pdf", as_attachment=True, download_name=filename)


def make_payment_planning_xlsx(p):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    abbr = p.get("planning_abbr") or "PNJ"
    representative_abbr = p.get("planning_representative_abbr") or abbr
    definitions = p.get("planning_definition_rows") or PAYMENT_PLANNING_PROFILES["pnj"]["definition_rows"]
    has_working_day_definition = bool(p.get("planning_has_working_day_definition"))
    effectiveness_items = list(p.get("planning_effectiveness_items") or PAYMENT_PLANNING_PROFILES["pnj"]["effectiveness_items"])
    while len(effectiveness_items) < 4:
        effectiveness_items.append("")

    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Planning"
    ws.sheet_view.showGridLines = False

    widths = {"A": 20, "B": 42, "C": 18, "D": 14, "E": 24, "F": 12, "G": 14}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="FFF2CC")
    input_fill = PatternFill("solid", fgColor="FCE4D6")
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    bold = Font(name="Times New Roman", size=11, bold=True)
    normal = Font(name="Times New Roman", size=11)
    title_font = Font(name="Times New Roman", size=13, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    def merge(row, start_col, end_col, value, fill=None, font=None, align=None):
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=start_col, value=value)
        cell.fill = fill or PatternFill(fill_type=None)
        cell.font = font or normal
        cell.alignment = align or left
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = border

    def set_row(row, values, fills=None):
        fills = fills or {}
        values = list(values)
        values = (values + [""] * 5)[:5]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal
            cell.alignment = left
            cell.border = border
            if fills.get(col):
                cell.fill = fills[col]

    try:
        transaction_date = datetime.strptime(
            f"{p.get('nam')}-{p.get('thang')}-{p.get('ngay')}", "%Y-%m-%d"
        )
    except (TypeError, ValueError):
        transaction_date = None
    ws["G1"] = p.get("planning_bk_numbers") or ""
    ws["G2"] = transaction_date
    ws["G2"].number_format = "dd/mm/yyyy"
    ws.column_dimensions["F"].hidden = True
    ws.column_dimensions["G"].hidden = True

    merge(1, 1, 5, "PHỤ LỤC SỐ 01: THOẢ THUẬN THU ĐỔI SẢN PHẨM", title_fill, title_font, center)
    merge(2, 1, 5, '="Kèm theo Bảng kê mua lại tài sản số: "&G1&" ngày "&DAY(G2)&"/"&MONTH(G2)&"/"&YEAR(G2)', None, normal, center)

    section_rows = {
        3: "I. THÔNG TIN CÁC BÊN",
        10: "Thông tin tài khoản nhận tiền",
        13: "II. SẢN PHẨM THU ĐỔI, GIÁ TRỊ THU ĐỔI VÀ PHƯƠNG ÁN NHẬN GIÁ TRỊ",
        21: "III. ĐỊNH NGHĨA VÀ CÁCH XÁC ĐỊNH THỜI HẠN",
        26: "IV. KẾ HOẠCH THANH TOÁN",
        34: "V. HÌNH THỨC VÀ THÔNG TIN THANH TOÁN",
        37: "VI. NGUYÊN TẮC THANH TOÁN",
        43: "VII. XÁC NHẬN VÀ CAM KẾT CỦA KHÁCH HÀNG",
        48: f"VIII. CAM KẾT CỦA {abbr}",
        52: "IX. HIỆU LỰC, THỨ TỰ ƯU TIÊN VÀ GIẢI QUYẾT PHÁT SINH",
        57: "X. XÁC NHẬN CỦA CÁC BÊN",
    }
    for row, text in section_rows.items():
        merge(row, 1, 5, text, section_fill, bold, left)

    schedule = p.get("planning_schedule") or []
    schedule_dates = [
        schedule[index].get("date", "") if index < len(schedule) else ""
        for index in range(5)
    ]
    rows = {
        4: ["DOANH NGHIỆP", p.get("planning_company_name", ""), "", "KHÁCH HÀNG", p.get("ma_kh", "")],
        5: ["Địa chỉ", p.get("planning_pnj_address", ""), "", "Họ và tên", p.get("ten_kh", "")],
        6: ["Mã số doanh nghiệp", p.get("planning_tax_code", ""), "", "Số CCCD/Hộ chiếu", p.get("cccd", "")],
        7: ["Đại diện/Người tiếp nhận", p.get("planning_representative_name", ""), "", "Ngày cấp/Nơi cấp", ""],
        8: ["Cửa hàng/Đơn vị", p.get("planning_store_name", ""), "", "Địa chỉ liên hệ", p.get("dia_chi", "")],
        9: [f"Điện thoại/Email {abbr}", p.get("planning_pnj_contact", ""), "", "Điện thoại/Email", p.get("sdt", "")],
        11: ["Chủ tài khoản", p.get("ten_tk") or p.get("ten_kh", ""), "", "Số tài khoản", p.get("so_tk", "")],
        12: ["Ngân hàng", p.get("ngan_hang", ""), "", "", ""],
        14: ["Sản phẩm thu đổi", f'="Khách Hàng đồng ý cho {abbr} thu đổi sản phẩm với Thông tin chi tiết được xác định theo Bảng kê mua lại tài sản số: "&G1&" ngày "&DAY(G2)&"/"&MONTH(G2)&"/"&YEAR(G2)&" theo phương án lựa chọn được ghi nhận tại Thoả Thuận này."', "", "", ""],
        15: ["Tổng giá trị thu đổi (VNĐ)", p.get("total_trade", 0), "", "", ""],
        16: ["Bằng chữ", p.get("total_trade_words", ""), "", "", ""],
        17: ["Phương án lựa chọn", "☐ Phương án 1  ☐ Phương án 2  ☐ Phương án 3", "", "", ""],
        18: [f"Giá trị quy đổi sang sản phẩm {abbr} (VNĐ)", p.get("product_conversion", 0), "", "", ""],
        19: ["Giá trị nhận bằng tiền (VNĐ)", "=B15-B18", "", "", ""],
        20: ["Nguyên tắc cấn trừ", f"Việc sử dụng Tổng giá trị thu đổi được thực hiện theo phương án do Khách Hàng lựa chọn tại Thỏa thuận này. Theo đó:\n1. Phần Giá trị quy đổi sang sản phẩm {abbr} (nếu có) được cấn trừ trực tiếp vào giá mua sản phẩm {abbr} theo Hóa đơn bán hàng tương ứng đính kèm Thoả Thuận này;\n2. Phần Giá trị nhận bằng tiền (nếu có) được {abbr} thanh toán cho Khách Hàng theo Kế hoạch thanh toán quy định tại Mục IV của Thỏa thuận này;\n3. Trường hợp sau khi cấn trừ, Giá trị quy đổi sang sản phẩm {abbr} thấp hơn giá thanh toán của sản phẩm {abbr}, Khách Hàng có trách nhiệm thanh toán cho {abbr} phần chênh lệch còn thiếu.", "", "", ""],
        22: ["Ngày T", definitions.get("ngay_t", ""), "", "", ""],
        23: ["Ngày làm việc" if has_working_day_definition else "Cách tính T+n", definitions.get("ngay_lam_viec" if has_working_day_definition else "cach_tinh", ""), "", "", ""],
        24: ["Cách tính T+n" if has_working_day_definition else "Hoàn tất thanh toán", definitions.get("cach_tinh" if has_working_day_definition else "hoan_tat", ""), "", "", ""],
        25: ["Hoàn tất thanh toán" if has_working_day_definition else "", definitions.get("hoan_tat", "") if has_working_day_definition else "", "", "", ""],
        27: ["Đợt", "Thời điểm dự kiến", "Ngày dự kiến", "Tỷ lệ", "Số tiền (VNĐ)"],
        28: [1, "T/T+1", schedule_dates[0], 0.1, "=ROUND($B$19*D28,0)"],
        29: [2, "T+30", schedule_dates[1], 0.2, "=ROUND($B$19*D29,0)"],
        30: [3, "T+60", schedule_dates[2], 0.25, "=ROUND($B$19*D30,0)"],
        31: [4, "T+90", schedule_dates[3], 0.25, "=ROUND($B$19*D31,0)"],
        32: [5, "T+120", schedule_dates[4], 0.2, "=B19-SUM(E28:E31)"],
        33: ["Tổng", "", "", 1, "=SUM(E28:E32)"],
        35: ["Hình thức", "☐ Chuyển khoản    ☐ Khác", "", "", ""],
        36: ["Nội dung chuyển khoản", "", "", "", ""],
        38: [1, "Kế hoạch thanh toán và phương án nhận giá trị thu đổi được hai bên tự nguyện thỏa thuận trên cơ sở đã được cung cấp đầy đủ thông tin; không làm thay đổi tổng giá trị thu đổi đã xác nhận, trừ khi có thỏa thuận khác bằng văn bản.", "", "", ""],
        39: [2, f"{abbr} thực hiện thanh toán đúng số tiền, thời hạn và phương thức đã xác nhận tại Phụ lục này. Trường hợp {abbr} chậm thanh toán do lỗi của mình, trong trường hợp Khách Hàng yêu cầu lãi chậm thanh toán, {abbr} phải thanh toán cho Khách Hàng khoản lãi chậm thanh toán bằng 0,01%/ngày, tính trên số tiền chậm thanh toán và tương ứng với thời gian chậm thanh toán thực tế.", "", "", ""],
        40: [3, f"Khách Hàng chịu trách nhiệm kiểm tra và cung cấp chính xác thông tin tài khoản. Nếu thông tin sai hoặc tài khoản không hợp lệ, {abbr} thông báo để Khách Hàng điều chỉnh; thời hạn thanh toán được tính lại từ ngày {abbr} nhận đủ thông tin hợp lệ.\n{abbr} không chịu trách nhiệm đối với hậu quả phát sinh trực tiếp từ thông tin sai do Khách Hàng cung cấp.", "", "", ""],
        41: [4, "Mọi sửa đổi, bổ sung của Thoả Thuận này phải được lập thành văn bản.", "", "", ""],
        42: [5, "Các khoản phí do ngân hàng của Khách Hàng thu (nếu có) được thực hiện theo chính sách của ngân hàng, trừ khi hai bên có thỏa thuận khác bằng văn bản.", "", "", ""],
        44: [1, f"Khách Hàng xác nhận đã được {abbr} giải thích đầy đủ về phương án lựa chọn, giá trị thu đổi, nguyên tắc cấn trừ, kế hoạch thanh toán, phương thức thanh toán và các thông tin khác liên quan tới Thoả Thuận này trước khi ký kết.", "", "", ""],
        45: [2, "Khách Hàng tự nguyện lựa chọn phương án nêu tại Thoả Thuận này; đã đọc, hiểu rõ và nhận một bản Thoả Thuận sau khi ký.", "", "", ""],
        46: [3, f"Khách Hàng cam kết thông tin cá nhân, liên hệ và tài khoản nhận thanh toán cung cấp cho {abbr} là chính xác, hợp pháp.", "", "", ""],
        47: [4, "Bảo mật các thông tin liên quan tới Thoả Thuận này.", "", "", ""],
        49: [1, f"{abbr} cam kết bàn giao sản phẩm quy đổi của {abbr}, thanh toán đúng tổng giá trị nhận bằng tiền và kế hoạch đã xác nhận, đồng thời cung cấp chứng từ hoặc thông tin đối chiếu thanh toán nếu Khách Hàng yêu cầu.", "", "", ""],
        50: [2, f"{abbr} bảo mật và xử lý thông tin cá nhân, thông tin tài khoản của Khách Hàng đúng mục đích giao dịch, theo quy định pháp luật và chính sách bảo vệ dữ liệu cá nhân của {abbr}.", "", "", ""],
        51: [3, f"{abbr} bố trí đầu mối tiếp nhận yêu cầu tra soát, điều chỉnh thông tin phản ánh liên quan việc thực hiện kế hoạch thanh toán.", "", "", ""],
        53: [1, effectiveness_items[0], "", "", ""],
        54: [2, effectiveness_items[1], "", "", ""],
        55: [3, effectiveness_items[2], "", "", ""],
        56: [4, effectiveness_items[3], "", "", ""],
        58: [f"\nKhách Hàng\n(Ký, ghi rõ họ tên)\n\n\n{p.get('seller_signature_name', '')}", "", f"ĐẠI DIỆN {representative_abbr}/NGƯỜI ĐƯỢC ỦY QUYỀN\n{p.get('buyer_signature_title', '')}\n(Ký, ghi rõ họ tên, chức danh)\n\n\n{p.get('buyer_signature_name', '')}", "", ""],
        59: ['=" Ngày ký "&DAY(G2)&"/"&MONTH(G2)&"/"&YEAR(G2)', "", '=" Ngày ký "&DAY(G2)&"/"&MONTH(G2)&"/"&YEAR(G2)', "", ""],
    }

    merge_rows = {14, 16, 17, 20, 22, 23, 24, 25, 38, 39, 40, 41, 42, 44, 45, 46, 47, 49, 50, 51, 53, 54, 55, 56}
    for row in range(4, 60):
        if row in section_rows:
            continue
        values = rows.get(row, ["", "", "", "", ""])
        set_row(row, values, {2: input_fill if row in {7, 8, 12, 17, 35} else None, 5: input_fill if row in {7, 8} else None})
        if row in merge_rows:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
            for col in range(2, 6):
                ws.cell(row=row, column=col).border = border

    for row in (27, 33):
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = header_fill
            ws.cell(row=row, column=col).font = bold
            ws.cell(row=row, column=col).alignment = center

    for row in range(28, 33):
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2).alignment = center
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=4).alignment = center
        ws.cell(row=row, column=5).alignment = right

    for row in list(range(38, 43)) + list(range(44, 48)) + list(range(49, 52)) + list(range(53, 57)):
        ws.cell(row=row, column=1).alignment = center

    for row in (15, 18, 19, 28, 29, 30, 31, 32, 33):
        ws.cell(row=row, column=5 if row >= 28 else 2).number_format = '#,##0'
    for row in (28, 29, 30, 31, 32, 33):
        ws.cell(row=row, column=4).number_format = '0%'
    for row in (28, 29, 30, 31, 32):
        ws.cell(row=row, column=3).number_format = 'dd/mm/yyyy'
    for row in (58, 59):
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=1).font = bold
        ws.cell(row=row, column=3).font = bold
    ws.merge_cells("A58:B58")
    ws.merge_cells("C58:E58")
    ws.merge_cells("A59:B59")
    ws.merge_cells("C59:E59")

    for row in range(1, 60):
        ws.row_dimensions[row].height = 18
    for row in (14, 20, 22, 23, 24, 25, 38, 39, 40, 41, 42, 44, 45, 46, 47, 49, 50, 51, 53, 54, 55, 56):
        ws.row_dimensions[row].height = 42
    ws.row_dimensions[58].height = 104

    ws.print_area = "A1:E59"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/api/payment-planning-xlsx/<int:phieu_id>")
def api_payment_planning_xlsx(phieu_id):
    """Download editable Payment Planning Excel workbook."""
    row = get_payment_planning_row(phieu_id)
    if not row:
        return "Không tìm thấy phiếu", 404
    p = prepare_payment_planning_for_output(row, get_settings())
    output = make_payment_planning_xlsx(p)
    filename = f"{p.get('planning_file_title')}.xlsx"
    response = send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.route("/api/history")
def api_history():
    """JSON list of phieu, newest first. Admin can see all users."""
    db = get_db()
    admin = is_admin()
    if admin:
        rows = db.execute("SELECT * FROM phieu ORDER BY id DESC").fetchall()
        dnck_rows = db.execute("SELECT * FROM dnck ORDER BY id DESC").fetchall()
        try:
            users = {u["id"]: (u.get("full_name") or u.get("username") or f"User {u['id']}") for u in shared_auth.list_users()}
        except Exception:
            users = {}
    else:
        rows = db.execute("SELECT * FROM phieu WHERE user_id = ? ORDER BY id DESC",
                          (current_user_id(),)).fetchall()
        dnck_rows = []
        users = {}
    result = []
    for row in rows:
        d = row_to_dict(row)
        d["source"] = "phieu"
        d["tong_ck_chu"] = so_thanh_chu(d["tong_ck"])
        try:
            d["chung_tu"] = json.loads(d["chung_tu_json"]) if d["chung_tu_json"] else []
        except (json.JSONDecodeError, TypeError):
            d["chung_tu"] = []
        d["pdf_token"] = create_pdf_token(d["id"], d.get("user_id") or current_user_id())
        d["can_manage"] = int(d.get("user_id") or 0) == int(current_user_id())
        d["can_update_da_trinh"] = admin or d["can_manage"]
        if admin:
            d["owner_name"] = users.get(d.get("user_id"), f"User {d.get('user_id')}")
        result.append(d)
    if admin:
        for row in dnck_rows:
            d = dnck_row_to_dict(row)
            mapped = {
                **d,
                "source": "dnck",
                "ma_kh": d.get("object_code", ""),
                "ten_kh": d.get("object_name", ""),
                "sdt": "",
                "cccd": d.get("identity_value", ""),
                "so_tk": d.get("account_number", ""),
                "ten_tk": d.get("account_name", ""),
                "ngan_hang": d.get("bank", ""),
                "so_bk": "DNCK",
                "tvv_code": "",
                "tvv_name": d.get("purpose", ""),
                "tong_ck": d.get("amount", 0),
                "tong_ck_chu": so_thanh_chu(d.get("amount", 0)),
                "status": "printed",
                "can_manage": True,
                "can_update_da_trinh": True,
                "owner_name": users.get(d.get("user_id"), f"User {d.get('user_id')}"),
                "pdf_token": "",
            }
            result.append(mapped)
    result.sort(key=lambda item: (str(item.get("created_at") or ""), int(item.get("id") or 0)), reverse=True)
    return jsonify({"ok": True, "data": result, "admin": admin})


@app.route("/api/delete/<int:phieu_id>", methods=["DELETE"])
def api_delete(phieu_id):
    """Delete a phieu."""
    db = get_db()
    row = get_owned_phieu(db, phieu_id)
    if not row:
        return jsonify({"ok": False, "error": "Kh\u00f4ng t\u00ecm th\u1ea5y phi\u1ebfu"}), 404

    db.execute("DELETE FROM phieu WHERE id = ? AND user_id = ?", (phieu_id, current_user_id()))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/parse-sap", methods=["POST"])
def api_parse_sap():
    """Parse SAP paste text and return structured chung_tu + calculated totals."""
    data = request.get_json(force=True)
    raw = data.get("sap_text", "") or data.get("raw_text", "")
    records = parse_sap_paste(raw)
    tong = calc_tong_ck(records)
    return jsonify({
        "ok": True,
        "chung_tu": records,
        "tong_ck": tong,
        "tong_ck_chu": so_thanh_chu(tong),
    })


@app.route("/api/calc-ngay-tt", methods=["POST"])
def api_calc_ngay_tt():
    """Calculate payment date based on current time."""
    ngay_tt = calc_ngay_tt()
    return jsonify({"ok": True, "ngay_tt": ngay_tt})


@app.route("/api/qr-url", methods=["POST"])
def api_qr_url():
    """Generate VietQR URL from bank info."""
    data = request.get_json(force=True)
    url = build_qr_url(
        data.get("ngan_hang", ""),
        normalize_account_number(data.get("so_tk")),
        amount=data.get("amount"),
        memo=data.get("memo"),
    )
    return jsonify({"ok": True, "qr_url": url})


@app.route("/api/so-thanh-chu", methods=["POST"])
def api_so_thanh_chu():
    """Convert number to Vietnamese words."""
    data = request.get_json(force=True)
    n = data.get("number", 0)
    return jsonify({"ok": True, "text": so_thanh_chu(n)})


@app.route("/api/bank-bins")
def api_bank_bins():
    """Return the bank BIN mapping."""
    return jsonify({"ok": True, "banks": BANK_BINS})


@app.route("/api/lookup-account", methods=["POST"])
def api_lookup_account():
    """Tra cứu tên chủ tài khoản qua MBBank interbank lookup."""
    data = request.get_json(force=True)
    bin_code = data.get("bin", "")
    account = normalize_account_number(data.get("accountNumber"))
    if not bin_code or not account:
        return jsonify({"ok": False, "error": "Thiếu BIN hoặc số tài khoản"})

    settings = get_settings()
    mb_user = settings.get("mb_username", "")
    mb_pass = settings.get("mb_password", "")
    mb_account = settings.get("mb_account", "")
    if not mb_user or not mb_pass:
        return jsonify({"ok": False, "error": "Chưa cấu hình MBBank. Vào Cài đặt để nhập."})

    # Find MBBank bankId from BIN (loaded from Excel)
    mb_bank_code = MB_BANK_MAP.get(str(bin_code), "")

    try:
        import mbbank
        mb = mbbank.MBBank(username=mb_user, password=mb_pass)

        # Auto-detect debitAccount from MBBank balance API
        debit = mb_account
        if not debit:
            try:
                bal = mb.getBalance()
                if bal.acct_list:
                    debit = bal.acct_list[0].acctNo
            except Exception:
                debit = mb_user

        if not mb_bank_code:
            return jsonify({"ok": False, "error": "Không tìm thấy mã ngân hàng MBBank cho BIN " + str(bin_code)})

        result = mb.getAccountName(
            accountNo=account,
            bankCode=mb_bank_code,
            debitAccount=debit,
        )
        if result.benName:
            return jsonify({"ok": True, "accountName": result.benName})
        else:
            return jsonify({"ok": False, "error": "Không tìm thấy tài khoản"})
    except Exception as e:
        err_msg = str(e)
        if "GW200" in err_msg:
            return jsonify({"ok": False, "error": "Số tài khoản không hợp lệ"})
        return jsonify({"ok": False, "error": f"Lỗi: {err_msg}"})


@app.route("/api/ocr-bk", methods=["POST"])
def api_ocr_bk():
    """OCR ảnh bảng kê SAP → trích xuất Số BK, Mã KH, CCCD, SĐT."""
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if request.content_length is not None and request.content_length > 10 * 1024 * 1024:
        return _customer_lookup_json({"ok": False, "error": "Ảnh vượt quá 10 MB."}, 413)
    if not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)

    import asyncio
    from chrome_lens_py import LensAPI

    if "image" not in request.files:
        return jsonify({"ok": False, "error": "Không có ảnh"})

    img_bytes = request.files["image"].read()
    if not img_bytes:
        return jsonify({"ok": False, "error": "Ảnh trống"})

    # Save temp file
    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp.write(img_bytes)
    tmp.close()

    try:
        import asyncio
        from chrome_lens_py import LensAPI
        api = LensAPI()

        async def do_ocr():
            return await api.process_image(tmp.name, ocr_language="vi", output_format="full_text")

        result = asyncio.run(do_ocr())

        # Extract text from result
        if isinstance(result, str):
            text = result
        elif isinstance(result, dict):
            if "full_text" in result:
                text = result["full_text"]
            elif "word_data" in result:
                # Build text from word_data preserving separators
                parts = []
                for w in result["word_data"]:
                    parts.append(w.get("word", "") + w.get("separator", " "))
                text = "".join(parts)
            else:
                text = str(result)
        else:
            text = str(result)
    except Exception as e:
        return jsonify({"ok": False, "error": f"OCR lỗi: {str(e)}"})
    finally:
        os.unlink(tmp.name)

    if not text:
        return jsonify({"ok": False, "error": "Không đọc được text từ ảnh"})

    so_bk = ""
    ma_kh = ""
    cccd = ""
    sdt = ""
    ten_kh = ""

    # Lấy MST để loại trừ khi tìm SĐT
    mst_match = re.search(r"MST[:\s]*(\d{10,14})", text)
    mst = mst_match.group(1) if mst_match else ""

    # Pattern 1: Số BK (44xxxxxxxx) + Mã KH — dạng "4403701048_100256722" hoặc tách space
    m = re.search(r"\b(44\d{8})[_\s]+(\d{6,12})\b", text)
    if m:
        so_bk = m.group(1)
        ma_kh = m.group(2)
    else:
        m2 = re.search(r"\b(44\d{8})\b", text)
        if m2:
            so_bk = m2.group(1)

    # Pattern 2: CCCD — 12 số liền bắt đầu bằng 0
    cccd_matches = re.findall(r"\b(0\d{11})\b", text)
    if cccd_matches:
        cccd = cccd_matches[0]
    else:
        # CCCD bị tách 2 phần, có thể đảo thứ tự: "459 046194003" hoặc "046194003 459"
        # Tìm số 0xxxxx (9 số) gần số ngắn (3 số) → ghép = 12
        long_parts = re.findall(r"\b(0\d{8})\b", text)
        short_parts = re.findall(r"\b(\d{3})\b", text)
        for lp in long_parts:
            for sp in short_parts:
                combined = lp + sp
                if len(combined) == 12:
                    cccd = combined
                    break
            if cccd:
                break

    # Pattern 3: SĐT — 10 số bắt đầu 07/08/09/03/05, loại MST và CCCD
    sdt_matches = re.findall(r"\b(0[3-9]\d{8})\b", text)
    for s in sdt_matches:
        if s != mst[:10] and s != cccd[:10] and not s.startswith("030"):
            sdt = s
            break

    # Pattern 4: Tên KH — giữa "1" và các từ địa chỉ/số
    # OCR đọc ngang bảng → tên có thể bị đảo: "THƯ NGUYỄN MINH PHƯỜNG"
    # Tên thật: NGUYỄN MINH THƯ
    addr_words = {"PHƯỜNG", "PHONG", "TP", "THÀNH", "SỐ", "ĐƯỜNG",
                  "QUẬN", "HUYỆN", "XÃ", "THÔN", "BÀ", "TRIỆU", "HÓA",
                  "HUẾ", "HUỂ", "THUẬN", "ĐIỆN", "TỔ", "DÂN", "PHỐ",
                  "CHIẾC", "KHU", "TDP", "LÀNG", "NGÕ", "HẺM", "ĐÔNG",
                  "TÂY", "NAM", "BẮC", "KIỆT", "ĐƯỜNG"}
    ho_vn = {"NGUYỄN", "TRẦN", "LÊ", "PHẠM", "HUỲNH", "HOÀNG", "PHAN",
             "VŨ", "VÕ", "ĐẶNG", "BÙI", "ĐỖ", "HỒ", "NGỌC", "DƯƠNG",
             "LÝ", "CHÂU", "ĐOÀN", "TRỊNH", "ĐINH", "LƯU", "VƯƠNG",
             "LƯƠNG", "THẠCH", "TÔN", "MAI", "TỐNG", "LƯƠNG", "HÀ"}

    name_m = re.search(r"\b1\s+([A-ZÀ-Ỹ][A-ZÀ-Ỹa-zà-ỹ\s]+?)(?:\s+\d|,|\s+\d+/)", text)
    if name_m:
        raw_name = name_m.group(1).strip()
        words = raw_name.split()
        # Lọc bỏ từ địa chỉ
        name_words = []
        for w in words:
            if w.upper() in addr_words:
                break
            name_words.append(w)

        # Reorder: nếu từ đầu tiên KHÔNG phải họ VN → có thể bị đảo
        # VD: "THƯ NGUYỄN MINH" → tìm họ VN trong danh sách → đưa lên đầu
        if name_words and name_words[0].upper() not in ho_vn:
            for i, w in enumerate(name_words):
                if w.upper() in ho_vn:
                    # Đưa từ họ lên đầu, từ trước đó xuống cuối (tên)
                    ten_rieng = name_words[:i]
                    ho_dem = name_words[i:]
                    name_words = ho_dem + ten_rieng
                    break

        ten_kh = " ".join(name_words) if name_words else ""

    return jsonify({
        "ok": True,
        "so_bk": so_bk,
        "ma_kh": ma_kh,
        "cccd": cccd,
        "sdt": sdt,
        "ten_kh": ten_kh,
    })


def make_template_tt_response(detail_rows, identity_value, filename_key):
    """Generate filled eOffice QT82 template Excel from normalized detail rows."""
    from openpyxl import load_workbook

    template_path = os.path.join(app.static_folder, "template_tt.xlsx")
    wb = load_workbook(template_path)
    ws = wb["Sheet1"]
    identity = str(identity_value or "")

    for i, ct in enumerate(detail_rows):
        r = 5 + i  # Row 5 onwards
        ws.cell(row=r, column=1, value=i + 1)
        loai = ct.get("loai") or ct.get("label") or ""
        ws.cell(row=r, column=2, value=loai)
        gia_tri = ct.get("gia_tri", ct.get("amount", 0))
        if float(gia_tri or 0) < 0:
            ws.cell(row=r, column=3, value=int(gia_tri))
        elif loai in ("Hóa đơn", "Phải thu khác", "Thuế TNCN"):
            ws.cell(row=r, column=3, value=-abs(int(gia_tri)))
        else:
            ws.cell(row=r, column=3, value=abs(int(gia_tri)))
        ws.cell(row=r, column=4, value=str(ct.get("so_ct") or ct.get("document") or ""))
        ws.cell(row=r, column=5, value=str(ct.get("identity") or identity))
        ws.cell(row=r, column=6, value=str(ct.get("note") or ""))
        ws.cell(row=r, column=7, value=True)

    # Fill remaining STT rows
    for i in range(len(detail_rows), 30):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=7, value=True)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_key = re.sub(r"[^0-9A-Za-z._-]", "_", str(filename_key or "phieu"))[:80]
    filename = f"Template - TT {safe_key}.xlsx"
    response = send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        max_age=0,
    )
    response.headers["Cache-Control"] = "no-store, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@app.route("/api/template-tt/<int:phieu_id>")
def api_template_tt(phieu_id):
    """Generate filled eOffice QT82 template Excel for a phieu."""
    if not is_admin():
        return "Bạn không có quyền tải dữ liệu QT82.", 403
    db = get_db()
    row = get_accessible_phieu(db, phieu_id)
    if not row:
        return "Không tìm thấy phiếu", 404

    d = row_to_dict(row)
    d["source"] = "phieu"
    try:
        chung_tu = json.loads(d["chung_tu_json"]) if d["chung_tu_json"] else []
    except (json.JSONDecodeError, TypeError):
        chung_tu = []
    cccd = re.sub(r"\D", "", d.get("cccd", ""))
    return make_template_tt_response(chung_tu, cccd, d.get("so_bk") or "phieu")


@app.route("/api/dnck/template-tt/<int:dnck_id>")
def api_dnck_template_tt(dnck_id):
    """Generate filled eOffice QT82 template Excel for DNCK."""
    if not is_admin():
        return "Bạn không có quyền tải dữ liệu QT82.", 403
    row = get_db().execute("SELECT * FROM dnck WHERE id = ?", (dnck_id,)).fetchone()
    if not row:
        return "Không tìm thấy đề nghị CK", 404
    d = dnck_row_to_dict(row)
    detail = [
        {
            "loai": item["label"],
            "gia_tri": item["amount"],
            "so_ct": item["document"],
            "identity": item.get("identity") or d.get("identity_value", ""),
            "note": item.get("note", ""),
        }
        for item in sanitize_dnck_detail(
            d.get("detail") or [],
            d.get("payment_tag") or "Thanh Toán Khác",
            int(round(float(d.get("amount", 0) or 0))),
            d.get("sap_document") or "DNCK",
        )
    ]
    return make_template_tt_response(detail, d.get("identity_value", ""), f"DNCK_{d.get('id')}")


@app.route("/api/banks")
def api_banks():
    """Return full bank list for dropdown search."""
    data = []
    admin = is_admin()
    for bank in BANK_LIST:
        item = {key: value for key, value in bank.items() if admin or key != "eoffice"}
        data.append(item)
    return jsonify({"ok": True, "data": data})


@app.route("/api/tvv")
def api_tvv():
    """Return TVV list from database."""
    db = get_db()
    rows = db.execute("SELECT id, ma, ten FROM tvv ORDER BY ma").fetchall()
    data = [{"id": r["id"], "ma": r["ma"], "ten": r["ten"]} for r in rows]
    return jsonify({"ok": True, "data": data})


def _central_plant_payload(data):
    province = str(data.get("province") or "").strip()
    plant = str(data.get("plant") or "").strip()
    if not province or len(province) > 80:
        raise ValueError("Tỉnh không hợp lệ.")
    if not re.fullmatch(r"\d{4}", plant):
        raise ValueError("Mã Plant phải gồm đúng 4 chữ số.")
    return province, plant


def _central_plants_admin_guard(require_json=True):
    denied = _customer_lookup_admin_required()
    if denied:
        return denied
    if (require_json and not request.is_json) or not _customer_lookup_is_same_origin():
        return _customer_lookup_json({"ok": False, "error": "Yêu cầu không hợp lệ."}, 400)
    csrf = request.headers.get("X-CSRF-Token", "")
    expected_csrf = session.get("customer_import_csrf", "")
    if not csrf or not expected_csrf or not secrets.compare_digest(csrf, expected_csrf):
        return _customer_lookup_json({"ok": False, "error": "Phiên xác nhận không hợp lệ."}, 403)
    return None


@app.route("/api/central-plants")
def api_central_plants():
    db = get_db()
    rows = db.execute(
        "SELECT id, province, plant FROM central_plants ORDER BY province COLLATE NOCASE, plant"
    ).fetchall()
    return _customer_lookup_json({"ok": True, "data": [dict(row) for row in rows]})


@app.route("/api/central-plants", methods=["POST"])
def api_central_plants_add():
    denied = _central_plants_admin_guard()
    if denied:
        return denied
    try:
        province, plant = _central_plant_payload(request.get_json(silent=True) or {})
        db = get_db()
        db.execute("INSERT INTO central_plants (province, plant) VALUES (?, ?)", (province, plant))
        db.commit()
    except ValueError as exc:
        return _customer_lookup_json({"ok": False, "error": str(exc)}, 400)
    except sqlite3.IntegrityError:
        return _customer_lookup_json({"ok": False, "error": "Mã Plant đã tồn tại."}, 409)
    return _customer_lookup_json({"ok": True})


@app.route("/api/central-plants/<int:plant_id>", methods=["PUT"])
def api_central_plants_update(plant_id):
    denied = _central_plants_admin_guard()
    if denied:
        return denied
    try:
        province, plant = _central_plant_payload(request.get_json(silent=True) or {})
        db = get_db()
        result = db.execute(
            "UPDATE central_plants SET province = ?, plant = ? WHERE id = ?",
            (province, plant, plant_id),
        )
        if result.rowcount == 0:
            return _customer_lookup_json({"ok": False, "error": "Không tìm thấy Plant."}, 404)
        db.commit()
    except ValueError as exc:
        return _customer_lookup_json({"ok": False, "error": str(exc)}, 400)
    except sqlite3.IntegrityError:
        return _customer_lookup_json({"ok": False, "error": "Mã Plant đã tồn tại."}, 409)
    return _customer_lookup_json({"ok": True})


@app.route("/api/central-plants/<int:plant_id>", methods=["DELETE"])
def api_central_plants_delete(plant_id):
    denied = _central_plants_admin_guard(require_json=False)
    if denied:
        return denied
    db = get_db()
    result = db.execute("DELETE FROM central_plants WHERE id = ?", (plant_id,))
    if result.rowcount == 0:
        return _customer_lookup_json({"ok": False, "error": "Không tìm thấy Plant."}, 404)
    db.commit()
    return _customer_lookup_json({"ok": True})


@app.route("/api/tvv", methods=["POST"])
def api_tvv_add():
    """Add a new TVV (admin only)."""
    if not is_admin():
        return jsonify({"ok": False, "error": "Chỉ admin mới được thao tác"}), 403
    data = request.get_json(force=True)
    ma = data.get("ma", "").strip()
    ten = data.get("ten", "").strip().upper()
    if not ma or not ten:
        return jsonify({"ok": False, "error": "Thiếu mã hoặc tên TVV"})
    db = get_db()
    db.execute("INSERT INTO tvv (ma, ten) VALUES (?, ?)", (ma, ten))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/tvv/<int:tvv_id>", methods=["DELETE"])
def api_tvv_delete(tvv_id):
    """Delete a TVV (admin only)."""
    if not is_admin():
        return jsonify({"ok": False, "error": "Chỉ admin mới được thao tác"}), 403
    db = get_db()
    db.execute("DELETE FROM tvv WHERE id = ?", (tvv_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/lydo-huy")
def api_lydo_huy():
    """Return lý do hủy list from database."""
    db = get_db()
    rows = db.execute("SELECT id, noi_dung FROM lydo_huy ORDER BY id").fetchall()
    data = [{"id": r["id"], "noi_dung": r["noi_dung"]} for r in rows]
    return jsonify({"ok": True, "data": data})


@app.route("/api/lydo-huy", methods=["POST"])
def api_lydo_huy_add():
    """Add a new lý do hủy (admin only)."""
    if not is_admin():
        return jsonify({"ok": False, "error": "Chỉ admin mới được thao tác"}), 403
    data = request.get_json(force=True)
    noi_dung = data.get("noi_dung", "").strip()
    if not noi_dung:
        return jsonify({"ok": False, "error": "Thiếu nội dung lý do"})
    db = get_db()
    db.execute("INSERT INTO lydo_huy (noi_dung) VALUES (?)", (noi_dung,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/api/lydo-huy/<int:ld_id>", methods=["DELETE"])
def api_lydo_huy_delete(ld_id):
    """Delete a lý do hủy (admin only)."""
    if not is_admin():
        return jsonify({"ok": False, "error": "Chỉ admin mới được thao tác"}), 403
    db = get_db()
    db.execute("DELETE FROM lydo_huy WHERE id = ?", (ld_id,))
    db.commit()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

# Always init DB when module loads
init_db()

if __name__ == "__main__":
    # Open browser after a short delay (the server needs a moment to start)
    import threading
    threading.Timer(1.5, lambda: webbrowser.open(f"http://localhost:{PORT}")).start()
    app.run(host="127.0.0.1", port=PORT, debug=True, use_reloader=False)

# -*- coding: utf-8 -*-

import json
import io
import re
import tempfile
import unittest
import zipfile
from pathlib import Path

import app as app_module
from customer_lookup import CustomerLookupStore
from employee_lookup import EmployeeLookupStore
from openpyxl import load_workbook


class EofficeQt82Tests(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.original_db_path = app_module.DB_PATH
        self.original_store = app_module._customer_lookup_store
        self.original_employee_store = app_module._employee_lookup_store
        app_module.DB_PATH = str(self.root / "phieu.db")
        app_module.init_db()
        self.store = CustomerLookupStore(self.root / "lookup.db", bytes(range(32)))
        self.store.initialize()
        app_module._customer_lookup_store = self.store
        self.employee_store = EmployeeLookupStore(self.root / "employee.db", bytes(range(32)))
        self.employee_store.initialize()
        app_module._employee_lookup_store = self.employee_store
        app_module.app.config.update(TESTING=True)
        self.client = app_module.app.test_client()

    def tearDown(self):
        app_module.DB_PATH = self.original_db_path
        app_module._customer_lookup_store = self.original_store
        app_module._employee_lookup_store = self.original_employee_store
        self.temp_dir.cleanup()

    def login(self, role="admin", user_id=1):
        with self.client.session_transaction() as session:
            session["user_id"] = user_id
            session["user_name"] = "ADMIN TEST"
            session["role"] = role

    def create_phieu(self, doc_num="", amount=1500000):
        payload = {
            "status": "draft",
            "ngay_lap": "2026-07-16",
            "ma_kh": "100000000",
            "ten_kh": "KHACH HANG TEST",
            "sdt": "0900000000",
            "cccd": "012345678901",
            "so_tk": "123456789",
            "ten_tk": "KHACH HANG TEST",
            "ngan_hang": "OCB",
            "so_bk": "4403000001",
            "plant": "1305",
            "tong_ck": amount,
            "chung_tu": [
                {
                    "loai": "Bảng kê",
                    "so_ct": "4403000001",
                    "doc_num": doc_num,
                    "gia_tri": amount,
                    "gio": "16/07/2026 10:00",
                }
            ],
        }
        response = self.client.post("/api/save", json=payload)
        self.assertEqual(response.status_code, 200)
        return response.get_json()["id"]

    @staticmethod
    def payload_from_html(html):
        match = re.search(
            r'<script type="application/json" id="qt82DraftPayload">(.*?)</script>',
            html,
            flags=re.DOTALL,
        )
        if not match:
            raise AssertionError("Không tìm thấy payload QT82 trong HTML.")
        return json.loads(match.group(1))

    def test_eoffice_and_template_are_really_admin_only(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(doc_num="2500000001")

        self.login(role="user", user_id=2)
        self.assertEqual(self.client.get("/eoffice").status_code, 403)
        self.assertEqual(self.client.get(f"/eoffice/{phieu_id}").status_code, 403)
        self.assertEqual(self.client.get(f"/api/template-tt/{phieu_id}").status_code, 403)
        self.assertEqual(self.client.get("/api/qt82-extension").status_code, 403)
        self.assertNotIn("eOffice QT82", self.client.get("/").get_data(as_text=True))
        history_html = self.client.get("/history").get_data(as_text=True)
        self.assertNotIn('title="eOffice QT82"', history_html)
        self.assertNotIn('title="Tải template TT"', history_html)

    def test_admin_can_download_current_qt82_extension_zip(self):
        self.login(role="admin")
        response = self.client.get("/api/qt82-extension")
        current_manifest = json.loads(
            (app_module.QT82_EXTENSION_DIR / "manifest.json").read_text(encoding="utf-8")
        )
        current_version = current_manifest["version"]

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers.get("Cache-Control"), "no-store, max-age=0")
        self.assertEqual(response.headers.get("X-Content-Type-Options"), "nosniff")
        self.assertIn(
            f"PNJ-QT82-Draft-Helper-v{current_version}.zip",
            response.headers.get("Content-Disposition", ""),
        )

        with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
            names = set(archive.namelist())
            expected = {
                f"PNJ-QT82-Draft-Helper/{name}"
                for name in app_module.QT82_EXTENSION_FILES
            }
            self.assertEqual(names, expected)
            manifest = json.loads(
                archive.read("PNJ-QT82-Draft-Helper/manifest.json").decode("utf-8")
            )
            self.assertEqual(manifest["version"], current_version)

    def test_extension_uses_workflow_catalog_before_filling_qt82(self):
        extension_dir = app_module.QT82_EXTENSION_DIR
        background = (extension_dir / "background.js").read_text(encoding="utf-8")
        filler = (extension_dir / "eoffice-fill.js").read_text(encoding="utf-8")
        manifest = json.loads((extension_dir / "manifest.json").read_text(encoding="utf-8"))

        self.assertEqual(manifest["version"], "0.1.18")
        self.assertIn("/workflow/sitepages/createworkflow.aspx?rcid=8&rscid=0&wid=0", background)
        self.assertIn('message.draft.openMode === "deeplink" ? message.draft.formUrl : CREATE_WORKFLOW_URL', background)
        self.assertIn('url.pathname === "/dnck"', background)
        self.assertIn('"https://dangkhoa.io.vn/dnck*"', json.dumps(manifest))
        self.assertIn("function fillApprovers(draft)", filler)
        self.assertIn("if (!draft.skipStoreManager)", filler)
        self.assertIn('message.type === "DISABLE_HELPER"', background)
        self.assertIn('text === "qt82 quy trinh thanh toan"', filler)
        self.assertIn("sessionStorage.setItem(HUB_CLICK_NONCE_KEY, nonce)", filler)
        self.assertIn("function isQt82FormPage()", filler)
        self.assertIn("function redirectForDraft(draft)", filler)
        self.assertIn("function directHrefFromTarget(target)", filler)
        self.assertIn("function dispatchDoubleClick(target)", filler)
        self.assertIn("function endCurrentSession(panel, disableHelper)", filler)
        self.assertIn('clear.textContent = activeDraft ? "Xóa dữ liệu tạm" : "Xong/Đóng"', filler)
        self.assertIn('disable.textContent = "Tắt helper"', filler)
        self.assertIn("handleDraftOnCurrentPage(activeDraft, false)", filler)
        self.assertNotIn('renderStatus("Tiện ích đã chạy. Đang tìm bản nháp QT82...', filler)
        self.assertNotIn("Không tìm thấy bản nháp QT82. Quay lại website", filler)
        html = self.client.get(f"/eoffice/{self.create_phieu()}").get_data(as_text=True)
        self.assertIn("preparingQt82 = false;", html)
        self.assertIn("if (prepareButton && qt82Draft && qt82Draft.ready) prepareButton.disabled = false;", html)

    def test_eoffice_index_redirects_to_latest_existing_phieu(self):
        self.login(role="admin")
        older_id = self.create_phieu(doc_num="2500000001")
        latest_id = self.create_phieu(doc_num="2500000002")
        dnck = self.client.post(
            "/api/dnck",
            json={
                "object_code": "104850022",
                "object_name": "HEPCO",
                "identity_value": "3300123456",
                "account_number": "123456789",
                "account_name": "HEPCO",
                "bank": "OCB",
                "request_content": "1305_THANH TOÁN CHI PHÍ KHÁC",
                "sap_document": "2500000003",
                "amount": 100000,
                "detail": [{"label": "Chi phí", "amount": 100000}],
            },
            headers={"Origin": "http://localhost"},
        )
        dnck_id = dnck.get_json()["id"]

        response = self.client.get("/eoffice", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith(f"/eoffice/{latest_id}"))

        response = self.client.get("/eoffice?mode=dnck", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith(f"/eoffice/dnck/{dnck_id}"))

        self.client.delete(f"/api/delete/{latest_id}")
        response = self.client.get("/eoffice", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith(f"/eoffice/{older_id}"))

    def test_eoffice_dnck_mode_renders_history_shell_without_dnck_rows(self):
        self.login(role="admin")
        self.create_phieu(doc_num="2500000001")

        response = self.client.get("/eoffice?mode=dnck", follow_redirects=False)
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("CK Khác", html)
        self.assertIn("Lịch sử ĐNCK khác", html)
        self.assertIn("Chọn một ĐNCK khác từ lịch sử để chuẩn bị QT82", html)
        self.assertIn("Chưa có ĐNCK khác.", html)

    def test_missing_sap_document_uses_1234_without_falling_back_to_bk(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(doc_num="")
        response = self.client.get(f"/eoffice/{phieu_id}")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers.get("Cache-Control"), "no-store, max-age=0")
        payload = self.payload_from_html(response.get_data(as_text=True))

        self.assertEqual(payload["sapDocument"], "1234")
        self.assertTrue(payload["sapPlaceholder"])
        self.assertEqual(payload["desiredDateMode"], "browser_today")
        self.assertNotEqual(payload["sapDocument"], "4403000001")
        self.assertEqual(payload["bankQuery"], "79333001-")
        self.assertTrue(payload["ready"])

    def test_explicit_sap_document_and_qt82_defaults_are_preserved(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(doc_num="2500000001")
        payload = self.payload_from_html(
            self.client.get(f"/eoffice/{phieu_id}").get_data(as_text=True)
        )

        self.assertEqual(payload["sapDocument"], "2500000001")
        self.assertFalse(payload["sapPlaceholder"])
        self.assertEqual(payload["purpose"], "Thanh toán cho khách hàng(Mua lại)")
        self.assertEqual(payload["currency"], "VND")
        self.assertEqual(payload["managerApproval"], "Có")
        self.assertEqual(payload["paymentMethod"], "Bank transfer – Chuyển khoản")
        self.assertEqual(payload["costGroup"], "Hàng hóa(ML)")
        self.assertEqual(payload["storeManagerQuery"], "my.hth")
        self.assertEqual(payload["detailDocuments"], ["4403000001"])
        self.assertEqual(
            payload["requestContent"],
            "1305_CK BK 4403000001 ngày 2026-07-16 cho KHACH HANG TEST - 1.500.000 VND",
        )
        self.assertEqual(payload["formUrl"], app_module.DEFAULT_QT82_FORM_URL)

    def test_admin_can_override_sap_document_for_qt82(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(doc_num="")

        response = self.client.post(
            f"/api/phieu/{phieu_id}/sap-document",
            json={"sap_document": "2500000001"},
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["sap_document"], "2500000001")

        html = self.client.get(f"/eoffice/{phieu_id}").get_data(as_text=True)
        payload = self.payload_from_html(html)
        self.assertEqual(payload["sapDocument"], "2500000001")
        self.assertFalse(payload["sapPlaceholder"])
        self.assertIn('id="btnSaveSapDocument"', html)
        self.assertIn("tr.addEventListener('dblclick'", html)

    def test_preflight_check_is_rendered_below_customer_and_qt82_actions(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(doc_num="2500000001")
        html = self.client.get(f"/eoffice/{phieu_id}").get_data(as_text=True)

        self.assertIn("CK Bảng Kê", html)
        self.assertIn("CK Khác", html)
        self.assertNotIn("ĐNCK bảng kê", html)
        customer_card = html.index('<i class="bi bi-clipboard-data"></i>')
        prepare_button = html.index('id="btnPrepareQt82"')
        preflight_card = html.index('<strong>Kiểm tra trước khi tạo QT82</strong>')

        self.assertLess(customer_card, prepare_button)
        self.assertLess(prepare_button, preflight_card)
        self.assertIn('<span class="badge bg-primary">', html)

    def test_history_navigation_derives_reverse_proxy_prefix_from_browser_path(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(doc_num="2500000001")
        html = self.client.get(f"/eoffice/{phieu_id}").get_data(as_text=True)

        self.assertIn("window.location.pathname.match", html)
        self.assertIn("const historyApiUrl = appPath('/api/history');", html)
        self.assertIn("p.source === 'dnck'", html)
        self.assertIn("appPath('/eoffice/dnck/' + p.id)", html)
        self.assertIn("appPath('/api/dnck/da-trinh/' + e.target.dataset.id)", html)
        self.assertIn("appPath('/api/da-trinh/' + e.target.dataset.id)", html)
        self.assertNotIn("link.href = '/eoffice/'", html)
        self.assertNotIn("const historyApiUrl = \"/api/history\";", html)
        self.assertIn("Không tải được lịch sử phiếu", html)

    def test_qt82_form_url_can_change_only_within_pnj_eoffice_workflow(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(doc_num="2500000001")
        new_url = (
            "https://eoffice.pnj.com.vn/workflow/SitePages/NewWorkflow.aspx"
            "?mode=1&LID=NEW-LIST-ID&wid=9999"
        )
        saved = self.client.post(
            "/api/settings",
            json={"qt82_form_url": new_url},
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(saved.status_code, 200)
        payload = self.payload_from_html(
            self.client.get(f"/eoffice/{phieu_id}").get_data(as_text=True)
        )
        self.assertEqual(payload["formUrl"], new_url)

        rejected = self.client.post(
            "/api/settings",
            json={"qt82_form_url": "https://example.com/workflow/fake"},
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(rejected.status_code, 400)

    def test_template_tt_is_xlsx_no_store_and_contains_expected_detail(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(doc_num="2500000001")

        response = self.client.get(f"/api/template-tt/{phieu_id}")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers.get("Cache-Control"), "no-store, max-age=0")
        self.assertEqual(response.headers.get("Pragma"), "no-cache")
        self.assertEqual(response.headers.get("X-Content-Type-Options"), "nosniff")
        self.assertTrue(response.data.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(response.data), data_only=False)
        sheet = workbook["Sheet1"]
        self.assertEqual(sheet.cell(row=5, column=2).value, "Bảng kê")
        self.assertEqual(sheet.cell(row=5, column=3).value, 1500000)
        self.assertEqual(sheet.cell(row=5, column=4).value, "4403000001")
        self.assertEqual(sheet.cell(row=5, column=5).value, "012345678901")

    def test_payment_planning_output_and_bk_save_copy_flow(self):
        self.login(role="admin")
        payload = {
            "status": "draft",
            "ngay_lap": "2026-07-21",
            "ma_kh": "100000001",
            "ten_kh": "NGUYEN VAN TEST",
            "dia_chi": "18A TRAN BINH TRONG, HUE",
            "sdt": "0900000001",
            "cccd": "012345678902",
            "so_tk": "123456789",
            "ten_tk": "NGUYEN VAN TEST",
            "ngan_hang": "OCB",
            "so_bk": "4403000002",
            "tvv_code": "11358",
            "tvv_name": "NGUYEN TVV",
            "plant": "1305",
            "tong_ck": 6000000,
            "nguoi_ki": "cht",
            "chung_tu": [
                {"loai": "Bảng kê", "so_ct": "4403000002", "bk_ref": "000551/07_1305", "gia_tri": 10000000, "gio": "21/07/2026 10:00"},
                {"loai": "Bảng kê", "so_ct": "4403000003", "bk_ref": "000552/07_1305", "gia_tri": 2000000, "gio": "21/07/2026 10:01"},
                {"loai": "Hóa đơn", "so_ct": "1400000001", "gia_tri": 4000000, "gio": "21/07/2026 10:05"},
            ],
        }
        created = self.client.post("/api/save", json=payload).get_json()
        phieu_id = created["id"]

        updated = self.client.post("/api/save", json={**payload, "phieu_id": phieu_id, "ten_kh": "TEN DA SUA"}).get_json()
        self.assertTrue(updated["updated"])
        self.assertEqual(updated["id"], phieu_id)
        self.assertEqual(self.client.get(f"/api/phieu/{phieu_id}").get_json()["phieu"]["ten_kh"], "TEN DA SUA")

        copied = self.client.post("/api/save", json={**payload, "phieu_id": phieu_id, "force_create": True}).get_json()
        self.assertNotEqual(copied["id"], phieu_id)

        html = self.client.get(f"/api/payment-planning/{phieu_id}").get_data(as_text=True)
        self.assertIn("Thoả thuận thu đổi sản phẩm", html)
        self.assertIn("Tải thoả thuận PDF", html)
        self.assertIn("Tải thoả thuận Excel", html)
        self.assertIn("0,01%/ngày", html)
        self.assertIn("000551/07_1305, 000552/07_1305", html)
        self.assertNotIn("000551/07_1305,000552/07_1305", html)
        self.assertNotIn("4403000002, 4403000003", html)
        self.assertIn("ngày 21 / 07 / 2026", html)
        self.assertIn("0300521758", html)
        self.assertNotIn("0300521758-023", html)
        self.assertIn("Thành phố Hồ Chí Minh, Việt Nam", html)
        self.assertIn("TTKH 27 Hà Nội,Huế", html)
        self.assertIn("18A TRAN BINH TRONG, HUE", html)
        self.assertIn("+ 84 (028) 39951703", html)
        self.assertIn("0900000001", html)
        self.assertNotIn("Chi tiết Phương án lựa chọn", html)
        self.assertIn("Khách Hàng", html)
        self.assertIn("ĐẠI DIỆN PNJ/NGƯỜI ĐƯỢC ỦY QUYỀN", html)
        self.assertIn('<div class="section-title">I. Thông tin các bên</div>', html)
        self.assertIn('<div class="section-title">IV. Kế hoạch thanh toán</div>', html)
        self.assertIn('<div class="page-one">', html)
        self.assertNotIn('class="section">I. Thông tin các bên', html)
        self.assertIn("Cửa Hàng Trưởng", html)
        self.assertIn('class="signature-block"', html)
        self.assertIn("signature-role-customer", html)
        self.assertIn('class="role-title role-title-pnj"', html)
        self.assertIn("<td>22/07/2026</td>", html)
        self.assertIn("<td>20/08/2026</td>", html)
        self.assertIn("Ngày ký: 21 / 07 / 2026", html)

        xlsx_response = self.client.get(f"/api/payment-planning-xlsx/{phieu_id}")
        self.assertEqual(xlsx_response.status_code, 200)
        self.assertIn("Thoa thuan thu doi", xlsx_response.headers.get("Content-Disposition", ""))
        workbook = load_workbook(io.BytesIO(xlsx_response.data), data_only=False)
        sheet = workbook["Payment Planning"]
        self.assertEqual(sheet["A1"].value, "PHỤ LỤC SỐ 01: THOẢ THUẬN THU ĐỔI SẢN PHẨM")
        self.assertIn("Bảng kê mua lại tài sản", sheet["A2"].value)
        self.assertEqual(sheet["G1"].value, "000551/07_1305, 000552/07_1305")
        self.assertEqual(sheet["B6"].value, "0300521758")
        self.assertIn("Việt Nam", sheet["B5"].value)
        self.assertEqual(sheet["B8"].value, "TTKH 27 Hà Nội,Huế")
        self.assertEqual(sheet["E8"].value, "18A TRAN BINH TRONG, HUE")
        self.assertEqual(sheet["B9"].value, "+ 84 (028) 39951703")
        self.assertEqual(sheet["E9"].value, "0900000001")
        self.assertNotIn("Chi tiết Phương án lựa chọn", sheet["B17"].value)
        self.assertEqual(sheet["B15"].value, 12000000)
        self.assertEqual(sheet["B18"].value, 6000000)
        self.assertEqual(sheet["B19"].value, "=B15-B18")
        self.assertEqual(sheet["C28"].value, "22/07/2026")
        self.assertEqual(sheet["C29"].value, "20/08/2026")
        self.assertEqual(sheet["C32"].value, "18/11/2026")
        self.assertIn("0,01%/ngày", sheet["B39"].value)
        self.assertIn("Khách Hàng", sheet["A58"].value)
        self.assertIn("ĐẠI DIỆN PNJ", sheet["C58"].value)
        self.assertIn("Cửa Hàng Trưởng", sheet["C58"].value)
        self.assertIn("DAY(G2)", sheet["A59"].value)

        manual_payload = {**payload, "force_create": True, "use_bk_ref": True, "so_bk": "4403000999"}
        manual = self.client.post("/api/save", json=manual_payload).get_json()
        manual_id = manual["id"]
        self.assertEqual(self.client.get(f"/api/phieu/{manual_id}").get_json()["phieu"]["use_bk_ref"], 1)

        manual_html = self.client.get(f"/api/payment-planning/{manual_id}").get_data(as_text=True)
        self.assertIn("000551/07_1305, 000552/07_1305", manual_html)
        self.assertNotIn("4403000002, 4403000003", manual_html)

        manual_xlsx_response = self.client.get(f"/api/payment-planning-xlsx/{manual_id}")
        self.assertEqual(manual_xlsx_response.status_code, 200)
        manual_workbook = load_workbook(io.BytesIO(manual_xlsx_response.data), data_only=False)
        manual_sheet = manual_workbook["Payment Planning"]
        self.assertEqual(manual_sheet["G1"].value, "000551/07_1305, 000552/07_1305")

        dated_payload = {**payload, "force_create": True, "show_payment_dates": True, "so_bk": "4403000888"}
        dated = self.client.post("/api/save", json=dated_payload).get_json()
        dated_id = dated["id"]
        self.assertEqual(self.client.get(f"/api/phieu/{dated_id}").get_json()["phieu"]["show_payment_dates"], 1)

        dated_html = self.client.get(f"/api/payment-planning/{dated_id}").get_data(as_text=True)
        for expected_date in ("22/07/2026", "20/08/2026", "19/09/2026", "19/10/2026", "18/11/2026"):
            self.assertIn(f"<td>{expected_date}</td>", dated_html)

        dated_xlsx_response = self.client.get(f"/api/payment-planning-xlsx/{dated_id}")
        self.assertEqual(dated_xlsx_response.status_code, 200)
        dated_workbook = load_workbook(io.BytesIO(dated_xlsx_response.data), data_only=False)
        dated_sheet = dated_workbook["Payment Planning"]
        self.assertEqual(dated_sheet["C28"].value, "22/07/2026")
        self.assertEqual(dated_sheet["C29"].value, "20/08/2026")
        self.assertEqual(dated_sheet["C32"].value, "18/11/2026")

    def test_bk_create_and_history_have_payment_planning_actions(self):
        self.login(role="admin")
        index_html = self.client.get("/").get_data(as_text=True)
        self.assertIn('data-store-mode="pnj"', index_html)
        self.assertIn('data-store-mode="cao"', index_html)
        self.assertIn('id="plant"', index_html)
        self.assertIn("plant: document.getElementById('plant').value", index_html)
        self.assertIn('id="btnCopyPhieu"', index_html)
        self.assertNotIn('id="useBkRef"', index_html)
        self.assertNotIn('id="showPaymentDates"', index_html)
        self.assertIn('id="sapTable"', index_html)
        self.assertIn("sap-bk-ref-cell", index_html)
        self.assertIn("use_bk_ref: useBkRefEnabled()", index_html)
        self.assertIn("show_payment_dates: showPaymentDatesEnabled()", index_html)
        self.assertIn("use_bk_ref_default", index_html)
        self.assertIn("show_payment_dates_default", index_html)
        self.assertIn("sap-bk-ref-status", index_html)
        self.assertIn("&lt;Đang tải số BK&gt;", index_html)
        self.assertIn("buildPurchaseOrderReferenceMapping", index_html)
        self.assertIn("normalizePurchaseOrderNumber", index_html)
        self.assertIn("Promise.all", index_html)
        self.assertIn("autoFillTransactionSuggestions", index_html)
        self.assertIn("auto_fill_transactions_default", index_html)
        self.assertIn("transactionLoadingStatus", index_html)
        self.assertIn("transactionAutoLoadChoiceModal", index_html)
        self.assertIn("handleTransactionManualIntent", index_html)
        self.assertIn("<Đang tải địa chỉ>", index_html)
        self.assertIn("Dữ liệu Bảng kê và Hoá đơn đang được tự động tải", index_html)
        self.assertIn("classList.toggle('use-bk-ref'", index_html)
        self.assertIn("download-planning-pdf", index_html)
        self.assertIn("download-planning-xlsx", index_html)
        self.assertIn("print-planning", index_html)
        self.assertIn("Thoả thuận thu đổi (pdf)", index_html)
        self.assertIn("Thoả thuận thu đổi (excel)", index_html)
        self.assertIn("Phiếu xác nhận CK (pdf)", index_html)
        self.assertNotIn("Phiếu XNTTTTCK (pdf)", index_html)
        self.assertIn("Tải thoả thuận PDF", index_html)
        self.assertIn("In thoả thuận", index_html)
        self.assertIn("attachCurrentPhieuId", index_html)
        self.assertIn("'Hóa đơn': '901'", index_html)
        self.assertIn("'Biên nhận cọc': '16'", index_html)
        self.assertIn("'HBTL': '990'", index_html)

        history_html = self.client.get("/history").get_data(as_text=True)
        self.assertIn("api/payment-planning-pdf/", history_html)
        self.assertIn("api/payment-planning-xlsx/", history_html)
        self.assertIn("Thoả thuận thu đổi (pdf)", history_html)
        self.assertIn("Thoả thuận thu đổi (excel)", history_html)
        self.assertIn("Phiếu xác nhận CK (pdf)", history_html)
        self.assertNotIn("Phiếu XNTTTTCK (pdf)", history_html)
        self.assertIn("Template TT (excel)", history_html)
        settings_html = self.client.get("/settings").get_data(as_text=True)
        self.assertIn("Tiền tố chứng từ", settings_html)
        self.assertIn('id="s_invoice_prefix"', settings_html)
        self.assertIn('id="s_deposit_prefix"', settings_html)
        self.assertIn('id="s_hbtl_prefix"', settings_html)
        self.assertIn('id="s_use_bk_ref_default"', settings_html)
        self.assertIn('id="s_show_payment_dates_default"', settings_html)
        self.assertIn('id="s_auto_fill_transactions_default"', settings_html)
        saved_settings = self.client.post(
            "/api/settings",
            json={
                "auto_fill_transactions_default": "1",
                "use_bk_ref_default": "0",
                "show_payment_dates_default": "1",
            },
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(saved_settings.status_code, 200)
        settings_json = self.client.get("/api/settings").get_json()["data"]
        self.assertEqual(settings_json["auto_fill_transactions_default"], "1")
        self.assertEqual(settings_json["use_bk_ref_default"], "0")
        self.assertEqual(settings_json["show_payment_dates_default"], "1")

    def test_cao_mode_uses_plant_2122_and_caf_payment_planning_text(self):
        self.login(role="admin")
        payload = {
            "status": "printed",
            "ngay_lap": "2026-07-22",
            "ma_kh": "100000003",
            "ten_kh": "KHACH HANG CAO",
            "dia_chi": "18A TRAN BINH TRONG, HUE",
            "sdt": "0900000003",
            "cccd": "012345678903",
            "so_tk": "123456789",
            "ten_tk": "KHACH HANG CAO",
            "ngan_hang": "OCB",
            "so_bk": "4403000003",
            "tvv_code": "11358",
            "tvv_name": "NGUYEN TVV",
            "plant": "2122",
            "tong_ck": 6000000,
            "nguoi_ki": "tvv",
            "chung_tu": [
                {"loai": "Bảng kê", "so_ct": "4403000003", "gia_tri": 10000000, "gio": "22/07/2026 16:20"},
                {"loai": "Hóa đơn", "so_ct": "9010000001", "gia_tri": 4000000, "gio": "22/07/2026 16:21"},
            ],
        }
        created = self.client.post("/api/save", json=payload).get_json()
        phieu_id = created["id"]

        phieu_json = self.client.get(f"/api/phieu/{phieu_id}").get_json()["phieu"]
        self.assertEqual(phieu_json["plant"], "2122")

        print_html = self.client.get(f"/api/print/{phieu_id}").get_data(as_text=True)
        self.assertIn("CH PNJ NEXT 27 Hà Nội, Huế - 2122_16:20_22/07/2026", print_html)

        planning_html = self.client.get(f"/api/payment-planning/{phieu_id}").get_data(as_text=True)
        self.assertIn("CÔNG TY TRÁCH NHIỆM HỮU HẠN MỘT THÀNH VIÊN THỜI TRANG CAO (CAF)", planning_html)
        self.assertIn("0309279212", planning_html)
        self.assertIn("<td>HỒ THỊ HÀ MY</td>", planning_html)
        self.assertIn("TTKH 27 Hà Nội,Huế", planning_html)
        self.assertIn("+ 84 (028) 39951703", planning_html)
        self.assertIn("Khách Hàng đồng ý cho CAF thu đổi sản phẩm", planning_html)
        self.assertIn("Giá trị quy đổi sang sản phẩm CAF", planning_html)
        self.assertIn("ĐẠI DIỆN CAF/NGƯỜI ĐƯỢC ỦY QUYỀN", planning_html)
        self.assertIn("Cửa Hàng Trưởng", planning_html)
        self.assertNotIn("NGUYEN TVV", planning_html)
        self.assertIn("Thỏa Thuận này chấm dứt", planning_html)
        self.assertNotIn("Ngày làm việc</td>", planning_html)
        self.assertNotIn("ngày làm việc thứ n", planning_html)

        xlsx_response = self.client.get(f"/api/payment-planning-xlsx/{phieu_id}")
        self.assertEqual(xlsx_response.status_code, 200)
        workbook = load_workbook(io.BytesIO(xlsx_response.data), data_only=False)
        sheet = workbook["Payment Planning"]
        self.assertEqual(sheet["B4"].value, "CÔNG TY TRÁCH NHIỆM HỮU HẠN MỘT THÀNH VIÊN THỜI TRANG CAO (CAF)")
        self.assertEqual(sheet["B6"].value, "0309279212")
        self.assertEqual(sheet["B7"].value, "HỒ THỊ HÀ MY")
        self.assertEqual(sheet["B8"].value, "TTKH 27 Hà Nội,Huế")
        self.assertEqual(sheet["A9"].value, "Điện thoại/Email CAF")
        self.assertEqual(sheet["B9"].value, "+ 84 (028) 39951703")
        self.assertIn("CAF thu đổi sản phẩm", sheet["B14"].value)
        self.assertIn("Cửa Hàng Trưởng", sheet["C58"].value)
        self.assertIn("ĐẠI DIỆN CAF/NGƯỜI ĐƯỢC ỦY QUYỀN", sheet["C58"].value)

    def test_bank_eoffice_code_is_admin_only_on_create_forms(self):
        self.login(role="user", user_id=2)
        banks = self.client.get("/api/banks").get_json()["data"]
        self.assertTrue(banks)
        self.assertNotIn("eoffice", banks[0])

        self.login(role="admin")
        banks = self.client.get("/api/banks").get_json()["data"]
        self.assertIn("eoffice", banks[0])
        index_html = self.client.get("/").get_data(as_text=True)
        dnck_html = self.client.get("/dnck").get_data(as_text=True)
        self.assertIn("bankEofficeCode", index_html)
        self.assertIn("bankEofficeCode", dnck_html)

    def test_dnck_object_lookup_fills_employee_demo_data(self):
        self.login(role="user", user_id=2)
        self.assertEqual(self.client.get("/api/dnck/object-lookup?code=E0124764").status_code, 403)

        self.login(role="admin")
        response = self.client.get("/api/dnck/object-lookup?code=E0124764")
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertTrue(data["ok"])
        self.assertTrue(data["found"])
        profile = data["data"]
        self.assertEqual(profile["object_code"], "E0124764")
        self.assertEqual(profile["object_name"], "CHÂU ĐĂNG KHOA")
        self.assertEqual(profile["account_number"], "109876756206")
        self.assertEqual(profile["bank"], "Vietinbank")
        self.assertEqual(profile["identity_value"], "046093004708")
        self.assertTrue(profile["bank_eoffice_code"])
        self.assertEqual(profile["account_name"], "CHÂU ĐĂNG KHOA")

        html = self.client.get("/dnck").get_data(as_text=True)
        self.assertIn("/api/dnck/object-lookup?code=", html)
        self.assertIn("objectLookupMessage", html)
        self.assertIn("selectBankFromLookup", html)

    def test_dnck_object_data_crud_api_and_modal_shell(self):
        self.login(role="user", user_id=2)
        self.assertEqual(self.client.get("/api/dnck/objects").status_code, 403)

        self.login(role="admin")
        html = self.client.get("/dnck").get_data(as_text=True)
        self.assertIn("Dữ liệu đối tượng", html)
        self.assertIn("objectDataModal", html)
        self.assertIn("/api/dnck/objects", html)

        create = self.client.post(
            "/api/dnck/objects",
            json={
                "object_code": "E0199999",
                "object_name": "NHAN VIEN CRUD",
                "account_number": "123000999",
                "bank": "Vietinbank",
                "identity_value": "046000000999",
                "bank_eoffice_code": "01201001-",
                "is_primary": True,
            },
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(create.status_code, 200)
        object_id = create.get_json()["id"]

        listing = self.client.get("/api/dnck/objects?q=E0199999").get_json()["data"]
        self.assertEqual(len(listing), 1)
        self.assertEqual(listing[0]["object_name"], "NHAN VIEN CRUD")
        self.assertTrue(listing[0]["is_primary"])

        update = self.client.put(
            f"/api/dnck/objects/{object_id}",
            json={
                "object_code": "E0199999",
                "object_name": "NHAN VIEN CRUD SUA",
                "account_number": "123000999",
                "bank": "Vietinbank",
                "identity_value": "046000000999",
                "bank_eoffice_code": "01201001-",
                "is_primary": True,
            },
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(update.status_code, 200)
        self.assertEqual(update.get_json()["data"]["object_name"], "NHAN VIEN CRUD SUA")

        delete = self.client.delete(f"/api/dnck/objects/{object_id}", headers={"Origin": "http://localhost"})
        self.assertEqual(delete.status_code, 200)
        self.assertTrue(delete.get_json()["ok"])

    def test_dnck_admin_only_save_payload_and_template(self):
        self.login(role="user", user_id=2)
        self.assertEqual(self.client.get("/dnck").status_code, 403)
        self.assertEqual(self.client.post("/api/dnck", json={}).status_code, 403)

        self.login(role="admin")
        save = self.client.post(
            "/api/dnck",
            json={
                "object_type": "vendor",
                "object_code": "104850022",
                "object_name": "HEPCO",
                "identity_value": "3300123456",
                "account_number": "123 456 789",
                "account_name": "HEPCO",
                "bank": "OCB",
                "purpose": "Thanh toán cho nhà cung cấp",
                "approval_level": "Cấp chi nhánh",
                "expense_type": "CSKH",
                "cost_group": "Khác",
                "request_content": "1305_THANH TOÁN CHI PHÍ KHÁC CH 27 HÀ NỘI - HUẾ",
                "sap_document": "2500000009",
                "amount": 250000,
                "approver_option": "hoang",
                "hashtags": "#cskh #tiepkhach #cskh",
                "cost_limit_ref": "QT82-DM-001",
                "reference_links": "https://www.meinvoice.vn/tra-cuu | Mã: TEST",
                "reference_note": "Ghi chú định mức chi phí",
                "detail": [
                    {
                        "label": "Bảng kê",
                        "amount": 300000,
                        "document": "4403902171",
                        "identity": "3300123456",
                        "note": "Dòng tăng",
                    },
                    {
                        "label": "Hóa đơn",
                        "amount": -50000,
                        "document": "9014589540",
                        "identity": "3300123456",
                        "note": "Dòng giảm",
                    },
                ],
            },
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(save.status_code, 200)
        data = save.get_json()
        dnck_id = data["id"]
        payload = data["qt82"]
        self.assertEqual(payload["source"], "dnck")
        self.assertEqual(payload["purpose"], "Thanh toán cho nhà cung cấp")
        self.assertTrue(payload["skipStoreManager"])
        self.assertEqual(payload["costGroup"], "Khác")
        self.assertEqual(payload["sapDocument"], "2500000009")
        self.assertEqual(payload["paymentAmount"], 250000)
        self.assertEqual(payload["customerId"], "3300123456")
        self.assertEqual([item["query"] for item in payload["approvers"]], ["hoang.vp", "my.hth"])
        self.assertTrue(payload["ready"])

        html = self.client.get(f"/dnck/{dnck_id}").get_data(as_text=True)
        self.assertIn("qt82DraftPayload", html)
        self.assertNotIn('name="payment_tag"', html)
        self.assertIn("saveDnckRecord", html)
        self.assertIn("forceCreate: true", html)
        self.assertIn("freshDraft", html)
        self.assertIn('placeholder="Diễn giải nhỏ"', html)
        self.assertNotIn("detail-label'><select", html)
        self.assertIn("handleDetailPaste", html)
        self.assertIn("handleDetailKeydown", html)
        self.assertIn("detail-remove", html)
        self.assertIn('id="amount" type="hidden"', html)
        self.assertIn("#cskh #tiepkhach", html)
        self.assertIn("QT82-DM-001", html)

        response = self.client.get(f"/api/dnck/template-tt/{dnck_id}")
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.data.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(response.data), data_only=False)
        sheet = workbook["Sheet1"]
        self.assertEqual(sheet.cell(row=5, column=2).value, "Bảng kê")
        self.assertEqual(sheet.cell(row=5, column=3).value, 300000)
        self.assertEqual(sheet.cell(row=5, column=4).value, "4403902171")
        self.assertEqual(sheet.cell(row=5, column=5).value, "3300123456")
        self.assertEqual(sheet.cell(row=5, column=6).value, "Dòng tăng")
        self.assertEqual(sheet.cell(row=6, column=2).value, "Hóa đơn")
        self.assertEqual(sheet.cell(row=6, column=3).value, -50000)
        self.assertEqual(sheet.cell(row=6, column=4).value, "9014589540")

        history = self.client.get("/api/history").get_json()["data"]
        dnck_history = next(item for item in history if item["source"] == "dnck" and item["id"] == dnck_id)
        self.assertEqual(dnck_history["ma_kh"], "104850022")
        self.assertEqual(dnck_history["ten_kh"], "HEPCO")
        self.assertEqual(dnck_history["so_bk"], "DNCK")
        self.assertEqual(dnck_history["hashtags"], ["cskh", "tiepkhach"])
        self.assertEqual(dnck_history["reference_links"], ["https://www.meinvoice.vn/tra-cuu | Mã: TEST"])
        history_html = self.client.get("/history").get_data(as_text=True)
        self.assertIn("historyTagFilters", history_html)
        self.assertIn("data-history-tag", history_html)
        self.assertIn("b[1] - a[1]", history_html)
        self.assertIn("deleteDnck", history_html)
        self.assertIn("b[1] - a[1]", html)

        eoffice_html = self.client.get(f"/eoffice/dnck/{dnck_id}").get_data(as_text=True)
        self.assertIn("qt82DraftPayload", eoffice_html)
        self.assertIn("api/dnck/template-tt", eoffice_html)
        self.assertIn("CK Bảng Kê", eoffice_html)
        self.assertIn("CK Khác", eoffice_html)
        self.assertNotIn("ĐNCK thanh toán khác", eoffice_html)
        self.assertIn("#cskh", eoffice_html)
        self.assertIn("#tiepkhach", eoffice_html)
        self.assertIn("Cấp duyệt", eoffice_html)
        self.assertIn("hoang.vp@pnj.com.vn", eoffice_html)
        self.assertIn("my.hth@pnj.com.vn", eoffice_html)
        self.assertNotIn("Loại chi phí", eoffice_html)
        self.assertIn("Nhóm chi phí", eoffice_html)
        self.assertIn("Số QT định mức / tham chiếu", eoffice_html)
        self.assertIn("Mã đối tượng", eoffice_html)
        self.assertIn("CCCD/MST", eoffice_html)
        self.assertIn('id="btnPrepareQt82"', eoffice_html)

    def test_dnck_update_copy_and_delete(self):
        self.login(role="admin")
        save = self.client.post(
            "/api/dnck",
            json={
                "object_code": "E0124764",
                "object_name": "CHÂU ĐĂNG KHOA",
                "identity_value": "046093004708",
                "account_number": "109876756206",
                "account_name": "CHÂU ĐĂNG KHOA",
                "bank": "Vietinbank",
                "purpose": "Thanh toán cho nhân viên",
                "approval_level": "Cấp chi nhánh",
                "expense_type": "Công tác phí",
                "cost_group": "Công tác",
                "request_content": "1305_THANH TOÁN CÔNG TÁC PHÍ",
                "sap_document": "2500000010",
                "detail": [{"label": "Công tác phí", "amount": 100000, "document": "HD01"}],
            },
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(save.status_code, 200)
        dnck_id = save.get_json()["id"]
        update = self.client.put(
            f"/api/dnck/{dnck_id}",
            json={
                "object_code": "E0124764",
                "object_name": "CHÂU ĐĂNG KHOA",
                "identity_value": "046093004708",
                "account_number": "109876756206",
                "account_name": "CHÂU ĐĂNG KHOA",
                "bank": "Vietinbank",
                "purpose": "Thanh toán cho nhân viên",
                "approval_level": "Cấp công ty",
                "expense_type": "CSKH",
                "cost_group": "Khác",
                "request_content": "1305_THANH TOÁN CSKH",
                "sap_document": "2500000011",
                "hashtags": ["cskh"],
                "detail": [{"label": "CSKH", "amount": 200000, "document": "HD02"}],
            },
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(update.status_code, 200)
        payload = update.get_json()["qt82"]
        self.assertEqual(payload["qt82Mode"], "thanh_toan_khac")
        self.assertEqual(payload["approvalLevel"], "Cấp công ty")
        self.assertIn("my.hth@pnj.com.vn", payload["approvalEmails"])
        self.assertEqual(payload["expenseType"], "CSKH")
        self.assertEqual(payload["paymentAmount"], 200000)

        copied = self.client.post(f"/api/dnck/{dnck_id}/copy", headers={"Origin": "http://localhost"})
        self.assertEqual(copied.status_code, 200)
        copy_id = copied.get_json()["id"]
        self.assertNotEqual(copy_id, dnck_id)

        delete = self.client.delete(f"/api/dnck/{dnck_id}", headers={"Origin": "http://localhost"})
        self.assertEqual(delete.status_code, 200)
        self.assertTrue(delete.get_json()["ok"])
        self.assertEqual(self.client.get(f"/eoffice/dnck/{dnck_id}").status_code, 404)

    def test_dnck_employee_code_infers_employee_and_omits_store_manager(self):
        self.login(role="admin")
        save = self.client.post(
            "/api/dnck",
            json={
                "object_code": "E0100001",
                "object_name": "NHAN VIEN TEST",
                "identity_value": "001200000000",
                "account_number": "A123",
                "account_name": "NHAN VIEN TEST",
                "bank": "OCB",
                "request_content": "1305_THANH TOÁN CHI PHÍ CÔNG TÁC CH 27 HÀ NỘI - HUẾ",
                "amount": 100000,
                "cost_group": "Công tác",
            },
            headers={"Origin": "http://localhost"},
        )
        self.assertEqual(save.status_code, 200)
        payload = save.get_json()["qt82"]
        self.assertEqual(payload["purpose"], "Thanh toán cho nhân viên")
        self.assertTrue(payload["skipStoreManager"])
        self.assertEqual(payload["costGroup"], "Công tác")

    def test_clean_html_print_token_hides_row_id_from_print_url(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(doc_num="2500000001")
        api_response = self.client.get("/api/history").get_json()
        phieu = next(item for item in api_response["data"] if item["id"] == phieu_id)
        token = phieu["pdf_token"]

        index_html = self.client.get("/").get_data(as_text=True)
        history_html = self.client.get("/history").get_data(as_text=True)
        self.assertIn("appUrl('/p/' + encodeURIComponent(data.pdf_token) + '?print=1')", index_html)
        self.assertIn("'p/' + encodeURIComponent(p.pdf_token)", history_html)

        self.login(role="user", user_id=2)
        response = self.client.get(f"/p/{token}?print=1")
        self.assertEqual(response.status_code, 200)
        html = response.get_data(as_text=True)
        self.assertIn("Phiếu xác nhận thông tin thanh toán chuyển khoản", html)
        self.assertIn(f'href="/api/pdf/{phieu_id}?token=', html)
        self.assertNotIn(f"/api/print/{phieu_id}", f"/p/{token}?print=1")

    def test_print_uses_five_stage_payment_schedule(self):
        self.login(role="admin")
        phieu_id = self.create_phieu(amount=143271123)

        html = self.client.get(f"/api/print/{phieu_id}").get_data(as_text=True)

        self.assertIn("1. T/T+1: 10% - tương ứng số tiền", html)
        self.assertIn("<strong>14,327,112 đồng</strong>", html)
        self.assertIn("2. T+30: 20% - tương ứng số tiền", html)
        self.assertIn("<strong>28,654,225 đồng</strong>", html)
        self.assertIn("3. T+60: 25% - tương ứng số tiền", html)
        self.assertIn("<strong>35,817,781 đồng</strong>", html)
        self.assertIn("4. T+90: 25% - tương ứng số tiền", html)
        self.assertIn("5. T+120: 20% - tương ứng số tiền", html)
        self.assertNotIn("Thanh toán trong ngày (T)", html)
        self.assertNotIn("Thanh toán vào ngày kế tiếp (T+1)", html)

    def test_payment_schedule_last_stage_absorbs_rounding_delta(self):
        schedule = app_module.build_payment_schedule(101)
        self.assertEqual([item["amount"] for item in schedule], [10, 20, 25, 25, 21])
        self.assertEqual(sum(item["amount"] for item in schedule), 101)


if __name__ == "__main__":
    unittest.main()
